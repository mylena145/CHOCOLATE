import customtkinter as ctk
from PIL import Image, ImageTk
import os
import tkinter as tk
from stock_management_page import SidebarFrame
from tkinter import filedialog, messagebox
from datetime import datetime, date
import database
import psycopg2
from database import PG_CONN

# Import des fonctions de base de données
try:
    from database import get_all_expeditions, add_expedition, update_expedition, delete_expedition, get_expedition_stats, search_expedition
    DB_AVAILABLE = True
except ImportError:
    print("⚠️ Base de données non disponible, utilisation des données de démonstration")
    DB_AVAILABLE = False

# Import optionnel pour éviter les erreurs
try:
    from responsive_utils import ThemeToggleButton
except ImportError:
    # Fallback si responsive_utils n'est pas disponible
    class ThemeToggleButton(ctk.CTkButton):
        def __init__(self, parent, app_instance, **kwargs):
            super().__init__(parent, text="☀️ Thème", width=100, height=32, command=lambda: print("Thème changé"), **kwargs)

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")


class GestionExpeditionsApp(ctk.CTkFrame):
    def __init__(self, master=None, user_info=None):
        try:
            super().__init__(master)
            self.master = master
            self.user_info = user_info
            self.previous_page = getattr(master, 'last_page', None) if hasattr(master, 'last_page') else None
            self.configure(fg_color="white")
            
            # Données d'expéditions (sera rempli depuis la BD ou données de démonstration)
            self.expeditions_data = []
            self.expedition_stats = {}
            
            # Charger les données depuis la base de données
            self.load_expeditions_data()
            
            self.planning_frame = None  # Pour garder une référence au bloc planning
            self._build_interface()
            self._refresh_planning_periodically()  # Démarre le rafraîchissement auto
            self._auto_refresh_interval = 5000  # 5 secondes
            self._auto_refresh()
        except Exception as e:
            self._show_error(str(e))
    
    def load_expeditions_data(self):
        """Charge les données d'expéditions depuis la base de données"""
        try:
            if DB_AVAILABLE:
                # Récupérer les vraies données depuis la BD
                self.expeditions_data = get_all_expeditions()
                self.expedition_stats = get_expedition_stats()
                print(f"✅ {len(self.expeditions_data)} expéditions chargées depuis la base de données")
            else:
                # Données de démonstration si la BD n'est pas disponible
                self.expeditions_data = [
                    {
                        'id': 1,
                        'number': 'BE-2024-087',
                        'client': 'Entreprise ABC',
                        'shippingDate': '2024-03-15',
                        'deliveryDate': '2024-03-16',
                        'actualDeliveryDate': None,
                        'carrier': 'DHL Express',
                        'priority': 'haute',
                        'packages': 3,
                        'totalWeight': 12.5,
                        'status': 'shipped',
                        'trackingNumber': '1234567890'
                    },
                    {
                        'id': 2,
                        'number': 'BE-2024-086',
                        'client': 'Société XYZ',
                        'shippingDate': '2024-03-15',
                        'deliveryDate': '2024-03-17',
                        'actualDeliveryDate': None,
                        'carrier': 'Chronopost',
                        'priority': 'moyenne',
                        'packages': 5,
                        'totalWeight': 25.8,
                        'status': 'preparing',
                        'trackingNumber': None
                    }
                ]
                self.expedition_stats = {
                    "total": len(self.expeditions_data),
                    "en_preparation": 1,
                    "aujourd_hui": 1,
                    "livrees": 0,
                    "transporteurs": {"DHL Express": 1, "Chronopost": 1}
                }
                print("⚠️ Utilisation des données de démonstration")
            
            # Mettre à jour l'affichage si l'interface est déjà construite
            if hasattr(self, 'main_content'):
                self.update_expeditions_display()
            
        except Exception as e:
            print(f"❌ Erreur lors du chargement des données: {e}")
            # Fallback avec données minimales
            self.expeditions_data = []
            self.expedition_stats = {"total": 0, "en_preparation": 0, "aujourd_hui": 0, "livrees": 0, "transporteurs": {}}
    
    def update_expeditions_display(self):
        """Met à jour l'affichage des expéditions en temps réel"""
        try:
            print("🔄 Mise à jour de l'affichage des expéditions...")
            
            # Mettre à jour les statistiques
            if hasattr(self, 'stats_frame'):
                self.update_stats_display()
            
            # Mettre à jour les blocs centraux (vue d'ensemble) AVANT le tableau
            if hasattr(self, 'central_blocks_frame'):
                self.update_central_blocks()
            
            # Mettre à jour le tableau APRÈS la vue d'ensemble
            if hasattr(self, 'table_frame'):
                self.update_table_display()
                
            print("✅ Affichage des expéditions mis à jour avec succès")
                
        except Exception as e:
            print(f"❌ Erreur lors de la mise à jour de l'affichage: {e}")
    
    def update_central_blocks(self):
        """Met à jour les blocs centraux (vue d'ensemble) en respectant l'ordre"""
        try:
            # Vérifier si l'attribut existe avant de l'utiliser
            if hasattr(self, 'central_blocks_frame') and self.central_blocks_frame:
                # Sauvegarder la position actuelle
                current_position = self.central_blocks_frame.pack_info()
                
                # Détruire l'ancienne frame
                self.central_blocks_frame.destroy()
                
                # Recréer la vue d'ensemble à la même position
                self.create_central_blocks()
                
                # S'assurer qu'elle reste avant le titre et le tableau
                if hasattr(self, 'table_title'):
                    self.central_blocks_frame.pack_configure(before=self.table_title)
                elif hasattr(self, 'table_frame'):
                    self.central_blocks_frame.pack_configure(before=self.table_frame)
            
        except Exception as e:
            print(f"❌ Erreur lors de la mise à jour des blocs centraux: {e}")
    
    def update_stats_display(self):
        """Met à jour l'affichage des statistiques"""
        try:
            # Mettre à jour les tuiles de statistiques
            if hasattr(self, 'total_expeditions_label'):
                self.total_expeditions_label.configure(text=str(self.expedition_stats.get("total", 0)))
            
            if hasattr(self, 'en_preparation_label'):
                self.en_preparation_label.configure(text=str(self.expedition_stats.get("en_preparation", 0)))
            
            if hasattr(self, 'aujourd_hui_label'):
                self.aujourd_hui_label.configure(text=str(self.expedition_stats.get("aujourd_hui", 0)))
            
            if hasattr(self, 'livrees_label'):
                self.livrees_label.configure(text=str(self.expedition_stats.get("livrees", 0)))
                
        except Exception as e:
            print(f"❌ Erreur lors de la mise à jour des stats: {e}")
    
    def update_table_display(self):
        """Met à jour l'affichage du tableau APRÈS la vue d'ensemble"""
        try:
            # Vider le tableau existant mais préserver le titre
            for widget in self.table_frame.winfo_children():
                widget.destroy()
            
            # Recréer le tableau avec les nouvelles données
            self.create_expeditions_table()
            
            # S'assurer que le tableau reste après le titre et la vue d'ensemble
            if hasattr(self, 'table_title') and hasattr(self, 'central_blocks_frame'):
                self.table_title.pack_configure(after=self.central_blocks_frame)
                self.table_frame.pack_configure(after=self.table_title)
            
        except Exception as e:
            print(f"❌ Erreur lors de la mise à jour du tableau: {e}")
    
    def refresh_data(self):
        """Rafraîchit les données depuis la base en temps réel"""
        try:
            print("🔄 Rafraîchissement des données depuis la base...")
            
            # Recharger les données depuis la BD
        self.load_expeditions_data()
            
            # Mettre à jour complètement l'affichage
            self.update_expeditions_display()
            
            # Afficher une notification de succès
        if hasattr(self, 'master') and hasattr(self.master, 'show_notification'):
                self.master.show_notification("✅ Données rafraîchies avec succès")
            else:
                print("✅ Données rafraîchies avec succès")
                
        except Exception as e:
            print(f"❌ Erreur lors du rafraîchissement: {e}")
            if hasattr(self, 'master') and hasattr(self.master, 'show_notification'):
                self.master.show_notification("❌ Erreur lors du rafraîchissement")
    
    def _get_urgent_expeditions(self):
        """Récupère les expéditions urgentes (priorité haute ou livraison aujourd'hui)"""
        try:
            urgent_expeditions = []
            today = date.today()
            
            for exp in self.expeditions_data:
                # Expéditions avec priorité haute
                if exp.get('priority') == 'haute':
                    urgent_expeditions.append(exp)
                # Expéditions à livrer aujourd'hui
                elif exp.get('deliveryDate') == today.strftime('%Y-%m-%d'):
                    urgent_expeditions.append(exp)
                # Limiter à 3 expéditions urgentes
                if len(urgent_expeditions) >= 3:
                    break
            
            return urgent_expeditions
        except Exception as e:
            print(f"❌ Erreur lors de la récupération des expéditions urgentes: {e}")
            return []
    
    def _get_today_planning(self):
        """Génère le planning du jour basé sur les expéditions réelles en temps réel"""
        try:
            today = date.today()
            today_str = today.strftime('%Y-%m-%d')
            
            # Récupérer les expéditions d'aujourd'hui depuis la BD en temps réel
            if DB_AVAILABLE:
                try:
                    # Requête directe à la BD pour les expéditions d'aujourd'hui
                    conn = psycopg2.connect(**PG_CONN)
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT 
                            be.id_bon_expedition,
                            be.client,
                            be.reference_commande,
                            be.date_livraison,
                            be.transporteurs,
                            be.observation,
                            be.liste_articles_livres
                        FROM sge_cre.bon_expeditions be
                        WHERE be.date_livraison = %s
                        ORDER BY be.transporteurs, be.id_bon_expedition
                    """, (today_str,))
                    
                    today_expeditions_db = cursor.fetchall()
                    conn.close()
                    
                    # Convertir au format attendu
                    today_expeditions = []
                    for exp in today_expeditions_db:
                        expedition_number = f"BE-2024-{exp[0]:03d}"
                        today_expeditions.append({
                            'id': exp[0],
                            'number': expedition_number,
                            'client': exp[1] if exp[1] else "Client non spécifié",
                            'carrier': exp[4] if exp[4] else "Non spécifié",
                            'observation': exp[5] if exp[5] else "",
                            'articles': exp[6] if exp[6] else ""
                        })
                    
                    print(f"📅 Planning temps réel : {len(today_expeditions)} expéditions pour aujourd'hui")
                    
                except Exception as e:
                    print(f"❌ Erreur requête BD planning: {e}")
                    # Fallback sur les données locales
                    today_expeditions = [exp for exp in self.expeditions_data if exp.get('deliveryDate') == today_str]
            else:
                # Mode démonstration
                today_expeditions = [exp for exp in self.expeditions_data if exp.get('deliveryDate') == today_str]
            
            if not today_expeditions:
                return []
            
            planning = []
            
            # Grouper par transporteur
            carriers = {}
            for exp in today_expeditions:
                carrier = exp.get('carrier', 'Non spécifié')
                if carrier not in carriers:
                    carriers[carrier] = []
                carriers[carrier].append(exp)
            
            # Créer le planning dynamique
            time_slots = [
                ("08:00 - 10:00", "Préparation commandes prioritaires"),
                ("10:00 - 12:00", "Enlèvement transporteurs"),
                ("14:00 - 16:00", "Préparation J+1"),
                ("16:00 - 18:00", "Clôture et vérifications")
            ]
            
            # Créer un planning basé sur les transporteurs réels
            carrier_list = list(carriers.keys())
            
            for i, (time, desc) in enumerate(time_slots):
                if i < len(carrier_list):
                    carrier_name = carrier_list[i]
                    exp_count = len(carriers[carrier_name])
                    
                    # Déterminer le statut selon l'heure actuelle
                    current_hour = datetime.now().hour
                    if time == "08:00 - 10:00" and current_hour < 10:
                        status = "En cours"
                        color = "#3867d6"
                    elif time == "10:00 - 12:00" and 10 <= current_hour < 12:
                        status = "En cours"
                        color = "#3867d6"
                    elif time == "14:00 - 16:00" and 14 <= current_hour < 16:
                        status = "En cours"
                        color = "#3867d6"
                    elif time == "16:00 - 18:00" and 16 <= current_hour < 18:
                        status = "En cours"
                        color = "#3867d6"
                    elif current_hour > 18:
                        status = "Terminé"
                        color = "#20bf6b"
                    else:
                        status = "Planifié"
                        color = "#f7b731"
                    
                    # Ajouter des détails sur les expéditions
                    exp_details = []
                    for exp in carriers[carrier_name][:3]:  # Limiter à 3 expéditions par créneau
                        exp_details.append(f"{exp['number']} ({exp['client']})")
                    
                    details_text = f"{desc} - {carrier_name}"
                    if exp_details:
                        details_text += f" - {', '.join(exp_details)}"
                    
                    planning.append({
                        "heure": time,
                        "desc": details_text,
                        "statut": f"{status} ({exp_count} colis)",
                        "color": color
                    })
                elif i == 0:  # Premier créneau même sans transporteur
                    planning.append({
                        "heure": time,
                        "desc": desc,
                        "statut": "Aucune expédition",
                        "color": "#6b7280"
                    })
            
            return planning
        except Exception as e:
            print(f"❌ Erreur lors de la génération du planning temps réel: {e}")
            return []
    
    def _get_carrier_stats(self):
        """Récupère les statistiques des transporteurs"""
        try:
            if not hasattr(self, 'expedition_stats') or not self.expedition_stats.get('transporteurs'):
                return []
            
            carriers = self.expedition_stats['transporteurs']
            carrier_list = []
            
            # Définir les couleurs pour les transporteurs
            colors = ["#20bf6b", "#3867d6", "#f7b731", "#e74c3c", "#9b59b6"]
            
            for i, (carrier_name, count) in enumerate(carriers.items()):
                if i >= len(colors):
                    break
                
                # Déterminer le statut basé sur le nombre de colis
                if count > 5:
                    status = "Actif"
                    color = colors[i]
                elif count > 0:
                    status = "En attente"
                    color = "#f7b731"
                else:
                    status = "Inactif"
                    color = "#6b7280"
                
                carrier_list.append({
                    "nom": carrier_name,
                    "statut": status,
                    "colis": str(count),
                    "color": color
                })
            
            return carrier_list
        except Exception as e:
            print(f"❌ Erreur lors de la récupération des stats transporteurs: {e}")
            return []

    def _build_interface(self):
        """Construction de l'interface"""
        # Topbar - TOUJOURS VISIBLE
        self._build_topbar()
        
        # Contenu principal
        self._build_main_content()

    def _build_topbar(self):
        """Construction de la barre supérieure - TOUJOURS VISIBLE"""
        # Topbar fixe en haut
        self.topbar = ctk.CTkFrame(self, height=80, fg_color='#F7F9FC', border_width=1, border_color='#E0E0E0')
        self.topbar.pack(side='top', fill='x', pady=(0, 0))
        
        # Logo et titre
        try:
            logo_img = Image.open('logo.png')
            logo = ctk.CTkImage(light_image=logo_img, size=(60, 60))
            ctk.CTkLabel(self.topbar, image=logo, text='').pack(side='left', padx=25, pady=10)
        except:
            ctk.CTkLabel(self.topbar, text='SGE', font=ctk.CTkFont(size=24, weight='bold'), text_color='#333333').pack(side='left', padx=25)
        
        # Titre principal
        title_frame = ctk.CTkFrame(self.topbar, fg_color='transparent')
        title_frame.pack(side='left', padx=(20, 0))
        
        ctk.CTkLabel(title_frame, text='Gestion des Expéditions', font=ctk.CTkFont(size=24, weight='bold'), text_color='#333333').pack(anchor='w')
        ctk.CTkLabel(title_frame, text="Gérez les bons d'expédition et les livraisons", font=ctk.CTkFont(size=16), text_color='#666666').pack(anchor='w')
        
        # Bouton de rafraîchissement
        refresh_button = ctk.CTkButton(
            self.topbar, 
            text="🔄 Rafraîchir", 
            width=120, 
            height=32, 
            fg_color="#3b82f6", 
            hover_color="#2563eb",
            text_color="white",
            font=ctk.CTkFont(size=13, weight="bold"),
            corner_radius=8,
            command=self.refresh_data
        )
        refresh_button.pack(side='right', padx=(10, 10), pady=10)
        
        # Bouton de thème
        try:
            self.theme_button = ThemeToggleButton(self.topbar, self.master)
            self.theme_button.pack(side='right', padx=(0, 20), pady=10)
        except Exception as e:
            print(f"Erreur bouton thème: {e}")
            # Bouton de thème simplifié
            theme_btn = ctk.CTkButton(self.topbar, text="☀️ Thème", width=100, height=32, command=self._toggle_theme)
            theme_btn.pack(side='right', padx=(0, 20), pady=10)
        
        # Boutons d'action
        buttons_frame = ctk.CTkFrame(self.topbar, fg_color='transparent')
        buttons_frame.pack(side='right', padx=25, pady=10)
        
        btns = [
            ("+ Nouvelle Expédition", "#2563eb", self.open_new_expedition_modal),
            ("Planning", "#a21caf", self.action_planning),
            ("Suivi", "#059669", self.open_tracking_modal),
            ("⭳ Exporter", "#22c55e", self.action_export),
        ]
        
        for txt, color, cmd in btns:
            btn = ctk.CTkButton(buttons_frame, text=txt, width=120, height=58, fg_color=color, hover_color=color, text_color="white", font=ctk.CTkFont(size=16, weight="bold"), corner_radius=12, command=cmd)
            btn.pack(side='left', padx=5)

        # Ajout du bouton retour intelligent en haut à gauche
        topbar = ctk.CTkFrame(self, fg_color="transparent")
        topbar.pack(fill="x", pady=(10, 0))
        btn_retour = ctk.CTkButton(
            topbar,
            text="← Retour",
            width=110,
            height=36,
            fg_color="#2563eb",
            hover_color="#1d4ed8",
            text_color="white",
            font=ctk.CTkFont(size=15, weight="bold"),
            corner_radius=8,
            command=self._go_back
        )
        btn_retour.pack(side="left", padx=18, pady=5)

    def _build_main_content(self):
        """Construction du contenu principal"""
        # Frame principal avec scroll
        self.main_content = ctk.CTkScrollableFrame(self, fg_color='white')
        self.main_content.pack(side='top', fill='both', expand=True, padx=20, pady=20)

        # Tuiles statistiques
        self.create_stat_tiles()

        # Blocs centraux
        self.create_central_blocks()

        # Tableau des expéditions
        self.create_expedition_table()

    def _toggle_theme(self):
        """Basculement simple du thème"""
        current = ctk.get_appearance_mode()
        new_theme = "dark" if current == "light" else "light"
        ctk.set_appearance_mode(new_theme)
        print(f"Thème changé vers: {new_theme}")

    def create_stat_tiles(self):
        """Création des tuiles statistiques"""
        # Titre de section
        title_label = ctk.CTkLabel(self.main_content, text="Statistiques", font=ctk.CTkFont(size=20, weight="bold"), text_color="#1f2937")
        title_label.pack(anchor="w", pady=(0, 15))
        
        # Frame pour les tuiles
        stats_frame = ctk.CTkFrame(self.main_content, fg_color="transparent")
        stats_frame.pack(fill="x", pady=(0, 20))
        
        # Utiliser les vraies statistiques si disponibles
        if hasattr(self, 'expedition_stats') and self.expedition_stats:
            stats = [
                {"label": "En Préparation", "value": str(self.expedition_stats.get("en_preparation", 0)), "desc": "À préparer", "color": "#f7b731", "icon": "📦"},
                {"label": "Expédiées Aujourd'hui", "value": str(self.expedition_stats.get("aujourd_hui", 0)), "desc": "En transit", "color": "#3867d6", "icon": "🚚"},
                {"label": "Livrées", "value": str(self.expedition_stats.get("livrees", 0)), "desc": "Ce mois", "color": "#20bf6b", "icon": "✅"},
                {"label": "Total", "value": str(self.expedition_stats.get("total", 0)), "desc": "Toutes les expéditions", "color": "#eb3b5a", "icon": "📊"},
            ]
        else:
            # Statistiques par défaut
            stats = [
                {"label": "En Préparation", "value": "0", "desc": "À préparer", "color": "#f7b731", "icon": "📦"},
                {"label": "Expédiées Aujourd'hui", "value": "0", "desc": "En transit", "color": "#3867d6", "icon": "🚚"},
                {"label": "Livrées", "value": "0", "desc": "Ce mois", "color": "#20bf6b", "icon": "✅"},
                {"label": "Total", "value": "0", "desc": "Toutes les expéditions", "color": "#eb3b5a", "icon": "📊"},
            ]
        
        # Créer une grille responsive
        for i, stat in enumerate(stats):
            # Frame de la tuile
            frame = ctk.CTkFrame(stats_frame, fg_color="white", border_width=1, border_color="#eee", corner_radius=8)
            frame.pack(side="left", expand=True, fill="x", padx=10, pady=5)
            
            # Contenu de la tuile
            content_frame = ctk.CTkFrame(frame, fg_color="transparent")
            content_frame.pack(padx=15, pady=15, fill="x")
            
            # Icône
            icon_frame = ctk.CTkFrame(content_frame, fg_color=stat["color"], width=40, height=40, corner_radius=8)
            icon_frame.pack(side="left", padx=(0, 15))
            icon_frame.pack_propagate(False)
            
            ctk.CTkLabel(icon_frame, text=stat["icon"], font=ctk.CTkFont(size=20), text_color="white", fg_color="transparent").pack(expand=True)
            
            # Texte
            text_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
            text_frame.pack(side="left", fill="x", expand=True)
            
            ctk.CTkLabel(text_frame, text=stat["label"], font=ctk.CTkFont(size=13), text_color="#444").pack(anchor="w")
            ctk.CTkLabel(text_frame, text=stat["value"], font=ctk.CTkFont(size=24, weight="bold"), text_color=stat["color"]).pack(anchor="w")
            ctk.CTkLabel(text_frame, text=stat["desc"], font=ctk.CTkFont(size=12), text_color=stat["color"]).pack(anchor="w")

    def create_central_blocks(self):
        """Création des blocs centraux : urgentes, planning du jour, transporteurs (sans doublon)"""
        # Détruire l'ancienne frame si elle existe
        if hasattr(self, 'central_blocks_frame') and self.central_blocks_frame:
            self.central_blocks_frame.destroy()
        
        # Nouvelle frame pour la vue d'ensemble
        self.central_blocks_frame = ctk.CTkFrame(self.main_content, fg_color="transparent")
        
        # S'assurer que la vue d'ensemble se place avant le tableau
        if hasattr(self, 'table_frame'):
            self.central_blocks_frame.pack(fill="x", pady=(0, 20), before=self.table_frame)
        else:
            self.central_blocks_frame.pack(fill="x", pady=(0, 20))

        # Titre de section
        title_label = ctk.CTkLabel(self.central_blocks_frame, text="Vue d'ensemble", font=ctk.CTkFont(size=20, weight="bold"), text_color="#1f2937")
        title_label.pack(anchor="w", pady=(20, 15))
        
        # Frame pour les blocs
        blocks_frame = ctk.CTkFrame(self.central_blocks_frame, fg_color="transparent")
        blocks_frame.pack(fill="x", pady=(0, 20))
        
        # Expéditions Urgentes
        urgentes_frame = ctk.CTkFrame(blocks_frame, fg_color="white", border_width=1, border_color="#eee", corner_radius=8)
        urgentes_frame.pack(side="left", expand=True, fill="both", padx=10, pady=5)
        ctk.CTkLabel(urgentes_frame, text="🚨 Expéditions Urgentes", font=ctk.CTkFont(size=18, weight="bold"), text_color="#b91c1c").pack(anchor="w", padx=15, pady=15)
        urgent_expeditions = self._get_urgent_expeditions()
        if urgent_expeditions:
            for exp in urgent_expeditions:
                card = ctk.CTkFrame(urgentes_frame, fg_color="#fef2f2", border_width=2, border_color="#fecaca", corner_radius=8)
                card.pack(fill="x", padx=15, pady=5)
                ctk.CTkLabel(card, text=exp["number"], font=ctk.CTkFont(size=16, weight="bold"), text_color="#b91c1c", fg_color="transparent").pack(anchor="w", padx=10, pady=(5,0))
                ctk.CTkLabel(card, text=f"{exp['client']} - {exp['carrier']}", font=ctk.CTkFont(size=15), text_color="#b91c1c", fg_color="transparent").pack(anchor="w", padx=10, pady=(2,0))
                ctk.CTkLabel(card, text=exp["priority"].upper(), font=ctk.CTkFont(size=14, weight="bold"), text_color="#b91c1c", fg_color="transparent").pack(anchor="e", padx=10, pady=(0,5))
        else:
            no_urgent_frame = ctk.CTkFrame(urgentes_frame, fg_color="#f9fafb", corner_radius=8)
            no_urgent_frame.pack(fill="x", padx=15, pady=5)
            ctk.CTkLabel(no_urgent_frame, text="✅ Aucune expédition urgente", font=ctk.CTkFont(size=14), text_color="#059669", fg_color="transparent").pack(pady=10)

        # Planning du Jour
        if hasattr(self, 'planning_frame') and self.planning_frame:
            self.planning_frame.destroy()
        self.planning_frame = ctk.CTkFrame(blocks_frame, fg_color="white", border_width=1, border_color="#eee", corner_radius=8)
        self.planning_frame.pack(side="left", expand=True, fill="both", padx=10, pady=5)
        ctk.CTkLabel(self.planning_frame, text="📅 Planning du Jour", font=ctk.CTkFont(size=18, weight="bold"), text_color="#3867d6").pack(anchor="w", padx=15, pady=15)
        today_planning = self._get_today_planning()
        if today_planning:
            for planning in today_planning:
                item_frame = ctk.CTkFrame(self.planning_frame, fg_color="transparent")
                item_frame.pack(fill="x", padx=15, pady=5)
                ctk.CTkLabel(item_frame, text=planning["heure"], font=ctk.CTkFont(size=16, weight="bold"), text_color="#222").pack(anchor="w")
                ctk.CTkLabel(item_frame, text=planning["desc"], font=ctk.CTkFont(size=15), text_color="#555").pack(anchor="w")
                ctk.CTkLabel(item_frame, text=planning["statut"], font=ctk.CTkFont(size=14), text_color=planning["color"]).pack(anchor="e")
        else:
            no_planning_frame = ctk.CTkFrame(self.planning_frame, fg_color="#f9fafb", corner_radius=8)
            no_planning_frame.pack(fill="x", padx=15, pady=5)
            ctk.CTkLabel(no_planning_frame, text="📅 Aucune expédition prévue aujourd'hui", font=ctk.CTkFont(size=14), text_color="#6b7280", fg_color="transparent").pack(pady=10)

        # Transporteurs
        transp_frame = ctk.CTkFrame(blocks_frame, fg_color="white", border_width=1, border_color="#eee", corner_radius=8)
        transp_frame.pack(side="left", expand=True, fill="both", padx=10, pady=5)
        ctk.CTkLabel(transp_frame, text="🚛 Transporteurs", font=ctk.CTkFont(size=18, weight="bold"), text_color="#20bf6b").pack(anchor="w", padx=15, pady=15)
        carrier_stats = self._get_carrier_stats()
        if carrier_stats:
            for carrier in carrier_stats:
                item_frame = ctk.CTkFrame(transp_frame, fg_color="transparent")
                item_frame.pack(fill="x", padx=15, pady=5)
                ctk.CTkLabel(item_frame, text=carrier["nom"], font=ctk.CTkFont(size=16, weight="bold"), text_color="#222").pack(anchor="w")
                ctk.CTkLabel(item_frame, text=f"Statut: {carrier['statut']}", font=ctk.CTkFont(size=15), text_color=carrier["color"]).pack(anchor="w")
                ctk.CTkLabel(item_frame, text=f"{carrier['colis']} colis", font=ctk.CTkFont(size=14), text_color="#666").pack(anchor="e")
        else:
            no_carrier_frame = ctk.CTkFrame(transp_frame, fg_color="#f9fafb", corner_radius=8)
            no_carrier_frame.pack(fill="x", padx=15, pady=5)
            ctk.CTkLabel(no_carrier_frame, text="🚛 Aucun transporteur configuré", font=ctk.CTkFont(size=14), text_color="#6b7280", fg_color="transparent").pack(pady=10)

    def create_expedition_table(self):
        """Création du tableau des expéditions APRÈS la vue d'ensemble"""
        # Titre du tableau - stocké comme attribut de classe
        self.table_title = ctk.CTkLabel(self.main_content, text="Liste des Expéditions", font=ctk.CTkFont(size=20, weight="bold"), text_color="#1f2937")
        
        # Frame du tableau
        self.table_frame = ctk.CTkFrame(self.main_content, fg_color="white", corner_radius=8, border_width=1, border_color="#e5e7eb")
        
        # S'assurer que le titre se place après la vue d'ensemble et avant le tableau
        if hasattr(self, 'central_blocks_frame'):
            self.table_title.pack(anchor="w", pady=(20, 10), after=self.central_blocks_frame)
            self.table_frame.pack(fill="x", pady=(0, 20), after=self.table_title)
        else:
            self.table_title.pack(anchor="w", pady=(20, 10))
        self.table_frame.pack(fill="x", pady=(0, 20))
        
        # Créer le contenu du tableau
        self.create_expeditions_table()
    
    def create_expeditions_table(self):
        """Crée le contenu du tableau des expéditions"""
        # Vider le tableau existant
        for widget in self.table_frame.winfo_children():
            widget.destroy()
        
        # En-têtes
        headers = ["N° Expédition", "Client", "Transporteur", "Priorité", "Statut", "Poids (kg)", "Actions"]
        header_frame = ctk.CTkFrame(self.table_frame, fg_color="#f8fafc")
        header_frame.pack(fill="x", padx=1, pady=1)
        
        for i, header in enumerate(headers):
            label = ctk.CTkLabel(header_frame, text=header, font=ctk.CTkFont(size=13, weight="bold"), text_color="#374151")
            label.grid(row=0, column=i, sticky="ew", padx=15, pady=12)
            header_frame.grid_columnconfigure(i, weight=1)
        
        # Message si aucune expédition
        if not self.expeditions_data:
            no_data_frame = ctk.CTkFrame(self.table_frame, fg_color="transparent")
            no_data_frame.pack(fill="x", padx=1, pady=20)
            
            ctk.CTkLabel(
                no_data_frame, 
                text="📦 Aucune expédition trouvée", 
                font=ctk.CTkFont(size=16, weight="bold"), 
                text_color="#6b7280"
            ).pack(pady=20)
            
            ctk.CTkLabel(
                no_data_frame, 
                text="Cliquez sur 'Nouvelle Expédition' pour commencer", 
                font=ctk.CTkFont(size=14), 
                text_color="#9ca3af"
            ).pack()
            return
        
        # Données
        for row_idx, expedition in enumerate(self.expeditions_data):
            row_frame = ctk.CTkFrame(self.table_frame, fg_color="transparent")
            row_frame.pack(fill="x", padx=1, pady=1)
            
            # Couleur de fond alternée
            if row_idx % 2 == 0:
                row_frame.configure(fg_color="#f9fafb")
            
            # Données de la ligne
            row_data = [
                expedition['number'],
                expedition['client'],
                expedition['carrier'],
                expedition['priority'].title(),
                expedition['status'].title(),
                f"{expedition['totalWeight']} kg"
            ]
            
            for col_idx, cell_data in enumerate(row_data):
                label = ctk.CTkLabel(row_frame, text=str(cell_data), font=ctk.CTkFont(size=12), text_color="#6b7280")
                label.grid(row=0, column=col_idx, sticky="ew", padx=15, pady=8)
                row_frame.grid_columnconfigure(col_idx, weight=1)
            
            # Boutons d'action
            actions_frame = ctk.CTkFrame(row_frame, fg_color="transparent")
            actions_frame.grid(row=0, column=6, sticky="ew", padx=15, pady=8)
            
            ctk.CTkButton(actions_frame, text="👁️", width=30, height=25, command=lambda e=expedition: self.view_expedition(e)).pack(side="left", padx=2)
            ctk.CTkButton(actions_frame, text="✏️", width=30, height=25, command=lambda e=expedition: self.edit_expedition(e)).pack(side="left", padx=2)
            ctk.CTkButton(actions_frame, text="🗑️", width=30, height=25, command=lambda e=expedition: self.delete_expedition(e)).pack(side="left", padx=2)

    def view_expedition(self, expedition):
        """Afficher les détails d'une expédition"""
        print(f"Affichage de l'expédition: {expedition['number']}")

    def edit_expedition(self, expedition):
        """Modifier une expédition"""
        print(f"Modification de l'expédition: {expedition['number']}")
        
        # Créer une fenêtre modale
        modal = ctk.CTkToplevel(self)
        modal.title("Modifier l'Expédition")
        modal.geometry("800x600")
        modal.configure(fg_color='white')
        modal.grab_set()
        modal.resizable(False, False)
        
        # Titre
        title_label = ctk.CTkLabel(modal, text=f"✏️ Modifier l'Expédition {expedition['number']}", font=ctk.CTkFont(size=24, weight="bold"), text_color="#1f2937")
        title_label.pack(pady=(20, 30))
        
        # Frame principal avec scroll
        main_frame = ctk.CTkScrollableFrame(modal, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # Formulaire
        form_frame = ctk.CTkFrame(main_frame, fg_color="white", border_width=1, border_color="#e5e7eb", corner_radius=8)
        form_frame.pack(fill="x", pady=(0, 20))
        
        # Titre du formulaire
        ctk.CTkLabel(form_frame, text="Informations de l'expédition", font=ctk.CTkFont(size=18, weight="bold"), text_color="#374151").pack(anchor="w", padx=20, pady=(20, 15))
        
        # Légende des champs obligatoires
        legend_frame = ctk.CTkFrame(form_frame, fg_color="#fef3c7", border_width=1, border_color="#f59e0b", corner_radius=8)
        legend_frame.pack(fill="x", padx=20, pady=(0, 15))
        ctk.CTkLabel(legend_frame, text="ℹ️ Les champs marqués d'un * sont obligatoires", font=ctk.CTkFont(size=12), text_color="#92400e").pack(padx=15, pady=8)
        
        # Grille pour les champs
        fields_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        fields_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        # Variables pour les champs (pré-remplies avec les données actuelles)
        client_var = tk.StringVar(value=expedition.get('client', ''))
        carrier_var = tk.StringVar(value=expedition.get('carrier', 'DHL Express'))
        priority_var = tk.StringVar(value=expedition.get('priority', 'moyenne'))
        weight_var = tk.StringVar(value=str(expedition.get('totalWeight', '')))
        packages_var = tk.StringVar(value=str(expedition.get('packages', '')))
        observation_var = tk.StringVar(value=expedition.get('observation', ''))
        status_var = tk.StringVar(value=expedition.get('status', 'preparing'))
        
        # Champs du formulaire avec indication des champs obligatoires
        fields = [
            ("Client *", client_var, "text", True),  # Obligatoire
            ("Transporteur", carrier_var, "combo", False, ["DHL Express", "Chronopost", "Colissimo", "UPS"]),
            ("Priorité", priority_var, "combo", False, ["haute", "moyenne", "basse"]),
            ("Statut", status_var, "combo", False, ["preparing", "in-transit", "delivered"]),
            ("Poids (kg) *", weight_var, "text", True),  # Obligatoire
            ("Nombre de colis *", packages_var, "text", True),  # Obligatoire
        ]
        
        for i, (label, var, field_type, is_required, *args) in enumerate(fields):
            row = i // 2
            col = i % 2
            
            # Frame pour le label et l'indicateur obligatoire
            label_frame = ctk.CTkFrame(fields_frame, fg_color="transparent")
            label_frame.grid(row=row*2, column=col, sticky="w", padx=(0, 10), pady=(15, 5))
            
            # Label avec couleur selon si obligatoire
            label_color = "#dc2626" if is_required else "#374151"  # Rouge si obligatoire
            label_text = ctk.CTkLabel(label_frame, text=label, font=ctk.CTkFont(size=14, weight="bold"), text_color=label_color)
            label_text.pack(side="left")
            
            # Champ avec bordure colorée si obligatoire
            if field_type == "text":
                entry = ctk.CTkEntry(fields_frame, textvariable=var, width=200, height=35, font=ctk.CTkFont(size=14))
                if is_required:
                    entry.configure(border_color="#dc2626", border_width=2)  # Bordure rouge pour obligatoire
                entry.grid(row=row*2+1, column=col, sticky="ew", padx=(0, 20), pady=(0, 15))
            elif field_type == "combo":
                combo = ctk.CTkOptionMenu(fields_frame, variable=var, values=args[0], width=200, height=35, font=ctk.CTkFont(size=14))
                combo.grid(row=row*2+1, column=col, sticky="ew", padx=(0, 20), pady=(0, 15))
            
            fields_frame.grid_columnconfigure(col, weight=1)
        
        # Champ observation (pleine largeur) - optionnel
        ctk.CTkLabel(fields_frame, text="Observation:", font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151").grid(row=len(fields)*2, column=0, columnspan=2, sticky="w", padx=(0, 10), pady=(15, 5))
        observation_entry = ctk.CTkEntry(fields_frame, textvariable=observation_var, width=400, height=35, font=ctk.CTkFont(size=14))
        observation_entry.grid(row=len(fields)*2+1, column=0, columnspan=2, sticky="ew", padx=(0, 20), pady=(0, 15))
        
        # Boutons d'action
        buttons_frame = ctk.CTkFrame(modal, fg_color="transparent")
        buttons_frame.pack(pady=(0, 20))
        
        def save_expedition_update():
            """Mettre à jour l'expédition"""
            # Vérifier les champs obligatoires avec messages spécifiques
            missing_fields = []
            if not client_var.get().strip():
                missing_fields.append("Client")
            if not weight_var.get().strip():
                missing_fields.append("Poids")
            if not packages_var.get().strip():
                missing_fields.append("Nombre de colis")
            
            if missing_fields:
                # Afficher une erreur détaillée
                error_text = f"❌ Champs obligatoires manquants : {', '.join(missing_fields)}"
                error_label = ctk.CTkLabel(modal, text=error_text, text_color="#ef4444", font=ctk.CTkFont(size=14, weight="bold"))
                error_label.pack(pady=10)
                modal.after(4000, error_label.destroy)
                return
            
            # Validation du poids
            poids_str = weight_var.get().strip()
            try:
                poids_float = float(poids_str) if poids_str else 0.0
                if poids_float < 0:
                    error_label = ctk.CTkLabel(modal, text="❌ Le poids ne peut pas être négatif", text_color="#ef4444", font=ctk.CTkFont(size=14, weight="bold"))
                    error_label.pack(pady=10)
                    modal.after(3000, error_label.destroy)
                    return
            except ValueError:
                error_label = ctk.CTkLabel(modal, text="❌ Le poids doit être un nombre valide", text_color="#ef4444", font=ctk.CTkFont(size=14, weight="bold"))
                error_label.pack(pady=10)
                modal.after(3000, error_label.destroy)
                return
            
            # Préparer les données pour la BD
            expedition_data = {
                'client': client_var.get(),
                'reference_commande': expedition.get('trackingNumber', ''),
                'date_livraison': expedition.get('deliveryDate', datetime.now().strftime('%Y-%m-%d')),
                'transporteurs': carrier_var.get(),
                'priorite': priority_var.get(),
                'poids': poids_str,
                'observation': observation_var.get(),
                'liste_articles_livres': expedition.get('articles', '')
            }
            
            print(f"🔄 Données de mise à jour: poids={poids_str} kg")
            
            try:
                if DB_AVAILABLE:
                    # Mettre à jour dans la base de données
                    success = update_expedition(expedition['id'], expedition_data)
                    if success:
                        print(f"✅ Expédition mise à jour dans la BD: {expedition['number']}")
                        # Recharger les données
                        self.load_expeditions_data()
                        # Mettre à jour l'affichage
                        self.update_expeditions_display()
                        # Mettre à jour les blocs centraux (notifications, planning, etc.)
                        self.update_central_blocks()
                        # Mettre à jour les statistiques
                        self.update_stats_display()
                        success_msg = "✅ Expédition mise à jour avec succès !"
                    else:
                        print("❌ Erreur lors de la mise à jour dans la BD")
                        success_msg = "❌ Erreur lors de la mise à jour"
                else:
                    # Mode démonstration
                    self._update_local_expedition(expedition['id'], expedition_data)
                    success_msg = "✅ Expédition mise à jour (mode démonstration)"
                
                # Afficher un message de succès
                success_label = ctk.CTkLabel(modal, text=success_msg, text_color="#10b981" if "succès" in success_msg else "#ef4444", font=ctk.CTkFont(size=16, weight="bold"))
                success_label.pack(pady=10)
                
                # Fermer la modal après 2 secondes
                modal.after(2000, modal.destroy)
                    
            except Exception as e:
                print(f"❌ Erreur lors de la mise à jour: {e}")
                error_label = ctk.CTkLabel(modal, text=f"❌ Erreur: {str(e)}", text_color="#ef4444", font=ctk.CTkFont(size=14))
                error_label.pack(pady=10)
                modal.after(3000, error_label.destroy)
        
        def cancel_update():
            """Annuler la modification"""
            modal.destroy()
        
        # Boutons
        ctk.CTkButton(buttons_frame, text="❌ Annuler", width=120, height=40, fg_color="#ef4444", hover_color="#dc2626", command=cancel_update).pack(side="left", padx=10)
        ctk.CTkButton(buttons_frame, text="✅ Mettre à jour", width=150, height=40, fg_color="#10b981", hover_color="#059669", command=save_expedition_update).pack(side="left", padx=10)

    def delete_expedition(self, expedition):
        """Supprimer une expédition"""
        print(f"Suppression de l'expédition: {expedition['number']}")
        
        # Créer une fenêtre de confirmation
        modal = ctk.CTkToplevel(self)
        modal.title("Confirmer la suppression")
        modal.geometry("500x300")
        modal.configure(fg_color='white')
        modal.grab_set()
        modal.resizable(False, False)
        
        # Titre
        title_label = ctk.CTkLabel(modal, text="🗑️ Supprimer l'Expédition", font=ctk.CTkFont(size=24, weight="bold"), text_color="#1f2937")
        title_label.pack(pady=(20, 20))
        
        # Message de confirmation
        message = f"Êtes-vous sûr de vouloir supprimer l'expédition {expedition['number']} ?\n\nCette action est irréversible."
        message_label = ctk.CTkLabel(modal, text=message, font=ctk.CTkFont(size=16), text_color="#6b7280", justify="center")
        message_label.pack(pady=(0, 30))
        
        # Détails de l'expédition
        details_frame = ctk.CTkFrame(modal, fg_color="#f3f4f6", corner_radius=8)
        details_frame.pack(fill="x", padx=20, pady=(0, 30))
        
        details_text = f"Client: {expedition.get('client', 'N/A')}\nTransporteur: {expedition.get('carrier', 'N/A')}\nStatut: {expedition.get('status', 'N/A')}"
        details_label = ctk.CTkLabel(details_frame, text=details_text, font=ctk.CTkFont(size=14), text_color="#374151", justify="left")
        details_label.pack(padx=15, pady=15)
        
        # Boutons d'action
        buttons_frame = ctk.CTkFrame(modal, fg_color="transparent")
        buttons_frame.pack(pady=(0, 20))
        
        def confirm_delete():
            """Confirmer la suppression"""
            try:
                if DB_AVAILABLE:
                    # Supprimer de la base de données
                    success = delete_expedition(expedition['id'])
                    if success:
                        print(f"✅ Expédition supprimée de la BD: {expedition['number']}")
                        # Recharger les données
                        self.load_expeditions_data()
                        # Mettre à jour l'affichage
                        self.update_expeditions_display()
                        # Mettre à jour les blocs centraux (notifications, planning, etc.)
                        self.update_central_blocks()
                        # Mettre à jour les statistiques
                        self.update_stats_display()
                        success_msg = "✅ Expédition supprimée avec succès !"
                    else:
                        print("❌ Erreur lors de la suppression dans la BD")
                        success_msg = "❌ Erreur lors de la suppression"
                else:
                    # Mode démonstration
                    self._delete_local_expedition(expedition['id'])
                    # Mettre à jour les blocs centraux (notifications, planning, etc.)
                    self.update_central_blocks()
                    # Mettre à jour les statistiques
                    self.update_stats_display()
                    success_msg = "✅ Expédition supprimée (mode démonstration)"
                
                # Fermer la modal de confirmation
                modal.destroy()
                
                # Afficher un message de succès
                success_modal = ctk.CTkToplevel(self)
                success_modal.title("Suppression terminée")
                success_modal.geometry("400x200")
                success_modal.configure(fg_color='white')
                success_modal.grab_set()
                success_modal.resizable(False, False)
                
                ctk.CTkLabel(success_modal, text=success_msg, text_color="#10b981" if "succès" in success_msg else "#ef4444", font=ctk.CTkFont(size=16, weight="bold")).pack(expand=True)
                success_modal.after(2000, success_modal.destroy)
                    
            except Exception as e:
                print(f"❌ Erreur lors de la suppression: {e}")
                modal.destroy()
                
                # Afficher un message d'erreur
                error_modal = ctk.CTkToplevel(self)
                error_modal.title("Erreur")
                error_modal.geometry("400x200")
                error_modal.configure(fg_color='white')
                error_modal.grab_set()
                error_modal.resizable(False, False)
                
                ctk.CTkLabel(error_modal, text=f"❌ Erreur: {str(e)}", text_color="#ef4444", font=ctk.CTkFont(size=16, weight="bold")).pack(expand=True)
                error_modal.after(3000, error_modal.destroy)
        
        def cancel_delete():
            """Annuler la suppression"""
            modal.destroy()
        
        # Boutons
        ctk.CTkButton(buttons_frame, text="❌ Annuler", width=120, height=40, fg_color="#6b7280", hover_color="#4b5563", command=cancel_delete).pack(side="left", padx=10)
        ctk.CTkButton(buttons_frame, text="🗑️ Supprimer", width=120, height=40, fg_color="#ef4444", hover_color="#dc2626", command=confirm_delete).pack(side="left", padx=10)

    def open_new_expedition_modal(self):
        """Ouvrir la modal de nouvelle expédition"""
        print("Ouverture de la modal nouvelle expédition")
        
        # Créer une fenêtre modale
        modal = ctk.CTkToplevel(self)
        modal.title("Nouvelle Expédition")
        modal.geometry("800x600")
        modal.configure(fg_color='white')
        modal.grab_set()  # Rendre la fenêtre modale
        modal.resizable(False, False)
        
        # Titre
        title_label = ctk.CTkLabel(modal, text="📦 Nouvelle Expédition", font=ctk.CTkFont(size=24, weight="bold"), text_color="#1f2937")
        title_label.pack(pady=(20, 30))
        
        # Frame principal avec scroll
        main_frame = ctk.CTkScrollableFrame(modal, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # Formulaire
        form_frame = ctk.CTkFrame(main_frame, fg_color="white", border_width=1, border_color="#e5e7eb", corner_radius=8)
        form_frame.pack(fill="x", pady=(0, 20))
        
        # Titre du formulaire
        ctk.CTkLabel(form_frame, text="Informations de l'expédition", font=ctk.CTkFont(size=18, weight="bold"), text_color="#374151").pack(anchor="w", padx=20, pady=(20, 15))
        
        # Légende des champs obligatoires
        legend_frame = ctk.CTkFrame(form_frame, fg_color="#fef3c7", border_width=1, border_color="#f59e0b", corner_radius=8)
        legend_frame.pack(fill="x", padx=20, pady=(0, 15))
        ctk.CTkLabel(legend_frame, text="ℹ️ Les champs marqués d'un * sont obligatoires", font=ctk.CTkFont(size=12), text_color="#92400e").pack(padx=15, pady=8)
        
        # Grille pour les champs
        fields_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        fields_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        # Variables pour les champs
        client_var = tk.StringVar(value="")
        carrier_var = tk.StringVar(value="DHL Express")
        priority_var = tk.StringVar(value="moyenne")
        weight_var = tk.StringVar(value="")
        packages_var = tk.StringVar(value="")
        observation_var = tk.StringVar(value="")
        
        # Champs du formulaire avec indication des champs obligatoires
        fields = [
            ("Client *", client_var, "text", True),  # Obligatoire
            ("Transporteur", carrier_var, "combo", False, ["DHL Express", "Chronopost", "Colissimo", "UPS"]),
            ("Priorité", priority_var, "combo", False, ["haute", "moyenne", "basse"]),
            ("Poids (kg) *", weight_var, "text", True),  # Obligatoire
            ("Nombre de colis *", packages_var, "text", True),  # Obligatoire
        ]
        
        for i, (label, var, field_type, is_required, *args) in enumerate(fields):
            row = i // 2
            col = i % 2
            
            # Frame pour le label et l'indicateur obligatoire
            label_frame = ctk.CTkFrame(fields_frame, fg_color="transparent")
            label_frame.grid(row=row*2, column=col, sticky="w", padx=(0, 10), pady=(15, 5))
            
            # Label avec couleur selon si obligatoire
            label_color = "#dc2626" if is_required else "#374151"  # Rouge si obligatoire
            label_text = ctk.CTkLabel(label_frame, text=label, font=ctk.CTkFont(size=14, weight="bold"), text_color=label_color)
            label_text.pack(side="left")
            
            # Champ avec bordure colorée si obligatoire
            if field_type == "text":
                entry = ctk.CTkEntry(fields_frame, textvariable=var, width=200, height=35, font=ctk.CTkFont(size=14))
                if is_required:
                    entry.configure(border_color="#dc2626", border_width=2)  # Bordure rouge pour obligatoire
                entry.grid(row=row*2+1, column=col, sticky="ew", padx=(0, 20), pady=(0, 15))
            elif field_type == "combo":
                combo = ctk.CTkOptionMenu(fields_frame, variable=var, values=args[0], width=200, height=35, font=ctk.CTkFont(size=14))
                combo.grid(row=row*2+1, column=col, sticky="ew", padx=(0, 20), pady=(0, 15))
            
            fields_frame.grid_columnconfigure(col, weight=1)
        
        # Champ observation (pleine largeur) - optionnel
        ctk.CTkLabel(fields_frame, text="Observation:", font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151").grid(row=len(fields)*2, column=0, columnspan=2, sticky="w", padx=(0, 10), pady=(15, 5))
        observation_entry = ctk.CTkEntry(fields_frame, textvariable=observation_var, width=400, height=35, font=ctk.CTkFont(size=14))
        observation_entry.grid(row=len(fields)*2+1, column=0, columnspan=2, sticky="ew", padx=(0, 20), pady=(0, 15))
        
        # Boutons d'action
        buttons_frame = ctk.CTkFrame(modal, fg_color="transparent")
        buttons_frame.pack(pady=(0, 20))
        
        def create_expedition():
            """Créer l'expédition"""
            # Vérifier les champs obligatoires avec messages spécifiques
            missing_fields = []
            if not client_var.get().strip():
                missing_fields.append("Client")
            if not weight_var.get().strip():
                missing_fields.append("Poids")
            if not packages_var.get().strip():
                missing_fields.append("Nombre de colis")
            
            if missing_fields:
                # Afficher une erreur détaillée
                error_text = f"❌ Champs obligatoires manquants : {', '.join(missing_fields)}"
                error_label = ctk.CTkLabel(modal, text=error_text, text_color="#ef4444", font=ctk.CTkFont(size=14, weight="bold"))
                error_label.pack(pady=10)
                modal.after(4000, error_label.destroy)
                return
            
            # Validation du poids
            poids_str = weight_var.get().strip()
            try:
                poids_float = float(poids_str) if poids_str else 0.0
                if poids_float < 0:
                    error_label = ctk.CTkLabel(modal, text="❌ Le poids ne peut pas être négatif", text_color="#ef4444", font=ctk.CTkFont(size=14, weight="bold"))
                    error_label.pack(pady=10)
                    modal.after(3000, error_label.destroy)
                    return
            except ValueError:
                error_label = ctk.CTkLabel(modal, text="❌ Le poids doit être un nombre valide", text_color="#ef4444", font=ctk.CTkFont(size=14, weight="bold"))
                error_label.pack(pady=10)
                modal.after(3000, error_label.destroy)
                return
            
            # Préparer les données pour la BD
            expedition_data = {
                'client': client_var.get().strip(),
                'reference_commande': f"REF{datetime.now().strftime('%H%M')[-2:]}",  # Référence courte (5 caractères max)
                'date_livraison': datetime.now().strftime('%Y-%m-%d'),
                'transporteurs': carrier_var.get(),
                'priorite': priority_var.get(),
                'poids': poids_str,
                'observation': observation_var.get().strip() or f"Poids: {poids_str}kg, Colis: {packages_var.get()}",
                'liste_articles_livres': ''
            }
            
            print(f"🔄 Données de création: poids={poids_str} kg")
            
            try:
                if DB_AVAILABLE:
                    # Créer dans la base de données
                    expedition_id = add_expedition(expedition_data)
                    if expedition_id:
                        print(f"✅ Expédition créée dans la BD avec l'ID: {expedition_id}")
                        # Recharger les données
                        self.load_expeditions_data()
                        # Mettre à jour l'affichage
                        self.update_expeditions_display()
                        # Mettre à jour les blocs centraux (notifications, planning, etc.)
                        self.update_central_blocks()
                        # Mettre à jour les statistiques
                        self.update_stats_display()
                        success_msg = "✅ Expédition créée avec succès !"
                    else:
                        print("❌ Erreur lors de la création dans la BD")
                        success_msg = "❌ Erreur lors de la création"
                else:
                    # Mode démonstration
                    _add_local_expedition(expedition_data)
                    success_msg = "✅ Expédition créée (mode démonstration)"
                
                # Afficher un message de succès
                success_label = ctk.CTkLabel(modal, text=success_msg, text_color="#10b981" if "succès" in success_msg else "#ef4444", font=ctk.CTkFont(size=16, weight="bold"))
                success_label.pack(pady=10)
                
                # Fermer la modal après 2 secondes
                modal.after(2000, modal.destroy)
                    
            except Exception as e:
                print(f"❌ Erreur lors de la création: {e}")
                # Fallback avec données locales
                _add_local_expedition(expedition_data)
                success_label = ctk.CTkLabel(modal, text="✅ Expédition créée (mode démonstration)", text_color="#10b981", font=ctk.CTkFont(size=16, weight="bold"))
                success_label.pack(pady=10)
                modal.after(2000, modal.destroy)
        
        def _add_local_expedition(expedition_data):
            """Ajoute une expédition en mode local (démonstration)"""
            # Convertir le poids en float
            poids = 0.0
            try:
                poids = float(expedition_data.get('poids', 0))
            except (ValueError, TypeError):
                poids = 0.0
            
            new_expedition = {
                'id': len(self.expeditions_data) + 1,
                'number': f'BE-2024-{len(self.expeditions_data) + 1:03d}',
                'client': expedition_data['client'],
                'shippingDate': datetime.now().strftime('%Y-%m-%d'),
                'deliveryDate': expedition_data['date_livraison'],
                'actualDeliveryDate': None,
                'carrier': expedition_data['transporteurs'],
                'priority': expedition_data.get('priorite', 'moyenne'),
                'packages': 1,
                'totalWeight': poids,
                'status': 'preparing',
                'trackingNumber': expedition_data['reference_commande'],
                'observation': expedition_data['observation'],
                'articles': expedition_data['liste_articles_livres']
            }
            
            # Ajouter à la liste
            self.expeditions_data.append(new_expedition)
            
            # Mettre à jour l'affichage principal
            self.update_expeditions_display()
            
            # Mettre à jour les blocs centraux (notifications, planning, etc.)
            self.update_central_blocks()
            
            # Mettre à jour les statistiques
            self.update_stats_display()
        
        def cancel_creation():
            """Annuler la création"""
            modal.destroy()
        
        # Boutons
        ctk.CTkButton(buttons_frame, text="❌ Annuler", width=120, height=40, fg_color="#ef4444", hover_color="#dc2626", command=cancel_creation).pack(side="left", padx=10)
        ctk.CTkButton(buttons_frame, text="✅ Créer l'expédition", width=150, height=40, fg_color="#10b981", hover_color="#059669", command=create_expedition).pack(side="left", padx=10)

    def action_planning(self):
        """Afficher le planning dynamique basé sur la BD"""
        print("Affichage du planning")
        
        # Créer une fenêtre modale pour le planning
        modal = ctk.CTkToplevel(self)
        modal.title("Planning des Expéditions")
        modal.geometry("1000x700")
        modal.configure(fg_color='white')
        modal.grab_set()
        modal.resizable(True, True)
        
        # Titre
        title_label = ctk.CTkLabel(modal, text="📅 Planning des Expéditions", font=ctk.CTkFont(size=24, weight="bold"), text_color="#1f2937")
        title_label.pack(pady=(20, 30))
        
        # Frame principal
        main_frame = ctk.CTkScrollableFrame(modal, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # Récupérer le planning dynamique depuis la BD
        planning = self._get_today_planning()
        if not planning:
            ctk.CTkLabel(main_frame, text="📅 Aucune expédition prévue aujourd'hui", font=ctk.CTkFont(size=18, weight="bold"), text_color="#6b7280").pack(pady=30)
            return
        
        # Afficher chaque créneau du planning
        for slot in planning:
            slot_frame = ctk.CTkFrame(main_frame, fg_color="white", border_width=1, border_color="#e5e7eb", corner_radius=8)
            slot_frame.pack(fill="x", pady=(0, 15))
            ctk.CTkLabel(slot_frame, text=slot.get("heure", ""), font=ctk.CTkFont(size=18, weight="bold"), text_color="#3867d6").pack(anchor="w", padx=20, pady=(15, 10))
            ctk.CTkLabel(slot_frame, text=slot["desc"], font=ctk.CTkFont(size=15), text_color="#374151").pack(anchor="w", padx=20, pady=(0, 5))
            ctk.CTkLabel(slot_frame, text=slot["statut"], font=ctk.CTkFont(size=14, weight="bold"), text_color=slot["color"]).pack(anchor="e", padx=20, pady=(0, 10))
        
        # Bouton de fermeture
        ctk.CTkButton(modal, text="Fermer", width=120, height=40, command=modal.destroy).pack(pady=20)

    def open_tracking_modal(self):
        """Ouvrir la modal de suivi avancé"""
        print("Ouverture de la modal de suivi avancé")
        
        # Créer une fenêtre modale pour le suivi
        modal = ctk.CTkToplevel(self)
        modal.title("🔍 Suivi Avancé des Expéditions")
        modal.geometry("1000x700")
        modal.configure(fg_color='white')
        modal.grab_set()
        modal.resizable(True, True)
        
        # Titre principal
        title_frame = ctk.CTkFrame(modal, fg_color="transparent")
        title_frame.pack(fill="x", padx=20, pady=(20, 10))
        
        ctk.CTkLabel(title_frame, text="🔍 Suivi Avancé des Expéditions", font=ctk.CTkFont(size=24, weight="bold"), text_color="#1f2937").pack(side="left")
        
        # Bouton de rafraîchissement
        refresh_btn = ctk.CTkButton(title_frame, text="🔄 Actualiser", width=100, height=32, fg_color="#3b82f6", hover_color="#2563eb", command=lambda: refresh_tracking())
        refresh_btn.pack(side="right")
        
        # Frame de recherche avancée
        search_frame = ctk.CTkFrame(modal, fg_color="#f8fafc", border_width=1, border_color="#e2e8f0", corner_radius=12)
        search_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        # Titre de recherche
        ctk.CTkLabel(search_frame, text="🔎 Recherche d'Expédition", font=ctk.CTkFont(size=18, weight="bold"), text_color="#374151").pack(anchor="w", padx=20, pady=(15, 10))
        
        # Frame pour les champs de recherche
        search_fields_frame = ctk.CTkFrame(search_frame, fg_color="transparent")
        search_fields_frame.pack(fill="x", padx=20, pady=(0, 15))
        
        # Variables de recherche
        search_var = tk.StringVar()
        status_filter_var = tk.StringVar(value="Tous")
        carrier_filter_var = tk.StringVar(value="Tous")
        
        # Numéro d'expédition
        ctk.CTkLabel(search_fields_frame, text="Numéro d'expédition:", font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151").grid(row=0, column=0, sticky="w", padx=(0, 10), pady=5)
        search_entry = ctk.CTkEntry(search_fields_frame, textvariable=search_var, placeholder_text="BE-2024-001", width=200, height=35)
        search_entry.grid(row=0, column=1, sticky="w", padx=(0, 20), pady=5)
        
        # Filtre par statut
        ctk.CTkLabel(search_fields_frame, text="Statut:", font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151").grid(row=0, column=2, sticky="w", padx=(0, 10), pady=5)
        status_combo = ctk.CTkOptionMenu(search_fields_frame, variable=status_filter_var, values=["Tous", "En préparation", "Expédiée", "En transit", "Livrée"], width=120, height=35)
        status_combo.grid(row=0, column=3, sticky="w", padx=(0, 20), pady=5)
        
        # Filtre par transporteur
        ctk.CTkLabel(search_fields_frame, text="Transporteur:", font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151").grid(row=0, column=4, sticky="w", padx=(0, 10), pady=5)
        carrier_combo = ctk.CTkOptionMenu(search_fields_frame, variable=carrier_filter_var, values=["Tous", "DHL Express", "Chronopost", "Colissimo", "UPS"], width=120, height=35)
        carrier_combo.grid(row=0, column=5, sticky="w", padx=(0, 20), pady=5)
        
        # Bouton de recherche
        search_btn = ctk.CTkButton(search_fields_frame, text="🔍 Rechercher", width=120, height=35, fg_color="#10b981", hover_color="#059669", command=lambda: perform_search())
        search_btn.grid(row=0, column=6, sticky="w", pady=5)
        
        # Frame pour les résultats
        results_frame = ctk.CTkFrame(modal, fg_color="transparent")
        results_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        def perform_search():
            """Effectuer la recherche avec filtres"""
            search_term = search_var.get().strip()
            status_filter = status_filter_var.get()
            carrier_filter = carrier_filter_var.get()
            
            # Nettoyer les résultats précédents
            for widget in results_frame.winfo_children():
                widget.destroy()
            
            try:
                if DB_AVAILABLE:
                    # Recherche dans la base de données avec filtres
                    found_expeditions = self._search_expeditions_db(search_term, status_filter, carrier_filter)
                    else:
                    # Recherche locale avec filtres
                    found_expeditions = self._search_expeditions_local(search_term, status_filter, carrier_filter)
                
                if found_expeditions:
                    show_search_results(found_expeditions)
                    else:
                    show_no_results()
                        
            except Exception as e:
                print(f"❌ Erreur lors de la recherche: {e}")
                show_no_results()
        
        def show_search_results(expeditions):
            """Afficher les résultats de recherche"""
            # Titre des résultats
            results_title = ctk.CTkLabel(results_frame, text=f"📋 Résultats ({len(expeditions)} expédition(s))", font=ctk.CTkFont(size=18, weight="bold"), text_color="#374151")
            results_title.pack(anchor="w", pady=(0, 15))
            
            # Frame scrollable pour les résultats
            scroll_frame = ctk.CTkScrollableFrame(results_frame, fg_color="transparent")
            scroll_frame.pack(fill="both", expand=True)
            
            for expedition in expeditions:
                # Carte d'expédition
                exp_card = ctk.CTkFrame(scroll_frame, fg_color="white", border_width=1, border_color="#e5e7eb", corner_radius=8)
                exp_card.pack(fill="x", pady=(0, 10))
                
                # En-tête de la carte
                header_frame = ctk.CTkFrame(exp_card, fg_color="transparent")
                header_frame.pack(fill="x", padx=15, pady=(15, 10))
                
                # Numéro et statut
                ctk.CTkLabel(header_frame, text=f"📦 {expedition['number']}", font=ctk.CTkFont(size=16, weight="bold"), text_color="#1f2937").pack(side="left")
                
                # Badge de statut
                status_color = self._get_status_color(expedition['status'])
                status_badge = ctk.CTkFrame(header_frame, fg_color=status_color, corner_radius=12)
                status_badge.pack(side="right", padx=(10, 0))
                ctk.CTkLabel(status_badge, text=expedition['status'].title(), font=ctk.CTkFont(size=12, weight="bold"), text_color="white").pack(padx=8, pady=2)
                
                # Informations principales
                info_frame = ctk.CTkFrame(exp_card, fg_color="transparent")
                info_frame.pack(fill="x", padx=15, pady=(0, 10))
                
                info_text = f"👤 Client: {expedition['client']} | 🚚 Transporteur: {expedition['carrier']} | ⚖️ Poids: {expedition['totalWeight']} kg"
                ctk.CTkLabel(info_frame, text=info_text, font=ctk.CTkFont(size=14), text_color="#6b7280").pack(anchor="w")
                
                # Boutons d'action
                actions_frame = ctk.CTkFrame(exp_card, fg_color="transparent")
                actions_frame.pack(fill="x", padx=15, pady=(0, 15))
                
                ctk.CTkButton(actions_frame, text="👁️ Voir détails", width=100, height=30, fg_color="#3b82f6", hover_color="#2563eb", command=lambda e=expedition: show_expedition_details(e)).pack(side="left", padx=(0, 10))
                ctk.CTkButton(actions_frame, text="📋 Timeline", width=100, height=30, fg_color="#10b981", hover_color="#059669", command=lambda e=expedition: show_timeline(e)).pack(side="left", padx=(0, 10))
                ctk.CTkButton(actions_frame, text="✏️ Modifier", width=100, height=30, fg_color="#f59e0b", hover_color="#d97706", command=lambda e=expedition: edit_from_tracking(e)).pack(side="left")
        
        def show_expedition_details(expedition):
            """Afficher les détails complets d'une expédition"""
            # Créer une nouvelle fenêtre pour les détails
            details_modal = ctk.CTkToplevel(modal)
            details_modal.title(f"Détails - {expedition['number']}")
            details_modal.geometry("800x600")
            details_modal.configure(fg_color='white')
            details_modal.grab_set()
            
            # Titre
            ctk.CTkLabel(details_modal, text=f"📦 {expedition['number']}", font=ctk.CTkFont(size=24, weight="bold"), text_color="#1f2937").pack(pady=(20, 30))
            
            # Frame principal avec scroll
            main_frame = ctk.CTkScrollableFrame(details_modal, fg_color="transparent")
            main_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
            
            # Informations détaillées
            details_frame = ctk.CTkFrame(main_frame, fg_color="white", border_width=1, border_color="#e5e7eb", corner_radius=8)
            details_frame.pack(fill="x", pady=(0, 20))
            
            ctk.CTkLabel(details_frame, text="📋 Informations détaillées", font=ctk.CTkFont(size=18, weight="bold"), text_color="#374151").pack(anchor="w", padx=20, pady=(20, 15))
            
            # Grille d'informations
            info_grid = ctk.CTkFrame(details_frame, fg_color="transparent")
            info_grid.pack(fill="x", padx=20, pady=(0, 20))
            
            info_data = [
                ("Client:", expedition['client']),
                ("Transporteur:", expedition['carrier']),
                ("Priorité:", expedition['priority'].title()),
                ("Statut:", expedition['status'].title()),
                ("Poids:", f"{expedition['totalWeight']} kg"),
                ("Nombre de colis:", str(expedition['packages'])),
                ("Date d'expédition:", expedition.get('shippingDate', 'Non spécifiée')),
                ("Date de livraison prévue:", expedition.get('deliveryDate', 'Non spécifiée')),
                ("Numéro de suivi:", expedition.get('trackingNumber', 'Non spécifié')),
                ("Observation:", expedition.get('observation', 'Aucune')),
            ]
            
            for i, (label, value) in enumerate(info_data):
                row = i // 2
                col = i % 2
                
                ctk.CTkLabel(info_grid, text=label, font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151").grid(row=row*2, column=col, sticky="w", padx=(0, 10), pady=(10, 5))
                ctk.CTkLabel(info_grid, text=value, font=ctk.CTkFont(size=14), text_color="#6b7280").grid(row=row*2+1, column=col, sticky="w", padx=(0, 20), pady=(0, 10))
            
            # Timeline de suivi
            timeline_frame = ctk.CTkFrame(main_frame, fg_color="#f9fafb", border_width=1, border_color="#e5e7eb", corner_radius=8)
            timeline_frame.pack(fill="x", pady=(0, 20))
            
            ctk.CTkLabel(timeline_frame, text="📅 Timeline de suivi", font=ctk.CTkFont(size=18, weight="bold"), text_color="#374151").pack(anchor="w", padx=20, pady=(20, 15))
            
            # Timeline dynamique basée sur le statut
            timeline_events = self._generate_timeline(expedition)
            
            for event in timeline_events:
                event_frame = ctk.CTkFrame(timeline_frame, fg_color="transparent")
                event_frame.pack(fill="x", padx=20, pady=5)
                
                # Icône de statut
                ctk.CTkLabel(event_frame, text=event["icon"], font=ctk.CTkFont(size=16)).pack(side="left")
                
                # Informations de l'événement
                event_info_frame = ctk.CTkFrame(event_frame, fg_color="transparent")
                event_info_frame.pack(side="left", fill="x", expand=True, padx=(10, 0))
                
                ctk.CTkLabel(event_info_frame, text=event["event"], font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151").pack(anchor="w")
                ctk.CTkLabel(event_info_frame, text=event["description"], font=ctk.CTkFont(size=12), text_color="#6b7280").pack(anchor="w")
                
                # Date
                ctk.CTkLabel(event_frame, text=event["date"], font=ctk.CTkFont(size=12), text_color="#9ca3af").pack(side="right")
            
            # Bouton de fermeture
            ctk.CTkButton(details_modal, text="Fermer", width=120, height=40, command=details_modal.destroy).pack(pady=20)
        
        def show_timeline(expedition):
            """Afficher uniquement la timeline d'une expédition"""
            timeline_modal = ctk.CTkToplevel(modal)
            timeline_modal.title(f"Timeline - {expedition['number']}")
            timeline_modal.geometry("600x500")
            timeline_modal.configure(fg_color='white')
            timeline_modal.grab_set()
            
            # Titre
            ctk.CTkLabel(timeline_modal, text=f"📅 Timeline - {expedition['number']}", font=ctk.CTkFont(size=20, weight="bold"), text_color="#1f2937").pack(pady=(20, 30))
            
            # Frame principal
            main_frame = ctk.CTkScrollableFrame(timeline_modal, fg_color="transparent")
            main_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
            
            # Timeline
            timeline_events = self._generate_timeline(expedition)
            
            for event in timeline_events:
                event_frame = ctk.CTkFrame(main_frame, fg_color="white", border_width=1, border_color="#e5e7eb", corner_radius=8)
                event_frame.pack(fill="x", pady=(0, 10))
                
                # En-tête de l'événement
                header_frame = ctk.CTkFrame(event_frame, fg_color="transparent")
                header_frame.pack(fill="x", padx=15, pady=(15, 10))
                
                ctk.CTkLabel(header_frame, text=event["icon"], font=ctk.CTkFont(size=18)).pack(side="left")
                ctk.CTkLabel(header_frame, text=event["event"], font=ctk.CTkFont(size=16, weight="bold"), text_color="#374151").pack(side="left", padx=(10, 0))
                ctk.CTkLabel(header_frame, text=event["date"], font=ctk.CTkFont(size=12), text_color="#9ca3af").pack(side="right")
                
                # Description
                ctk.CTkLabel(event_frame, text=event["description"], font=ctk.CTkFont(size=14), text_color="#6b7280").pack(anchor="w", padx=15, pady=(0, 15))
            
            # Bouton de fermeture
            ctk.CTkButton(timeline_modal, text="Fermer", width=120, height=40, command=timeline_modal.destroy).pack(pady=20)
        
        def edit_from_tracking(expedition):
            """Modifier une expédition depuis le suivi"""
            modal.destroy()  # Fermer la modal de suivi
            self.edit_expedition(expedition)  # Ouvrir la modal d'édition
        
        def show_no_results():
            """Afficher message si aucun résultat"""
            no_results_frame = ctk.CTkFrame(results_frame, fg_color="white", border_width=1, border_color="#e5e7eb", corner_radius=8)
            no_results_frame.pack(fill="both", expand=True)
            
            ctk.CTkLabel(no_results_frame, text="🔍 Aucune expédition trouvée", font=ctk.CTkFont(size=18, weight="bold"), text_color="#6b7280").pack(expand=True, pady=(50, 10))
            ctk.CTkLabel(no_results_frame, text="Essayez avec d'autres critères de recherche", font=ctk.CTkFont(size=14), text_color="#9ca3af").pack()
        
        def refresh_tracking():
            """Actualiser les données de suivi"""
            print("🔄 Actualisation du suivi...")
            self.load_expeditions_data()
            perform_search()
        
        # Recherche initiale pour afficher toutes les expéditions
        perform_search()
        
        # Bouton de fermeture
        ctk.CTkButton(modal, text="Fermer", width=120, height=40, command=modal.destroy).pack(pady=20)

    def action_export(self):
        """Exporter les données"""
        print("Export des données")
        
        # Créer une fenêtre modale pour l'export
        modal = ctk.CTkToplevel(self)
        modal.title("Export des Expéditions")
        modal.geometry("700x600")
        modal.configure(fg_color='white')
        modal.grab_set()
        modal.resizable(False, False)
        
        # Titre
        title_label = ctk.CTkLabel(modal, text="⭳ Export des Expéditions", font=ctk.CTkFont(size=24, weight="bold"), text_color="#1f2937")
        title_label.pack(pady=(20, 30))
        
        # Frame principal
        main_frame = ctk.CTkFrame(modal, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # Options d'export
        options_frame = ctk.CTkFrame(main_frame, fg_color="white", border_width=1, border_color="#e5e7eb", corner_radius=8)
        options_frame.pack(fill="x", pady=(0, 20))
        
        ctk.CTkLabel(options_frame, text="Options d'export", font=ctk.CTkFont(size=18, weight="bold"), text_color="#374151").pack(anchor="w", padx=20, pady=(20, 15))
        
        # Variables
        format_var = tk.StringVar(value="Excel")
        date_range_var = tk.StringVar(value="Toutes")
        status_var = tk.StringVar(value="Tous")
        
        # Format d'export
        format_frame = ctk.CTkFrame(options_frame, fg_color="transparent")
        format_frame.pack(fill="x", padx=20, pady=(0, 15))
        
        ctk.CTkLabel(format_frame, text="Format:", font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151").pack(side="left")
        format_combo = ctk.CTkOptionMenu(format_frame, variable=format_var, values=["Excel", "CSV", "PDF"], width=150)
        format_combo.pack(side="left", padx=(10, 0))
        
        # Période
        period_frame = ctk.CTkFrame(options_frame, fg_color="transparent")
        period_frame.pack(fill="x", padx=20, pady=(0, 15))
        
        ctk.CTkLabel(period_frame, text="Période:", font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151").pack(side="left")
        period_combo = ctk.CTkOptionMenu(period_frame, variable=date_range_var, values=["Toutes", "Aujourd'hui", "Cette semaine", "Ce mois"], width=150)
        period_combo.pack(side="left", padx=(10, 0))
        
        # Statut
        status_frame = ctk.CTkFrame(options_frame, fg_color="transparent")
        status_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        ctk.CTkLabel(status_frame, text="Statut:", font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151").pack(side="left")
        status_combo = ctk.CTkOptionMenu(status_frame, variable=status_var, values=["Tous", "En préparation", "Expédiées", "Livrées"], width=150)
        status_combo.pack(side="left", padx=(10, 0))
        
        # Résumé
        summary_frame = ctk.CTkFrame(main_frame, fg_color="#f0f9ff", border_width=1, border_color="#0ea5e9", corner_radius=8)
        summary_frame.pack(fill="x", pady=(0, 20))
        
        ctk.CTkLabel(summary_frame, text="📊 Résumé de l'export", font=ctk.CTkFont(size=16, weight="bold"), text_color="#0c4a6e").pack(anchor="w", padx=15, pady=(15, 10))
        
        summary_text = f"• Format: {format_var.get()}\n• Période: {date_range_var.get()}\n• Statut: {status_var.get()}\n• Nombre d'expéditions: {len(self.expeditions_data)}"
        ctk.CTkLabel(summary_frame, text=summary_text, font=ctk.CTkFont(size=14), text_color="#0c4a6e", justify="left").pack(anchor="w", padx=15, pady=(0, 15))
        
        # Zone de statut
        status_display_frame = ctk.CTkFrame(main_frame, fg_color="white", border_width=1, border_color="#e5e7eb", corner_radius=8)
        status_display_frame.pack(fill="x", pady=(0, 20))
        
        status_label = ctk.CTkLabel(status_display_frame, text="⏳ Prêt à exporter...", font=ctk.CTkFont(size=14), text_color="#6b7280")
        status_label.pack(pady=15)
        
        # Boutons d'action
        buttons_frame = ctk.CTkFrame(modal, fg_color="transparent")
        buttons_frame.pack(pady=(0, 20))
        
        def perform_export():
            """Effectuer l'export réel"""
            try:
                # Mettre à jour le statut
                status_label.configure(text="🔄 Export en cours...", text_color="#f59e0b")
                modal.update()
                
                # Filtrer les données selon les critères
                filtered_data = self._filter_export_data(date_range_var.get(), status_var.get())
                
                # Générer le fichier selon le format
                export_format = format_var.get().lower()
                file_path = self._generate_export_file(filtered_data, export_format)
                
                if file_path:
                    # Afficher le succès
                    status_label.configure(text=f"✅ Export terminé !\n📁 Fichier: {file_path}", text_color="#10b981")
                    
                    # Bouton pour ouvrir le dossier
                    open_folder_btn = ctk.CTkButton(
                        status_display_frame, 
                        text="📂 Ouvrir le dossier", 
                        width=150, 
                        height=35, 
                        fg_color="#3b82f6", 
                        hover_color="#2563eb",
                        command=lambda: self._open_export_folder(file_path)
                    )
                    open_folder_btn.pack(pady=(10, 0))
                    
                    # Fermer la modal après 5 secondes
                    modal.after(5000, modal.destroy)
                else:
                    status_label.configure(text="❌ Erreur lors de l'export", text_color="#ef4444")
                    
            except Exception as e:
                print(f"❌ Erreur export: {e}")
                status_label.configure(text=f"❌ Erreur: {str(e)}", text_color="#ef4444")
        
        def cancel_export():
            """Annuler l'export"""
            modal.destroy()
        
        # Boutons centrés et alignés avec couleurs bien visibles
        ctk.CTkButton(
            buttons_frame, 
            text="❌ Annuler", 
            width=140, 
            height=40, 
            fg_color="#dc2626", 
            hover_color="#b91c1c", 
            text_color="white",
            font=ctk.CTkFont(size=14, weight="bold"),
            corner_radius=8,
            command=cancel_export
        ).pack(side="left", padx=(0, 15))
        
        ctk.CTkButton(
            buttons_frame, 
            text="⭳ Exporter", 
            width=140, 
            height=40, 
            fg_color="#059669", 
            hover_color="#047857", 
            text_color="white",
            font=ctk.CTkFont(size=14, weight="bold"),
            corner_radius=8,
            command=perform_export
        ).pack(side="left", padx=(15, 0))
        
        # Centrer les boutons dans le frame
        buttons_frame.pack_configure(anchor="center")

    def _filter_export_data(self, date_range, status_filter):
        """Filtrer les données d'export selon les critères"""
        from datetime import datetime, timedelta
        
        filtered_data = []
        today = datetime.now().date()
        
        for exp in self.expeditions_data:
            # Filtre par période
            if date_range != "Toutes":
                delivery_date = exp.get('deliveryDate')
                if delivery_date:
                    try:
                        delivery_date_obj = datetime.strptime(delivery_date, '%Y-%m-%d').date()
                        
                        if date_range == "Aujourd'hui" and delivery_date_obj != today:
                            continue
                        elif date_range == "Cette semaine":
                            week_start = today - timedelta(days=today.weekday())
                            week_end = week_start + timedelta(days=6)
                            if not (week_start <= delivery_date_obj <= week_end):
                                continue
                        elif date_range == "Ce mois":
                            if delivery_date_obj.month != today.month or delivery_date_obj.year != today.year:
                                continue
                    except:
                        continue
            
            # Filtre par statut
            if status_filter != "Tous":
                delivery_date = exp.get('deliveryDate')
                if delivery_date:
                    try:
                        delivery_date_obj = datetime.strptime(delivery_date, '%Y-%m-%d').date()
                        
                        if delivery_date_obj < today:
                            current_status = "Livrées"
                        elif delivery_date_obj == today:
                            current_status = "Expédiées"
                        else:
                            current_status = "En préparation"
                    except:
                        current_status = "En préparation"
                else:
                    current_status = "En préparation"
                
                if current_status != status_filter:
                    continue
            
            filtered_data.append(exp)
        
        return filtered_data

    def _generate_export_file(self, data, format_type):
        """Générer le fichier d'export selon le format"""
        import os
        from datetime import datetime
        
        # Créer le dossier exports s'il n'existe pas
        export_dir = "exports"
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)
        
        # Nom du fichier avec timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format_type == "excel":
            return self._export_to_excel(data, export_dir, timestamp)
        elif format_type == "csv":
            return self._export_to_csv(data, export_dir, timestamp)
        elif format_type == "pdf":
            return self._export_to_pdf(data, export_dir, timestamp)
        else:
            return None

    def _export_to_excel(self, data, export_dir, timestamp):
        """Exporter vers Excel"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Préparer les données
            export_data = []
            for exp in data:
                export_data.append({
                    'Numéro': exp.get('number', ''),
                    'Client': exp.get('client', ''),
                    'Date d\'expédition': exp.get('shippingDate', ''),
                    'Date de livraison': exp.get('deliveryDate', ''),
                    'Transporteur': exp.get('carrier', ''),
                    'Priorité': exp.get('priority', ''),
                    'Statut': exp.get('status', ''),
                    'Poids (kg)': exp.get('totalWeight', 0),
                    'Nombre de colis': exp.get('packages', 1),
                    'Numéro de suivi': exp.get('trackingNumber', ''),
                    'Observation': exp.get('observation', '')
                })
            
            # Créer le DataFrame
            df = pd.DataFrame(export_data)
            
            # Nom du fichier
            filename = f"expeditions_{timestamp}.xlsx"
            filepath = os.path.join(export_dir, filename)
            
            # Créer le workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Expéditions"
            
            # Ajouter les données
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Styliser l'en-tête
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Ajuster la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Sauvegarder
            wb.save(filepath)
            print(f"✅ Export Excel créé: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"❌ Erreur export Excel: {e}")
            return None

    def _export_to_csv(self, data, export_dir, timestamp):
        """Exporter vers CSV"""
        try:
            import pandas as pd
            
            # Préparer les données
            export_data = []
            for exp in data:
                export_data.append({
                    'Numéro': exp.get('number', ''),
                    'Client': exp.get('client', ''),
                    'Date d\'expédition': exp.get('shippingDate', ''),
                    'Date de livraison': exp.get('deliveryDate', ''),
                    'Transporteur': exp.get('carrier', ''),
                    'Priorité': exp.get('priority', ''),
                    'Statut': exp.get('status', ''),
                    'Poids (kg)': exp.get('totalWeight', 0),
                    'Nombre de colis': exp.get('packages', 1),
                    'Numéro de suivi': exp.get('trackingNumber', ''),
                    'Observation': exp.get('observation', '')
                })
            
            # Créer le DataFrame
            df = pd.DataFrame(export_data)
            
            # Nom du fichier
            filename = f"expeditions_{timestamp}.csv"
            filepath = os.path.join(export_dir, filename)
            
            # Sauvegarder
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            print(f"✅ Export CSV créé: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"❌ Erreur export CSV: {e}")
            return None

    def _export_to_pdf(self, data, export_dir, timestamp):
        """Exporter vers PDF"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            
            # Nom du fichier
            filename = f"expeditions_{timestamp}.pdf"
            filepath = os.path.join(export_dir, filename)
            
            # Créer le document
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            story = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Centré
            )
            
            # Titre
            title = Paragraph("Rapport des Expéditions", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
            # En-têtes
            headers = ['Numéro', 'Client', 'Date Livraison', 'Transporteur', 'Priorité', 'Statut', 'Poids']
            
            # Données
            table_data = [headers]
            for exp in data:
                row = [
                    exp.get('number', ''),
                    exp.get('client', ''),
                    exp.get('deliveryDate', ''),
                    exp.get('carrier', ''),
                    exp.get('priority', ''),
                    exp.get('status', ''),
                    f"{exp.get('totalWeight', 0)} kg"
                ]
                table_data.append(row)
            
            # Créer la table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            story.append(table)
            
            # Construire le PDF
            doc.build(story)
            print(f"✅ Export PDF créé: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"❌ Erreur export PDF: {e}")
            return None

    def _open_export_folder(self, file_path):
        """Ouvrir le dossier contenant le fichier exporté"""
        try:
            import os
            import subprocess
            import platform
            
            folder_path = os.path.dirname(os.path.abspath(file_path))
            
            if platform.system() == "Windows":
                subprocess.run(["explorer", folder_path])
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", folder_path])
            else:  # Linux
                subprocess.run(["xdg-open", folder_path])
                
        except Exception as e:
            print(f"❌ Erreur ouverture dossier: {e}")

    def apply_theme(self, theme):
        """Appliquer le thème à la page"""
        try:
            print(f"Thème appliqué à la page Expéditions: {theme}")
            
            # Couleurs selon le thème
            if theme == "dark":
                bg_color = "#1a1a1a"
                card_bg = "#2d2d2d"
                text_color = "#ffffff"
                border_color = "#555555"
            else:
                bg_color = "#ffffff"
                card_bg = "#ffffff"
                text_color = "#222222"
                border_color = "#e0e0e0"
            
            # Appliquer aux widgets principaux
            self.configure(fg_color=bg_color)
            
            # Appliquer récursivement aux enfants
            self._apply_theme_recursive(self, theme)
            
        except Exception as e:
            print(f"Erreur lors de l'application du thème: {e}")
    
    def _apply_theme_recursive(self, widget, theme):
        """Applique le thème de manière récursive"""
        try:
            is_dark = theme == "dark"
            
            # Couleurs adaptatives
            bg_color = "#1a1a1a" if is_dark else "#f7fafd"
            card_bg = "#2d2d2d" if is_dark else "white"
            text_color = "#ffffff" if is_dark else "#222222"
            border_color = "#555555" if is_dark else "#e0e0e0"
            
            # Appliquer aux widgets principaux
            if isinstance(widget, ctk.CTkFrame):
                widget.configure(fg_color=card_bg, border_color=border_color)
            elif isinstance(widget, ctk.CTkLabel):
                widget.configure(text_color=text_color)
            
            # Appliquer récursivement aux enfants
            for child in widget.winfo_children():
                self._apply_theme_recursive(child, theme)
                
        except Exception as e:
            # Ignorer les erreurs pour éviter les boucles infinies
            pass

    def _refresh_planning_periodically(self):
        """Rafraîchit le planning toutes les 30 secondes en temps réel"""
        try:
            if hasattr(self, 'planning_frame') and self.planning_frame:
                self.create_central_blocks()  # Reconstruit le bloc planning
        except Exception as e:
            print(f"Erreur lors du rafraîchissement du planning: {e}")
        self.after(30000, self._refresh_planning_periodically)  # 30 secondes

    def _auto_refresh(self):
        self.refresh_data()
        self.after(self._auto_refresh_interval, self._auto_refresh)

    def _go_back(self):
        """Retourne à la page précédente si possible, sinon dashboard"""
        if hasattr(self.master, 'show_dashboard') and self.previous_page is None:
            self.master.show_dashboard(self.user_info)
        elif hasattr(self.master, 'show_' + str(self.previous_page)):
            getattr(self.master, 'show_' + str(self.previous_page))()
        elif hasattr(self.master, 'show_dashboard'):
            self.master.show_dashboard(self.user_info)
        else:
            messagebox.showinfo("Retour", "Impossible de revenir en arrière. Redémarrez l'application.")

    def _show_error(self, message):
        for widget in self.winfo_children():
            widget.destroy()
        error_frame = ctk.CTkFrame(self, fg_color="#fee2e2", corner_radius=12)
        error_frame.pack(fill="both", expand=True, padx=60, pady=60)
        ctk.CTkLabel(error_frame, text="❌ Erreur lors de l'affichage de la page expédition", font=ctk.CTkFont(size=20, weight="bold"), text_color="#b91c1c").pack(pady=(30, 10))
        ctk.CTkLabel(error_frame, text=message, font=ctk.CTkFont(size=15), text_color="#b91c1c").pack(pady=(0, 20))
        ctk.CTkButton(error_frame, text="🏠 Retour à l'accueil", fg_color="#2563eb", text_color="white", font=ctk.CTkFont(size=15, weight="bold"), command=self._go_back).pack(pady=10)

    def _update_local_expedition(self, expedition_id, expedition_data):
        """Met à jour une expédition en mode local (démonstration)"""
        print(f"🔄 Mise à jour locale de l'expédition ID: {expedition_id}")
        
        for i, exp in enumerate(self.expeditions_data):
            if exp['id'] == expedition_id:
                # Convertir le poids en float
                poids = 0.0
                try:
                    poids = float(expedition_data.get('poids', 0))
                except (ValueError, TypeError):
                    poids = 0.0
                
                self.expeditions_data[i].update({
                    'client': expedition_data['client'],
                    'carrier': expedition_data['transporteurs'],
                    'priority': expedition_data.get('priorite', 'moyenne'),
                    'observation': expedition_data['observation'],
                    'trackingNumber': expedition_data['reference_commande'],
                    'totalWeight': poids
                })
                print(f"✅ Expédition {expedition_id} mise à jour avec le poids: {poids} kg")
                break
        
        # Mettre à jour l'affichage principal
        self.update_expeditions_display()
        
        # Mettre à jour les blocs centraux (notifications, planning, etc.)
        self.update_central_blocks()
        
        # Mettre à jour les statistiques
        self.update_stats_display()

    def _delete_local_expedition(self, expedition_id):
        """Supprime une expédition en mode local (démonstration)"""
        print(f"🗑️ Suppression locale de l'expédition ID: {expedition_id}")
        
        # Supprimer de la liste principale
        self.expeditions_data = [exp for exp in self.expeditions_data if exp['id'] != expedition_id]
        
        # Mettre à jour l'affichage principal
        self.update_expeditions_display()
        
        # Mettre à jour les blocs centraux (notifications, planning, etc.)
        self.update_central_blocks()
        
        # Mettre à jour les statistiques
        self.update_stats_display()
        
        print(f"✅ Expédition {expedition_id} supprimée de toutes les parties de l'interface")

    def _search_expeditions_db(self, search_term, status_filter, carrier_filter):
        """Rechercher des expéditions dans la base de données avec filtres"""
        try:
            conn = psycopg2.connect(**PG_CONN)
            cursor = conn.cursor()
            
            # Requête corrigée - pas de colonne statut dans bon_expeditions
            query = """
                SELECT be.reference_commande, be.client, be.date_livraison, 
                       be.transporteurs, be.priorite, be.observation, be.liste_articles_livres,
                       c.poids, c.dimension
                FROM sge_cre.bon_expeditions be
                LEFT JOIN sge_cre.colis c ON be.id_colis = c.id_colis
                WHERE 1=1
            """
            params = []
            
            # Filtre par terme de recherche (client ou référence)
            if search_term:
                query += " AND (be.client ILIKE %s OR be.reference_commande ILIKE %s)"
                params.extend([f"%{search_term}%", f"%{search_term}%"])
            
            # Filtre par transporteur
            if carrier_filter and carrier_filter != "Tous":
                query += " AND be.transporteurs = %s"
                params.append(carrier_filter)
            
            cursor.execute(query, params)
            results = cursor.fetchall()
            
            # Convertir en format standard
            expeditions = []
            for row in results:
                # Déterminer le statut basé sur la date de livraison
                delivery_date = row[2]
                today = datetime.now().date()
                
                if delivery_date:
                    if delivery_date < today:
                        status = "delivered"
                    elif delivery_date == today:
                        status = "in-transit"
                    else:
                        status = "preparing"
                else:
                    status = "preparing"
                
                expeditions.append({
                    'number': row[0],  # reference_commande
                    'client': row[1] or "Client non spécifié",
                    'deliveryDate': row[2].strftime('%Y-%m-%d') if row[2] else None,
                    'carrier': row[3] or "Non spécifié",
                    'priority': row[4] or "moyenne",
                    'observation': row[5] or "",
                    'articles': row[6] or "",
                    'status': status,
                    'totalWeight': float(row[7]) if row[7] else 0.0,
                    'packages': 1,
                    'trackingNumber': row[0],
                    'shippingDate': datetime.now().strftime('%Y-%m-%d')
                })
            
            cursor.close()
            conn.close()
            
            return expeditions
            
        except Exception as e:
            print(f"❌ Erreur recherche BD: {e}")
            return []
    
    def _search_expeditions_local(self, search_term, status_filter, carrier_filter):
        """Rechercher des expéditions en mode local avec filtres"""
        filtered_expeditions = []
        
        for exp in self.expeditions_data:
            # Filtre par terme de recherche (numéro ou client)
            if search_term:
                search_upper = search_term.upper()
                if (search_upper not in exp['number'].upper() and 
                    search_upper not in exp['client'].upper()):
                    continue
            
            # Filtre par statut (déterminé par la date de livraison)
            if status_filter and status_filter != "Tous":
                delivery_date = exp.get('deliveryDate')
                if delivery_date:
                    try:
                        delivery_date_obj = datetime.strptime(delivery_date, '%Y-%m-%d').date()
                        today = datetime.now().date()
                        
                        if delivery_date_obj < today:
                            current_status = "delivered"
                        elif delivery_date_obj == today:
                            current_status = "in-transit"
                        else:
                            current_status = "preparing"
                    except:
                        current_status = "preparing"
                else:
                    current_status = "preparing"
                
                status_mapping = {
                    "En préparation": "preparing",
                    "Expédiée": "shipped",
                    "En transit": "in-transit",
                    "Livrée": "delivered"
                }
                if status_filter in status_mapping and current_status != status_mapping[status_filter]:
                    continue
            
            # Filtre par transporteur
            if carrier_filter and carrier_filter != "Tous" and exp['carrier'] != carrier_filter:
                continue
            
            filtered_expeditions.append(exp)
        
        return filtered_expeditions
    
    def _get_status_color(self, status):
        """Obtenir la couleur du badge de statut"""
        status_colors = {
            'preparing': '#f59e0b',      # Orange
            'shipped': '#3b82f6',        # Bleu
            'in-transit': '#8b5cf6',     # Violet
            'delivered': '#10b981',      # Vert
            'cancelled': '#ef4444'       # Rouge
        }
        return status_colors.get(status, '#6b7280')  # Gris par défaut
    
    def _generate_timeline(self, expedition):
        """Générer une timeline dynamique basée sur le statut de l'expédition"""
        timeline = []
        
        # Date de création (toujours présente)
        creation_date = expedition.get('shippingDate', datetime.now().strftime('%Y-%m-%d'))
        timeline.append({
            "icon": "✅",
            "event": "Expédition créée",
            "description": f"L'expédition {expedition['number']} a été créée dans le système",
            "date": f"{creation_date} 08:30"
        })
        
        # Événements basés sur le statut
        status = expedition.get('status', 'preparing')
        
        if status in ['preparing', 'shipped', 'in-transit', 'delivered']:
            timeline.append({
                "icon": "📦",
                "event": "En préparation",
                "description": "Les articles sont préparés et emballés",
                "date": f"{creation_date} 10:15"
            })
        
        if status in ['shipped', 'in-transit', 'delivered']:
            timeline.append({
                "icon": "🚚",
                "event": "Expédiée",
                "description": f"L'expédition a été confiée à {expedition['carrier']}",
                "date": f"{creation_date} 14:20"
            })
        
        if status in ['in-transit', 'delivered']:
            timeline.append({
                "icon": "🔄",
                "event": "En transit",
                "description": "L'expédition est en cours de livraison",
                "date": f"{datetime.now().strftime('%Y-%m-%d')} 09:00"
            })
        
        if status == 'delivered':
            timeline.append({
                "icon": "🎉",
                "event": "Livrée",
                "description": "L'expédition a été livrée avec succès",
                "date": f"{datetime.now().strftime('%Y-%m-%d')} 11:30"
            })
        else:
            # Livraison prévue
            delivery_date = expedition.get('deliveryDate', creation_date)
            timeline.append({
                "icon": "⏳",
                "event": "Livraison prévue",
                "description": f"Livraison prévue le {delivery_date}",
                "date": f"{delivery_date} 11:30"
            })
        
        return timeline


if __name__ == "__main__":
    root = ctk.CTk()
    root.title("Gestion des Expéditions (test autonome)")
    root.geometry("1400x900")
    frame = GestionExpeditionsApp(master=root)
    frame.pack(fill="both", expand=True)
    root.mainloop()
