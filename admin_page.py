import customtkinter as ctk
import tkinter as tk
from sidebar import SidebarFrame
import datetime
from database import list_users, delete_user, update_user, reset_password, deactivate_user

# Paramètres de connexion PostgreSQL
PG_CONN = dict(
    dbname="postgres",
    user="postgres",
    password="postgres123",
    host="localhost",
    port="5432",
    client_encoding="utf8",
    options="-c client_encoding=utf8",
    connect_timeout=10
)

# Gestion sécurisée des imports optionnels
try:
    from tkcalendar import DateEntry
    HAS_TKCALENDAR = True
except ImportError:
    HAS_TKCALENDAR = False
    print("tkcalendar non disponible - fonctionnalités de calendrier désactivées")

try:
    from babel.dates import format_datetime
    HAS_BABEL = True
except ImportError:
    HAS_BABEL = False
    print("babel non disponible - formatage des dates en français désactivé")

try:
    from CTkToolTip import CTkToolTip
    HAS_TOOLTIP = True
except ImportError:
    HAS_TOOLTIP = False
    print("CTkToolTip non disponible - tooltips désactivés")

try:
    import locale
    locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
    HAS_LOCALE = True
except:
    HAS_LOCALE = False
    print("Locale française non disponible")

import tkinter.messagebox as mbox

class AdminFrame(ctk.CTkFrame):
    
    def get_downloads_folder(self):
        """Obtient le dossier Téléchargements du système"""
        import os
        import platform
        
        # Essayer différents chemins selon le système d'exploitation
        system = platform.system()
        
        if system == "Windows":
            # Windows - utiliser le dossier Downloads de l'utilisateur
            downloads_path = os.path.expanduser("~/Downloads")
            if not os.path.exists(downloads_path):
                downloads_path = os.path.expanduser("~/Téléchargements")
        elif system == "Darwin":  # macOS
            downloads_path = os.path.expanduser("~/Downloads")
        else:  # Linux
            downloads_path = os.path.expanduser("~/Downloads")
            if not os.path.exists(downloads_path):
                downloads_path = os.path.expanduser("~/Téléchargements")
        
        # Fallback vers le dossier exports local si le dossier téléchargements n'existe pas
        if not os.path.exists(downloads_path):
            downloads_path = "exports"
        
        return downloads_path

    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        
        # Configuration responsive
        self.responsive_config = {}
        if hasattr(parent, 'get_responsive_config'):
            self.responsive_config = parent.get_responsive_config()
        
        self.create_widgets()
        self.pack(fill="both", expand=True)
        
        # Lier le redimensionnement
        self.bind("<Configure>", self._on_frame_resize)

    def on_window_resize(self, width, height):
        """Méthode appelée quand la fenêtre est redimensionnée"""
        if hasattr(self.parent, 'get_responsive_config'):
            self.responsive_config = self.parent.get_responsive_config()
        self._adapt_layout_to_size(width, height)
    
    def _on_frame_resize(self, event):
        """Gère le redimensionnement du frame"""
        if event.width > 100 and event.height > 100:
            self._adapt_layout_to_size(event.width, event.height)
    
    def _adapt_layout_to_size(self, width, height):
        """Adapte la mise en page selon la taille"""
        # Adapter la table selon la largeur
        if width < 1200:
            # Mode compact - masquer certaines colonnes
            self._configure_table_compact(True)
        else:
            # Mode normal - afficher toutes les colonnes
            self._configure_table_compact(False)
    
    def _configure_table_compact(self, compact):
        """Configure la table en mode compact ou normal"""
        # Cette méthode peut être implémentée pour adapter la table
        pass

    def create_widgets(self):
        user_info = getattr(self.parent, 'user_info', None)
        if user_info is None:
            user_info = {
                'prenom': 'Utilisateur',
                'nom': 'Test',
                'role': 'Admin',
                'email': 'test@example.com',
                'matricule': '12345'
            }
        self.sidebar = SidebarFrame(self, self.parent)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.set_active_button("Admin")
        self.main_content = ctk.CTkFrame(self, fg_color="#f7fafd")
        self.main_content.pack(side="right", fill="both", expand=True)
        self._build_topbar()
        self._build_stats_cards()
        self._build_tabs()
        self._show_tab("Utilisateurs")
        self._update_time()

    def _build_topbar(self):
        topbar = ctk.CTkFrame(self.main_content, fg_color="white", height=70)
        topbar.pack(fill="x", pady=(18, 0), padx=24)
        topbar.grid_columnconfigure(0, weight=1)
        title = ctk.CTkLabel(topbar, text="Administration du Système", font=ctk.CTkFont(size=22, weight="bold"), text_color="#222")
        title.grid(row=0, column=0, sticky="w", pady=(8,0))
        subtitle = ctk.CTkLabel(topbar, text="Gestion des utilisateurs, paramètres et configurations", font=ctk.CTkFont(size=14), text_color="#666")
        subtitle.grid(row=1, column=0, sticky="w")
        self.time_label = ctk.CTkLabel(topbar, text="", font=ctk.CTkFont(size=13), text_color="#666")
        self.time_label.grid(row=0, column=2, rowspan=2, sticky="e", padx=(0,10))
        btn = ctk.CTkButton(topbar, text="👤  Nouvel Utilisateur", fg_color="#3b82f6", hover_color="#2563eb", text_color="white", corner_radius=8, font=ctk.CTkFont(size=14, weight="bold"), width=180, height=36, command=self._open_user_modal)
        btn.grid(row=0, column=1, rowspan=2, sticky="e", padx=(0,20))

    def _update_time(self):
        now = datetime.datetime.now()
        if HAS_LOCALE:
            txt = now.strftime("%A %d %B %Y à %H:%M")
        else:
            txt = now.strftime("%d/%m/%Y %H:%M")
        self.time_label.configure(text=txt)
        self.after(1000, self._update_time)

    def _get_real_time_stats(self):
        """Récupère les statistiques en temps réel depuis la base de données"""
        try:
            import psycopg2
            conn = psycopg2.connect(**PG_CONN)
            cursor = conn.cursor()
            
            # Compter tous les utilisateurs
            cursor.execute("SELECT COUNT(*) FROM sge_cre.individus")
            users_actifs = cursor.fetchone()[0]
            
            # Compter les administrateurs
            cursor.execute("SELECT COUNT(*) FROM sge_cre.individus WHERE role LIKE '%informatique%'")
            admins = cursor.fetchone()[0]
            
            # Compter les responsables
            cursor.execute("SELECT COUNT(*) FROM sge_cre.individus WHERE role LIKE '%Responsable%'")
            responsables = cursor.fetchone()[0]
            
            # Compter les utilisateurs limités (magasiniers, emballeurs, livreurs, etc.)
            cursor.execute("""
                SELECT COUNT(*) FROM sge_cre.individus 
                WHERE role IN ('Magasinier', 'Emballeur', 'Livreur', 'Agent de logistique', 'Garde de sécurité')
            """)
            users_limites = cursor.fetchone()[0]
            
            # Calculer les nouveaux utilisateurs ce mois
            cursor.execute("""
                SELECT COUNT(*) FROM sge_cre.individus 
                WHERE date_creation >= CURRENT_DATE - INTERVAL '30 days'
            """)
            nouveaux_mois = cursor.fetchone()[0]
            
            conn.close()
            
            return {
                'users_actifs': users_actifs,
                'admins': admins,
                'responsables': responsables,
                'users_limites': users_limites,
                'nouveaux_mois': nouveaux_mois
            }
        except Exception as e:
            print(f"Erreur lors de la récupération des statistiques: {e}")
            # Retourner des valeurs par défaut en cas d'erreur
            return {
                'users_actifs': 0,
                'admins': 0,
                'responsables': 0,
                'users_limites': 0,
                'nouveaux_mois': 0
            }

    def _build_stats_cards(self):
        stats_frame = ctk.CTkFrame(self.main_content, fg_color="#f7fafd")
        stats_frame.pack(fill="x", pady=(10, 8), padx=24)
        
        # Récupérer les données en temps réel
        stats = self._get_real_time_stats()
        
        cards_data = [
            ("Utilisateurs Actifs", str(stats['users_actifs']), "#10b981", "✅", f"+{stats['nouveaux_mois']} ce mois", "#10b981"),
            ("Administrateurs", str(stats['admins']), "#DC2626", "🛡️", "Accès complet", "#DC2626"),
            ("Responsables", str(stats['responsables']), "#2563EB", "👔", "Accès étendu", "#2563EB"),
            ("Utilisateurs Limités", str(stats['users_limites']), "#D97706", "👷", "Accès limité", "#D97706")
        ]
        
        # Stocker les références des cartes pour pouvoir les mettre à jour
        self.stats_cards = {}
        
        for i, (title, value, color, icon, sub, subcolor) in enumerate(cards_data):
            card = ctk.CTkFrame(stats_frame, fg_color="white", corner_radius=16, border_width=1, border_color="#e0e0e0", width=220, height=110)
            card.grid(row=0, column=i, padx=10, sticky="nsew")
            
            # Créer les labels avec des références pour pouvoir les mettre à jour
            icon_label = ctk.CTkLabel(card, text=icon, font=ctk.CTkFont(size=28), text_color=color)
            icon_label.place(x=12, y=10)
            
            title_label = ctk.CTkLabel(card, text=title, font=ctk.CTkFont(size=13), text_color="#666")
            title_label.place(x=54, y=12)
            
            value_label = ctk.CTkLabel(card, text=value, font=ctk.CTkFont(size=28, weight="bold"), text_color="#222")
            value_label.place(x=54, y=36)
            
            sub_label = ctk.CTkLabel(card, text=sub, font=ctk.CTkFont(size=12), text_color=subcolor)
            sub_label.place(x=54, y=72)
            
            # Stocker les références pour mise à jour
            self.stats_cards[title] = {
                'value_label': value_label,
                'sub_label': sub_label
            }
        
        # Programmer la mise à jour automatique toutes les 30 secondes
        self._schedule_stats_update()

    def _schedule_stats_update(self):
        """Programme la mise à jour automatique des statistiques"""
        def update_stats():
            try:
                stats = self._get_real_time_stats()
                
                # Mettre à jour les cartes
                if hasattr(self, 'stats_cards'):
                    # Utilisateurs Actifs
                    if 'Utilisateurs Actifs' in self.stats_cards:
                        self.stats_cards['Utilisateurs Actifs']['value_label'].configure(text=str(stats['users_actifs']))
                        self.stats_cards['Utilisateurs Actifs']['sub_label'].configure(text=f"+{stats['nouveaux_mois']} ce mois")
                    
                    # Administrateurs
                    if 'Administrateurs' in self.stats_cards:
                        self.stats_cards['Administrateurs']['value_label'].configure(text=str(stats['admins']))
                    
                    # Responsables
                    if 'Responsables' in self.stats_cards:
                        self.stats_cards['Responsables']['value_label'].configure(text=str(stats['responsables']))
                    
                    # Utilisateurs Limités
                    if 'Utilisateurs Limités' in self.stats_cards:
                        self.stats_cards['Utilisateurs Limités']['value_label'].configure(text=str(stats['users_limites']))
                
                # Programmer la prochaine mise à jour
                self.after(30000, update_stats)  # 30 secondes
                
            except Exception as e:
                print(f"Erreur lors de la mise à jour des statistiques: {e}")
                # Réessayer dans 30 secondes même en cas d'erreur
                self.after(30000, update_stats)
        
        # Démarrer la première mise à jour
        self.after(30000, update_stats)

    def _update_stats_immediately(self):
        """Met à jour immédiatement les statistiques"""
        try:
            stats = self._get_real_time_stats()
            
            # Mettre à jour les cartes
            if hasattr(self, 'stats_cards'):
                # Utilisateurs Actifs
                if 'Utilisateurs Actifs' in self.stats_cards:
                    self.stats_cards['Utilisateurs Actifs']['value_label'].configure(text=str(stats['users_actifs']))
                    self.stats_cards['Utilisateurs Actifs']['sub_label'].configure(text=f"+{stats['nouveaux_mois']} ce mois")
                
                # Administrateurs
                if 'Administrateurs' in self.stats_cards:
                    self.stats_cards['Administrateurs']['value_label'].configure(text=str(stats['admins']))
                
                # Responsables
                if 'Responsables' in self.stats_cards:
                    self.stats_cards['Responsables']['value_label'].configure(text=str(stats['responsables']))
                
                # Utilisateurs Limités
                if 'Utilisateurs Limités' in self.stats_cards:
                    self.stats_cards['Utilisateurs Limités']['value_label'].configure(text=str(stats['users_limites']))
                    
        except Exception as e:
            print(f"Erreur lors de la mise à jour immédiate des statistiques: {e}")

    def _build_tabs(self):
        tab_frame = ctk.CTkFrame(self.main_content, fg_color="white")
        tab_frame.pack(fill="x", pady=(10, 0), padx=24)
        self.tabs = {}
        tab_labels = [
            ("Utilisateurs", "👥"),
            ("Rôles & Permissions", "🔖"),
            ("Matricules", "🆔"),
            ("Paramètres Système", "⚙️"),
            ("Journal d'Activité", "📝"),
            ("CLI", "💻")
        ]
        for tab, icon in tab_labels:
            b = ctk.CTkButton(tab_frame, text=f"{icon} {tab}", fg_color="#f7fafd", hover_color="#e3e8ee", text_color="#222", corner_radius=6, font=ctk.CTkFont(size=13, weight="bold"), width=190, height=32, command=lambda t=tab: self._show_tab(t))
            b.pack(side="left", padx=(0, 8))
            self.tabs[tab] = b
        self.tab_contents = {}
        for tab, _ in tab_labels:
            frame = ctk.CTkFrame(self.main_content, fg_color="white")
            frame.pack(fill="both", expand=True, padx=24, pady=(0, 24))
            frame.pack_forget()
            self.tab_contents[tab] = frame
        self._build_users_tab(self.tab_contents["Utilisateurs"])
        self._build_roles_tab(self.tab_contents["Rôles & Permissions"])
        self._build_matricules_tab(self.tab_contents["Matricules"])
        self._build_settings_tab(self.tab_contents["Paramètres Système"])
        self._build_audit_tab(self.tab_contents["Journal d'Activité"])
        self._build_cli_tab(self.tab_contents["CLI"])

    def _show_tab(self, tab):
        for t, btn in self.tabs.items():
            if t == tab:
                btn.configure(fg_color="#3b82f6", text_color="white")
            else:
                btn.configure(fg_color="#f7fafd", text_color="#222")
        for t, frame in self.tab_contents.items():
            if t == tab:
                frame.pack(fill="both", expand=True, padx=24, pady=(0, 24))
            else:
                frame.pack_forget()

    def _build_users_tab(self, parent):
        # Barre recherche et titre
        top = ctk.CTkFrame(parent, fg_color="white")
        top.pack(fill="x", pady=(10, 0))
        ctk.CTkLabel(top, text="Liste des Utilisateurs", font=ctk.CTkFont(size=16, weight="bold"), text_color="#222").pack(side="left", padx=8)
        # Bouton moderne d'ajout d'utilisateur
        add_btn = ctk.CTkButton(top, text="➕ Nouvel utilisateur", fg_color="#3b82f6", hover_color="#2563eb", text_color="#fff", corner_radius=8, font=ctk.CTkFont(size=14, weight="bold"), height=38, command=self._open_user_modal)
        add_btn.pack(side="right", padx=(8, 0))
        search = ctk.CTkEntry(top, placeholder_text="Rechercher un utilisateur...", width=260)
        search.pack(side="right", padx=8)
        search.bind('<KeyRelease>', lambda e: self._refresh_users(table, search.get()))
        # Tableau utilisateurs
        table = ctk.CTkFrame(parent, fg_color="white", border_width=1, border_color="#e0e0e0")
        table.pack(fill="both", expand=True, pady=(8, 0))
        
        # Stocker la référence de la table pour l'actualisation automatique
        self.users_table = table
        
        self._refresh_users(table, "")
        # Pagination fictive
        pag = ctk.CTkFrame(parent, fg_color="white")
        pag.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(pag, text="Affichage de 1 à 10 sur 24 utilisateurs", font=ctk.CTkFont(size=12), text_color="#666").pack(side="left", padx=8)
        for txt, color in [("Précédent", "#666"), ("1", "#fff"), ("2", "#666"), ("Suivant", "#666")]:
            ctk.CTkButton(pag, text=txt, fg_color="#3b82f6" if txt=="1" else "#f7fafd", text_color="#222" if txt!="1" else "#fff", width=38, height=28, corner_radius=6).pack(side="left", padx=2)

    def _refresh_users(self, table, search_term):
        for w in table.winfo_children():
            w.destroy()
        header = ctk.CTkFrame(table, fg_color="white")
        header.pack(fill="x")
        for col in ["Nom", "Email", "Rôle", "Dernière Connexion", "Statut", "Actions"]:
            ctk.CTkLabel(header, text=col, font=ctk.CTkFont(size=13, weight="bold"), text_color="#666", width=160, anchor="w").pack(side="left", padx=2)
        users = list_users()
        for user in users:
            uid, nom, prenom, email, role, matricule, actif, last_login = user
            if search_term and search_term.lower() not in (nom+prenom+email+role+matricule).lower():
                continue
            # Format date en français
            try:
                dt = datetime.datetime.strptime(last_login, "%d/%m/%Y %H:%M")
                if HAS_BABEL:
                    last_login_fr = format_datetime(dt, "d MMMM yyyy HH:mm", locale="fr_FR")
                else:
                    last_login_fr = dt.strftime("%d/%m/%Y %H:%M")
            except Exception:
                last_login_fr = last_login
            row = ctk.CTkFrame(table, fg_color="#f7fafd")
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=f"👤 {prenom} {nom}\n{matricule}", font=ctk.CTkFont(size=13, weight="bold"), text_color="#222", width=160, anchor="w").pack(side="left", padx=2)
            ctk.CTkLabel(row, text=email, font=ctk.CTkFont(size=13), text_color="#222", width=160, anchor="w").pack(side="left", padx=2)
            ctk.CTkLabel(row, text=role, font=ctk.CTkFont(size=13), text_color="#7c3aed" if role=="Super Administrateur" else "#2563eb" if role=="Administrateur" else "#059669" if role=="Gestionnaire Entrepôt" else "#ef4444", width=160, anchor="w").pack(side="left", padx=2)
            ctk.CTkLabel(row, text=last_login_fr, font=ctk.CTkFont(size=13), text_color="#222", width=160, anchor="w").pack(side="left", padx=2)
            ctk.CTkLabel(row, text="🟢 Actif" if actif else "🔴 Inactif", font=ctk.CTkFont(size=13), text_color="#10b981" if actif else "#ef4444", width=160, anchor="w").pack(side="left", padx=2)
            actions = ctk.CTkFrame(row, fg_color="#f7fafd")
            actions.pack(side="left", padx=2)
            ctk.CTkButton(actions, text="✏️", width=32, height=32, fg_color="#3b82f6", text_color="#fff", corner_radius=8, command=lambda u=user: self._edit_user(u, table, search_term)).pack(side="left", padx=2)
            ctk.CTkButton(actions, text="🔑", width=32, height=32, fg_color="#10b981", text_color="#fff", corner_radius=8, command=lambda u=user: self._reset_user_pw(u, table, search_term)).pack(side="left", padx=2)
            ctk.CTkButton(actions, text="🚫", width=32, height=32, fg_color="#ef4444", text_color="#fff", corner_radius=8, command=lambda u=user: self._confirm_delete_user(u, table, search_term)).pack(side="left", padx=2)

    def _edit_user(self, user, table, search_term):
        uid, nom, prenom, email, role, matricule, actif, last_login = user
        modal = ctk.CTkToplevel(self)
        modal.title("Éditer l'utilisateur")
        modal.geometry("500x650")
        modal.grab_set()
        modal.resizable(False, False)
        modal.configure(fg_color="#f0f6ff")  # Fond doux bleu clair

        # Barre de titre colorée avec icône
        title_bar = ctk.CTkFrame(modal, fg_color="#2563eb", corner_radius=18)
        title_bar.pack(fill="x", pady=(0, 0))
        ctk.CTkLabel(
            title_bar, text="✏️ Modifier un utilisateur",
            font=ctk.CTkFont(size=23, weight="bold"),
            text_color="#fff"
        ).pack(pady=18)

        # Illustration ou icône décorative
        ctk.CTkLabel(
            modal, text="📝", font=ctk.CTkFont(size=48), text_color="#2563eb"
        ).pack(pady=(10, 0))

        form = ctk.CTkScrollableFrame(modal, fg_color="#f7fafd", corner_radius=22, width=460, height=470)
        form.pack(fill="both", expand=True, padx=22, pady=8)
        entries = {}
        errors = {}
        champs = [
            ("Nom *", nom),
            ("Prénom *", prenom),
            ("Email *", email),
            ("Rôle *", role),
            ("Matricule *", matricule),
            ("Adresse *", ""),
            ("Téléphone", "")
        ]
        role_values = [
            "Le responsable des stocks",
            "Le magasinier",
            "Les emballeurs",
            "Le responsable de la logistique",
            "Les agents de logistique",
            "Les livreurs",
            "Le responsable informatique",
            "Les techniciens informatiques",
            "Le responsable de la sécurité physique",
            "Les gardes de sécurité"
        ]
        for label, val in champs:
            row = ctk.CTkFrame(form, fg_color="transparent")
            row.pack(fill="x", pady=(14,0))
            # Astérisque rouge pour les champs obligatoires
            if "*" in label:
                label_widget = ctk.CTkLabel(row, text=label, font=ctk.CTkFont(size=16, weight="bold"), text_color="#374151")
                label_widget.pack(side="left", padx=(0,2))
                ctk.CTkLabel(row, text="*", font=ctk.CTkFont(size=16, weight="bold"), text_color="#ef4444").pack(side="left")
            else:
                ctk.CTkLabel(row, text=label, font=ctk.CTkFont(size=16, weight="bold"), text_color="#374151").pack(side="left", padx=(0,10))
            if label == "Rôle *":
                e = ctk.CTkComboBox(row, values=role_values, width=270, font=ctk.CTkFont(size=15))
                e.set(val)
            else:
                e = ctk.CTkEntry(row, height=40, font=ctk.CTkFont(size=15))
                e.insert(0, val)
            e.pack(side="right", fill="x", expand=True)
            entries[label] = e
            err = ctk.CTkLabel(form, text="", text_color="#ef4444", font=ctk.CTkFont(size=12))
            err.pack(anchor="w", padx=10)
            errors[label] = err
        # Gestion dynamique des erreurs : efface l'erreur dès que l'utilisateur modifie le champ
        def clear_error(event, label):
            errors[label].configure(text="")
            entries[label].configure(border_color="#D1D5DB")
        for label in entries:
            entries[label].bind('<Key>', lambda e, l=label: clear_error(e, l))
            if label == "Rôle *":
                entries[label].bind("<<ComboboxSelected>>", lambda e, l=label: clear_error(e, l))
        btns = ctk.CTkFrame(modal, fg_color="transparent")
        btns.pack(fill="x", pady=18, padx=20)
        ctk.CTkButton(
            btns, text="❌ Annuler", fg_color="#f87171", hover_color="#ef4444", text_color="#fff",
            corner_radius=10, height=44, font=ctk.CTkFont(size=16, weight="bold"),
            command=modal.destroy
        ).pack(side="left", padx=12, expand=True, fill="x")
        def save():
            # --- ÉDITION UTILISATEUR ---
            if 'uid' in locals() or 'uid' in globals():
                nom = entries["Nom *"].get()
                prenom = entries["Prénom *"].get()
                email = entries["Email *"].get()
                role = entries["Rôle *"].get()
                matricule = entries["Matricule *"].get()
                valid = True
                for err in errors.values(): err.configure(text="")
                for field, entry in entries.items(): entry.configure(border_color="#D1D5DB")
                # Validation des champs obligatoires
                if not nom:
                    errors["Nom *"].configure(text="Champ obligatoire")
                    entries["Nom *"].configure(border_color="#ef4444")
                    valid = False
                if not prenom:
                    errors["Prénom *"].configure(text="Champ obligatoire")
                    entries["Prénom *"].configure(border_color="#ef4444")
                    valid = False
                if not email or "@" not in email:
                    errors["Email *"].configure(text="Email invalide")
                    entries["Email *"].configure(border_color="#ef4444")
                    valid = False
                if not matricule:
                    errors["Matricule *"].configure(text="Champ obligatoire")
                    entries["Matricule *"].configure(border_color="#ef4444")
                    valid = False
                if not role:
                    errors["Rôle *"].configure(text="Champ obligatoire")
                    entries["Rôle *"].configure(border_color="#ef4444")
                    valid = False
                if not valid:
                    return
                try:
                    from database import update_user
                    update_user(uid, nom, prenom, email, role, "", "", matricule)
                    modal.destroy()
                    self._refresh_users(table, search_term)
                except Exception as e:
                    errors["Email *"].configure(text=f"Erreur : {e}")
                    entries["Email *"].configure(border_color="#ef4444")
            # --- CRÉATION UTILISATEUR ---
            else:
                nom = entries["Nom *"].get()
                prenom = entries["Prénom *"].get()
                email = entries["Email *"].get()
                matricule = entries["Matricule *"].get()
                role = entries["Rôle *"].get()
                mdp = entries["Mot de passe *"].get()
                conf = entries["Confirmation *"].get()
                adresse = entries["Adresse *"].get()
                telephone = entries["Téléphone"].get()
                valid = True
                for err in errors.values(): err.configure(text="")
                for field, entry in entries.items(): entry.configure(border_color="#D1D5DB")
                # Validation des champs obligatoires
                if not nom:
                    errors["Nom *"].configure(text="Champ obligatoire")
                    entries["Nom *"].configure(border_color="#ef4444")
                    valid = False
                if not prenom:
                    errors["Prénom *"].configure(text="Champ obligatoire")
                    entries["Prénom *"].configure(border_color="#ef4444")
                    valid = False
                if not email or "@" not in email:
                    errors["Email *"].configure(text="Email invalide")
                    entries["Email *"].configure(border_color="#ef4444")
                    valid = False
                if not matricule:
                    errors["Matricule *"].configure(text="Champ obligatoire")
                    entries["Matricule *"].configure(border_color="#ef4444")
                    valid = False
                if not mdp or not conf:
                    errors["Mot de passe *"].configure(text="Champ obligatoire")
                    errors["Confirmation *"].configure(text="Champ obligatoire")
                    entries["Mot de passe *"].configure(border_color="#ef4444")
                    entries["Confirmation *"].configure(border_color="#ef4444")
                    valid = False
                if mdp != conf:
                    errors["Confirmation *"].configure(text="Les mots de passe ne correspondent pas")
                    entries["Confirmation *"].configure(border_color="#ef4444")
                    valid = False
                if not role:
                    errors["Rôle *"].configure(text="Champ obligatoire")
                    entries["Rôle *"].configure(border_color="#ef4444")
                    valid = False
                if not adresse:
                    errors["Adresse *"].configure(text="Champ obligatoire")
                    entries["Adresse *"].configure(border_color="#ef4444")
                    valid = False
                # Validation et unicité du matricule
                if matricule:
                    is_valid, validation_msg = MatriculeManager.validate_matricule(matricule)
                    if not is_valid:
                        errors["Matricule *"].configure(text=validation_msg)
                        entries["Matricule *"].configure(border_color="#ef4444")
                        valid = False
                    else:
                        is_available, availability_msg = MatriculeManager.is_matricule_available(matricule)
                        if not is_available:
                            errors["Matricule *"].configure(text=availability_msg)
                            entries["Matricule *"].configure(border_color="#ef4444")
                            valid = False
                        expected_role = MatriculeManager.get_role_from_matricule(matricule)
                        if expected_role and expected_role != role:
                            errors["Matricule *"].configure(text=f"Ce matricule correspond au rôle '{expected_role}'")
                            entries["Matricule *"].configure(border_color="#ef4444")
                            valid = False
                    if not valid:
                        return
                try:
                    import psycopg2, datetime
                    conn = psycopg2.connect(**PG_CONN)
                    cursor = conn.cursor()
                    cursor.execute("""
                        INSERT INTO sge_cre.individus (nom, prenom, email, password, role, matricule, adresse, telephone)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        nom,
                        prenom,
                        email,
                        mdp,  # mot de passe en clair (à hasher en production)
                        role,
                        matricule,
                        adresse,
                        telephone
                    ))
                    conn.commit()
                    conn.close()
                    # Popup de succès
                    success_modal = ctk.CTkToplevel(modal)
                    success_modal.title("🎉 Succès")
                    success_modal.geometry("480x420")
                    success_modal.grab_set()
                    success_modal.resizable(False, False)
                    success_modal.configure(fg_color="#ffffff")
                    # Centrer la modale
                    success_modal.update_idletasks()
                    x = (success_modal.winfo_screenwidth() // 2) - (480 // 2)
                    y = (success_modal.winfo_screenheight() // 2) - (420 // 2)
                    success_modal.geometry(f"480x420+{x}+{y}")
                    # Conteneur principal
                    main_container = ctk.CTkFrame(success_modal, fg_color="#ffffff", corner_radius=24, border_width=3, border_color="#e5e7eb")
                    main_container.pack(fill="both", expand=True, padx=25, pady=25)
                    # Animation de succès
                    success_circle = ctk.CTkFrame(main_container, fg_color="#dcfce7", corner_radius=60, width=120, height=120, border_width=3, border_color="#bbf7d0")
                    success_circle.pack(pady=(35, 25))
                    ctk.CTkLabel(success_circle, text="🎉", font=ctk.CTkFont(size=52), text_color="#16a34a").pack(expand=True)
                    ctk.CTkLabel(main_container, text="🎊 Utilisateur Créé avec Succès !", font=ctk.CTkFont(size=26, weight="bold"), text_color="#1f2937").pack(pady=(0, 18))
                    ctk.CTkLabel(main_container, text=f"✨ L'utilisateur {prenom} {nom} a été ajouté avec succès au système de gestion d'entrepôts.", font=ctk.CTkFont(size=15), text_color="#6b7280", wraplength=400).pack(pady=(0, 28))
                    info_frame = ctk.CTkFrame(main_container, fg_color="#f8fafc", corner_radius=16, border_width=2, border_color="#e2e8f0")
                    info_frame.pack(fill="x", padx=25, pady=(0, 28))
                    matricule_container = ctk.CTkFrame(info_frame, fg_color="transparent")
                    matricule_container.pack(fill="x", padx=18, pady=(15, 8))
                    ctk.CTkLabel(matricule_container, text="🆔 Matricule", font=ctk.CTkFont(size=15, weight="bold"), text_color="#374151").pack(side="left")
                    ctk.CTkLabel(matricule_container, text=matricule, font=ctk.CTkFont(size=20, weight="bold"), text_color="#059669").pack(side="right")
                    role_container = ctk.CTkFrame(info_frame, fg_color="transparent")
                    role_container.pack(fill="x", padx=18, pady=(8, 15))
                    ctk.CTkLabel(role_container, text="👤 Rôle", font=ctk.CTkFont(size=15, weight="bold"), text_color="#374151").pack(side="left")
                    ctk.CTkLabel(role_container, text=role, font=ctk.CTkFont(size=17), text_color="#6b7280").pack(side="right")
                    ctk.CTkButton(main_container, text="🎯 Parfait ! Continuer", fg_color="#10b981", hover_color="#059669", text_color="white", corner_radius=16, height=52, font=ctk.CTkFont(size=17, weight="bold"), border_width=2, border_color="#34d399", command=lambda: [success_modal.destroy(), modal.destroy()]).pack(pady=(0, 25), padx=25, fill="x")
                    success_modal.after(5000, lambda: [success_modal.destroy(), modal.destroy()])
                except Exception as e:
                    errors["Email *"].configure(text=f"Erreur : {e}")
                    entries["Email *"].configure(border_color="#ef4444")
                if hasattr(self, 'tab_contents') and "Utilisateurs" in self.tab_contents:
                    for w in self.tab_contents["Utilisateurs"].winfo_children():
                        w.destroy()
                    self._build_users_tab(self.tab_contents["Utilisateurs"])
                    self._update_stats_immediately()
                    self._log_activity("Création", f"Création de l'utilisateur {prenom} {nom}", f"{prenom} {nom}", f"Matricule: {matricule}, Rôle: {role}")
                
    def _confirm_delete_user(self, user, table, search_term):
        uid = user[0]
        modal = ctk.CTkToplevel(self)
        modal.title("Confirmation")
        modal.geometry("400x200")
        modal.grab_set()
        modal.resizable(False, False)
        ctk.CTkLabel(modal, text="🚫 Désactiver/Supprimer", font=ctk.CTkFont(size=18, weight="bold"), text_color="#ef4444").pack(pady=(22,8))
        ctk.CTkLabel(modal, text="Voulez-vous vraiment désactiver ou supprimer cet utilisateur ?", font=ctk.CTkFont(size=14), text_color="#d32f2f", wraplength=340, justify="center").pack(pady=(0,12), padx=20)
        btns = ctk.CTkFrame(modal, fg_color="white")
        btns.pack(pady=(0,18))
        ctk.CTkButton(btns, text="Oui, supprimer", fg_color="#ef4444", hover_color="#b91c1c", text_color="#fff", corner_radius=8, height=38, command=lambda: self._delete_user(uid, modal, table, search_term)).pack(side="left", padx=12)
        ctk.CTkButton(btns, text="Non, annuler", fg_color="#90A4AE", hover_color="#78909C", text_color="#fff", corner_radius=8, height=38, command=modal.destroy).pack(side="left", padx=12)

    def _delete_user(self, uid, modal, table, search_term):
        try:
            delete_user(uid)
            modal.destroy()
            self._refresh_users(table, search_term)
        except Exception as e:
            print(f"Erreur lors de la suppression de l'utilisateur : {e}")

    def _build_roles_tab(self, parent):
        ctk.CTkLabel(parent, text="Gestion des Rôles et Permissions", font=ctk.CTkFont(size=19, weight="bold"), text_color="#2563eb").pack(pady=(18, 2))
        ctk.CTkLabel(parent, text="Visualisez, recherchez et comprenez les droits de chaque rôle du système.", font=ctk.CTkFont(size=13), text_color="#666").pack(pady=(0, 10))
        topbar = ctk.CTkFrame(parent, fg_color="white")
        topbar.pack(fill="x", padx=30, pady=(0, 8))
        search_var = tk.StringVar()
        search_entry = ctk.CTkEntry(topbar, placeholder_text="Rechercher un rôle...", textvariable=search_var, width=220)
        search_entry.pack(side="left", padx=(0, 12))
        add_btn = ctk.CTkButton(topbar, text="➕ Ajouter un rôle", fg_color="#3b82f6", text_color="#fff", corner_radius=8, height=36)
        add_btn.pack(side="right")
        # Table responsive
        table_frame = ctk.CTkFrame(parent, fg_color="#f7fafd", corner_radius=14)
        table_frame.pack(fill="both", expand=True, padx=30, pady=10)
        def render_table():
            for w in table_frame.winfo_children():
                w.destroy()
            header = ctk.CTkFrame(table_frame, fg_color="white")
            header.pack(fill="x")
            for col, w in zip(["Rôle", "Description", "Permissions"], [140, 260, 420]):
                ctk.CTkLabel(header, text=col, font=ctk.CTkFont(size=14, weight="bold"), text_color="#666", width=w, anchor="w").pack(side="left", padx=2)
            roles = [
                ("Super Administrateur", "Accès complet à toutes les fonctionnalités du système", [
                    ("Gestion utilisateurs", "Créer, éditer, supprimer tous les comptes"),
                    ("Gestion rôles", "Définir et modifier les rôles et permissions"),
                    ("Paramètres système", "Configuration complète du système"),
                    ("Sauvegardes", "Gestion des sauvegardes et restauration"),
                    ("Audit complet", "Accès à tous les logs d'activité")]),
                ("Administrateur", "Gestion complète des utilisateurs et paramètres système", [
                    ("Gestion utilisateurs", "Créer, éditer, supprimer des comptes"),
                    ("Gestion rôles", "Définir les rôles et permissions"),
                    ("Paramètres système", "Configuration du système"),
                    ("Audit", "Consulter les logs d'activité")]),
                ("Gestionnaire Entrepôt", "Gestion des stocks, réceptions, expéditions et rapports", [
                    ("Gestion stocks", "Créer, modifier, supprimer des produits"),
                    ("Réceptions", "Gérer les bons de réception"),
                    ("Expéditions", "Gérer les bons d'expédition"),
                    ("Emballages", "Gérer les matériaux d'emballage"),
                    ("Entrepôts", "Gérer les zones et cellules"),
                    ("Rapports", "Créer et exporter des rapports")]),
                ("Responsable Réception", "Gestion des réceptions et contrôle qualité", [
                    ("Réceptions", "Créer et valider les réceptions"),
                    ("Contrôle qualité", "Vérifier les marchandises reçues"),
                    ("Stocks", "Consulter les niveaux de stock"),
                    ("Emballages", "Consulter les emballages"),
                    ("Rapports", "Consulter les rapports")]),
                ("Opérateur Stock", "Opérations de stockage et inventaire", [
                    ("Stocks", "Consulter et modifier les stocks"),
                    ("Réceptions", "Enregistrer les réceptions"),
                    ("Entrepôts", "Consulter les zones de stockage")]),
                ("Expéditeur", "Gestion des expéditions et livraisons", [
                    ("Expéditions", "Créer et valider les expéditions"),
                    ("Livraisons", "Suivre les livraisons"),
                    ("Emballages", "Gérer les emballages"),
                    ("Rapports", "Consulter les rapports")]),
                ("Consultant", "Consultation des rapports et statistiques", [
                    ("Dashboard", "Consulter le tableau de bord"),
                    ("Stocks", "Consulter les stocks"),
                    ("Réceptions", "Consulter les réceptions"),
                    ("Expéditions", "Consulter les expéditions"),
                    ("Rapports", "Consulter et exporter les rapports")]),
                ("Stagiaire", "Accès limité en lecture seule", [
                    ("Dashboard", "Consulter le tableau de bord"),
                    ("Stocks", "Consulter les stocks"),
                    ("Réceptions", "Consulter les réceptions"),
                    ("Expéditions", "Consulter les expéditions")])
            ]
            filtered = [r for r in roles if search_var.get().lower() in r[0].lower() or search_var.get().lower() in r[1].lower() or any(search_var.get().lower() in p[0].lower() for p in r[2])]
            if not filtered:
                ctk.CTkLabel(table_frame, text="Aucun rôle ne correspond à votre recherche.", font=ctk.CTkFont(size=13), text_color="#ef4444").pack(pady=30)
                return
            for role, desc, perms in filtered:
                row = ctk.CTkFrame(table_frame, fg_color="#fff", corner_radius=10)
                row.pack(fill="x", pady=4, padx=2)
                # Couleurs selon le niveau d'accès
                if role == "Super Administrateur":
                    badge_color = "#DC2626"
                elif role == "Administrateur":
                    badge_color = "#7C3AED"
                elif role == "Gestionnaire Entrepôt":
                    badge_color = "#2563EB"
                elif role == "Responsable Réception":
                    badge_color = "#059669"
                elif role == "Opérateur Stock":
                    badge_color = "#D97706"
                elif role == "Expéditeur":
                    badge_color = "#F59E0B"
                elif role == "Consultant":
                    badge_color = "#6B7280"
                else:
                    badge_color = "#9CA3AF"
                badge = ctk.CTkLabel(row, text=role, font=ctk.CTkFont(size=13, weight="bold"), text_color="#fff", fg_color=badge_color, corner_radius=8, width=120, anchor="center")
                badge.pack(side="left", padx=8, pady=8)
                ctk.CTkLabel(row, text=desc, font=ctk.CTkFont(size=13), text_color="#222", width=260, anchor="w", wraplength=250, justify="left").pack(side="left", padx=2)
                perms_frame = ctk.CTkFrame(row, fg_color="#fff")
                perms_frame.pack(side="left", padx=2)
                for p, tip in perms:
                    lbl = ctk.CTkLabel(perms_frame, text=f"✔️ {p}", font=ctk.CTkFont(size=12), text_color="#2563eb", anchor="w")
                    lbl.pack(anchor="w", pady=1)
                    if HAS_TOOLTIP:
                        CTkToolTip(lbl, message=tip)
        render_table()
        search_entry.bind('<KeyRelease>', lambda e: render_table())

    def _build_matricules_tab(self, parent):
        """Construit l'onglet de gestion des matricules"""
        from matricule_manager import MatriculeManager
        
        # Stocker la référence du parent pour les mises à jour
        self.matricules_parent = parent
        
        # Titre et description
        ctk.CTkLabel(parent, text="Gestion des Matricules", font=ctk.CTkFont(size=19, weight="bold"), text_color="#2563eb").pack(pady=(18, 2))
        ctk.CTkLabel(parent, text="Système de génération automatique des matricules basé sur les rôles", font=ctk.CTkFont(size=13), text_color="#666").pack(pady=(0, 10))
        
        # Barre d'outils
        topbar = ctk.CTkFrame(parent, fg_color="white")
        topbar.pack(fill="x", padx=30, pady=(0, 8))
        
        # Bouton de rafraîchissement
        refresh_btn = ctk.CTkButton(topbar, text="🔄 Rafraîchir", fg_color="#10b981", text_color="#fff", corner_radius=8, height=36, 
                                   command=lambda: self._refresh_matricules_tab(parent))
        refresh_btn.pack(side="right", padx=(8, 0))
        
        # Bouton de test
        test_btn = ctk.CTkButton(topbar, text="🧪 Tester", fg_color="#f59e0b", text_color="#fff", corner_radius=8, height=36,
                                command=self._test_matricule_generation)
        test_btn.pack(side="right", padx=8)
        
        # Conteneur principal avec scroll
        main_container = ctk.CTkScrollableFrame(parent, fg_color="white", corner_radius=14)
        main_container.pack(fill="both", expand=True, padx=30, pady=10)
        
        # Section 1: Mapping des rôles vers préfixes
        mapping_frame = ctk.CTkFrame(main_container, fg_color="#f7fafd", corner_radius=12)
        mapping_frame.pack(fill="x", pady=(0, 20))
        
        ctk.CTkLabel(mapping_frame, text="📋 Mapping des Rôles vers Préfixes", 
                    font=ctk.CTkFont(size=16, weight="bold"), text_color="#2563eb").pack(pady=(15, 10))
        
        # Grille des rôles et préfixes
        roles_grid = ctk.CTkFrame(mapping_frame, fg_color="transparent")
        roles_grid.pack(fill="x", padx=20, pady=(0, 15))
        
        # En-têtes
        headers = ctk.CTkFrame(roles_grid, fg_color="transparent")
        headers.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(headers, text="Rôle", font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151", width=200, anchor="w").pack(side="left")
        ctk.CTkLabel(headers, text="Préfixe", font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151", width=80, anchor="w").pack(side="left")
        ctk.CTkLabel(headers, text="Exemple", font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151", width=100, anchor="w").pack(side="left")
        
        # Afficher les rôles et leurs préfixes
        for role_name, prefix in MatriculeManager.ROLE_PREFIXES.items():
            if role_name == "default":
                continue
                
            row = ctk.CTkFrame(roles_grid, fg_color="white", corner_radius=8)
            row.pack(fill="x", pady=2)
            
            ctk.CTkLabel(row, text=role_name, font=ctk.CTkFont(size=13), text_color="#374151", width=200, anchor="w").pack(side="left", padx=10, pady=8)
            
            # Badge pour le préfixe
            prefix_badge = ctk.CTkLabel(row, text=prefix, font=ctk.CTkFont(size=12, weight="bold"), 
                                      text_color="#fff", fg_color="#3b82f6", corner_radius=6, width=60, anchor="center")
            prefix_badge.pack(side="left", padx=10, pady=8)
            
            ctk.CTkLabel(row, text=f"{prefix}001", font=ctk.CTkFont(size=13), text_color="#6b7280", width=100, anchor="w").pack(side="left", padx=10, pady=8)
        
        # Section 2: Statistiques d'utilisation
        stats_frame = ctk.CTkFrame(main_container, fg_color="#f7fafd", corner_radius=12)
        stats_frame.pack(fill="x", pady=(0, 20))
        
        ctk.CTkLabel(stats_frame, text="📊 Statistiques d'Utilisation", 
                    font=ctk.CTkFont(size=16, weight="bold"), text_color="#2563eb").pack(pady=(15, 10))
        
        # Récupérer les statistiques
        stats = MatriculeManager.get_statistics()
        
        if stats:
            stats_grid = ctk.CTkFrame(stats_frame, fg_color="transparent")
            stats_grid.pack(fill="x", padx=20, pady=(0, 15))
            
            # En-têtes des statistiques
            stats_headers = ctk.CTkFrame(stats_grid, fg_color="transparent")
            stats_headers.pack(fill="x", pady=(0, 8))
            ctk.CTkLabel(stats_headers, text="Rôle", font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151", width=200, anchor="w").pack(side="left")
            ctk.CTkLabel(stats_headers, text="Utilisateurs", font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151", width=100, anchor="w").pack(side="left")
            ctk.CTkLabel(stats_headers, text="Matricules", font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151", width=200, anchor="w").pack(side="left")
            
            for role, count in stats.items():
                row = ctk.CTkFrame(stats_grid, fg_color="white", corner_radius=8)
                row.pack(fill="x", pady=2)
                
                ctk.CTkLabel(row, text=role, font=ctk.CTkFont(size=13), text_color="#374151", width=200, anchor="w").pack(side="left", padx=10, pady=8)
                
                # Badge pour le nombre d'utilisateurs
                count_badge = ctk.CTkLabel(row, text=str(count), font=ctk.CTkFont(size=12, weight="bold"), 
                                         text_color="#fff", fg_color="#10b981", corner_radius=6, width=80, anchor="center")
                count_badge.pack(side="left", padx=10, pady=8)
                
                # Afficher les matricules existants
                matricules = MatriculeManager.get_all_matricules_by_role(role)
                matricules_text = ", ".join(matricules[:5])  # Limiter à 5 matricules
                if len(matricules) > 5:
                    matricules_text += f" ... (+{len(matricules)-5})"
                
                ctk.CTkLabel(row, text=matricules_text, font=ctk.CTkFont(size=12), text_color="#6b7280", width=200, anchor="w").pack(side="left", padx=10, pady=8)
        else:
            ctk.CTkLabel(stats_frame, text="Aucune donnée disponible", font=ctk.CTkFont(size=13), text_color="#6b7280").pack(pady=20)
        
        # Section 3: Outils de validation
        tools_frame = ctk.CTkFrame(main_container, fg_color="#f7fafd", corner_radius=12)
        tools_frame.pack(fill="x", pady=(0, 20))
        
        ctk.CTkLabel(tools_frame, text="🔧 Outils de Validation", 
                    font=ctk.CTkFont(size=16, weight="bold"), text_color="#2563eb").pack(pady=(15, 10))
        
        # Zone de test de matricule
        test_container = ctk.CTkFrame(tools_frame, fg_color="white", corner_radius=8)
        test_container.pack(fill="x", padx=20, pady=(0, 15))
        
        ctk.CTkLabel(test_container, text="Tester un matricule:", font=ctk.CTkFont(size=14, weight="bold"), text_color="#374151").pack(anchor="w", padx=15, pady=(15, 5))
        
        test_input_frame = ctk.CTkFrame(test_container, fg_color="transparent")
        test_input_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        test_entry = ctk.CTkEntry(test_input_frame, placeholder_text="Entrez un matricule (ex: AD001)", height=36)
        test_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        test_result_label = ctk.CTkLabel(test_container, text="", font=ctk.CTkFont(size=12), text_color="#6b7280")
        test_result_label.pack(anchor="w", padx=15, pady=(0, 15))
        
        def test_matricule():
            matricule = test_entry.get().strip().upper()
            if not matricule:
                test_result_label.configure(text="Veuillez entrer un matricule", text_color="#ef4444")
                return
            
            # Validation du format
            is_valid, validation_msg = MatriculeManager.validate_matricule(matricule)
            if not is_valid:
                test_result_label.configure(text=f"❌ Format invalide: {validation_msg}", text_color="#ef4444")
                return
            
            # Identification du rôle
            role = MatriculeManager.get_role_from_matricule(matricule)
            if role:
                test_result_label.configure(text=f"✅ Matricule valide - Rôle: {role}", text_color="#10b981")
            else:
                test_result_label.configure(text="⚠️ Matricule valide mais rôle non reconnu", text_color="#f59e0b")
        
        test_btn = ctk.CTkButton(test_input_frame, text="Tester", fg_color="#3b82f6", text_color="#fff", corner_radius=8, height=36, command=test_matricule)
        test_btn.pack(side="right")
        
        # Stocker les références pour le rafraîchissement
        parent._test_entry = test_entry
        parent._test_result_label = test_result_label

    def _refresh_matricules_tab(self, parent):
        """Rafraîchit l'onglet matricules"""
        # Vider le contenu existant
        for widget in parent.winfo_children():
            widget.destroy()
        
        # Reconstruire l'onglet
        self._build_matricules_tab(parent)

    def _test_matricule_generation(self):
        """Teste la génération de matricules pour différents rôles"""
        from matricule_manager import MatriculeManager
        
        # Créer une fenêtre de test
        test_window = ctk.CTkToplevel(self)
        test_window.title("Test de Génération de Matricules")
        test_window.geometry("500x400")
        test_window.grab_set()
        
        ctk.CTkLabel(test_window, text="🧪 Test de Génération de Matricules", 
                    font=ctk.CTkFont(size=18, weight="bold"), text_color="#2563eb").pack(pady=(20, 10))
        
        # Zone de résultats
        results_frame = ctk.CTkScrollableFrame(test_window, fg_color="white", corner_radius=12)
        results_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Tester différents rôles
        test_roles = [
            "Le responsable des stocks",
            "Le magasinier",
            "Les emballeurs",
            "Le responsable de la logistique"
        ]
        
        for role in test_roles:
            matricule = MatriculeManager.generate_matricule(role)
            prefix = MatriculeManager.get_role_prefix(role)
            
            row = ctk.CTkFrame(results_frame, fg_color="#f7fafd", corner_radius=8)
            row.pack(fill="x", pady=2, padx=5)
            
            ctk.CTkLabel(row, text=f"Rôle: {role}", font=ctk.CTkFont(size=13, weight="bold"), text_color="#374151").pack(anchor="w", padx=10, pady=5)
            ctk.CTkLabel(row, text=f"Préfixe: {prefix} → Matricule: {matricule}", font=ctk.CTkFont(size=12), text_color="#6b7280").pack(anchor="w", padx=10, pady=(0, 5))

    def _build_settings_tab(self, parent):
        # Détecter le thème actuel
        current_theme = ctk.get_appearance_mode()
        is_dark = current_theme == "dark"
        
        # Couleurs adaptatives
        bg_color = "#404040" if is_dark else "#f7fafd"
        text_color = "#ffffff" if is_dark else "#222222"
        secondary_text = "#cccccc" if is_dark else "#666666"
        border_color = "#555555" if is_dark else "#e0e0e0"
        
        ctk.CTkLabel(parent, text="Paramètres du Système", 
                    font=ctk.CTkFont(size=19, weight="bold"), 
                    text_color="#00ff00" if is_dark else "#2563eb").pack(pady=(18, 2))
        
        ctk.CTkLabel(parent, text="Configurez les options globales du système. Les modifications sont immédiates.", 
                    font=ctk.CTkFont(size=13), 
                    text_color=secondary_text).pack(pady=(0, 10))
        
        # Variables pour stocker les états des switches
        self.settings_switches = {}
        
        settings = [
            ("Mode sombre", "Active le thème sombre pour tous les utilisateurs.", "dark_mode"),
            ("Notifications email", "Envoie des alertes par email pour les événements critiques.", "email_notifications"),
            ("Sauvegarde automatique", "Sauvegarde les données toutes les 10 minutes.", "auto_backup"),
            ("Maintenance", "Met le système en mode maintenance (accès restreint).", "maintenance_mode"),
            ("Rafraîchissement auto", "Rafraîchit les données automatiquement toutes les 30s.", "auto_refresh"),
            ("Logs détaillés", "Active les logs détaillés pour le débogage.", "detailed_logs")
        ]
        
        for label, tip, setting_key in settings:
            row = ctk.CTkFrame(parent, fg_color=bg_color, corner_radius=10, border_width=1, border_color=border_color)
            row.pack(fill="x", padx=40, pady=8)
            
            ctk.CTkLabel(row, text=label, 
                        font=ctk.CTkFont(size=14, weight="bold"), 
                        text_color=text_color, width=220, anchor="w").pack(side="left", padx=8)
            
            # Créer le switch avec fonction de callback
            switch = ctk.CTkSwitch(row, text="", width=60)
            switch.pack(side="left", padx=8)
            
            # Configurer la commande après avoir créé le switch
            switch.configure(command=lambda key=setting_key, sw=switch: self._on_setting_changed(key, sw))
            
            # Stocker la référence du switch
            self.settings_switches[setting_key] = switch
            
            # Charger l'état actuel depuis la base de données ou les préférences
            self._load_setting_state(setting_key, switch)
            
            if HAS_TOOLTIP:
                CTkToolTip(switch, message=tip)
            else:
                ctk.CTkLabel(row, text=tip, 
                            font=ctk.CTkFont(size=12), 
                            text_color=secondary_text, anchor="w").pack(side="left", padx=8)
        
        # Boutons d'action
        buttons_frame = ctk.CTkFrame(parent, fg_color=bg_color, corner_radius=10, border_width=1, border_color=border_color)
        buttons_frame.pack(fill="x", padx=40, pady=15)
        
        ctk.CTkButton(buttons_frame, text="💾 Sauvegarder", 
                     fg_color="#00ff00" if is_dark else "#10b981", 
                     text_color="#000000" if is_dark else "#ffffff", 
                     corner_radius=8, height=35, 
                     command=self._save_all_settings).pack(side="left", padx=10, pady=10)
        
        ctk.CTkButton(buttons_frame, text="🔄 Réinitialiser", 
                     fg_color="#ff0000" if is_dark else "#ef4444", 
                     text_color="#ffffff", 
                     corner_radius=8, height=35, 
                     command=self._reset_settings).pack(side="left", padx=10, pady=10)
        
        ctk.CTkButton(buttons_frame, text="📊 Statistiques", 
                     fg_color="#0000ff" if is_dark else "#3b82f6", 
                     text_color="#ffffff", 
                     corner_radius=8, height=35, 
                     command=self._show_settings_stats).pack(side="right", padx=10, pady=10)
        
        ctk.CTkLabel(parent, text="Les modifications sont appliquées immédiatement.", 
                    font=ctk.CTkFont(size=12, slant="italic"), 
                    text_color="#888888").pack(pady=(18, 0))

    def _load_setting_state(self, setting_key, switch):
        """Charge l'état d'un paramètre depuis la base de données"""
        try:
            # Pour l'instant, on utilise des valeurs par défaut
            # Plus tard, on peut charger depuis une table settings
            default_states = {
                "dark_mode": True,  # Mode sombre activé par défaut
                "email_notifications": False,
                "auto_backup": True,
                "maintenance_mode": False,
                "auto_refresh": True,
                "detailed_logs": False
            }
            
            is_enabled = default_states.get(setting_key, False)
            switch.select() if is_enabled else switch.deselect()
            
        except Exception as e:
            print(f"Erreur lors du chargement du paramètre {setting_key}: {e}")
            switch.deselect()

    def _on_setting_changed(self, setting_key, switch):
        """Appelé quand un paramètre change"""
        is_enabled = switch.get()
        
        try:
            if setting_key == "dark_mode":
                self._apply_dark_mode(is_enabled)
            elif setting_key == "email_notifications":
                self._toggle_email_notifications(is_enabled)
            elif setting_key == "auto_backup":
                self._toggle_auto_backup(is_enabled)
            elif setting_key == "maintenance_mode":
                self._toggle_maintenance_mode(is_enabled)
            elif setting_key == "auto_refresh":
                self._toggle_auto_refresh(is_enabled)
            elif setting_key == "detailed_logs":
                self._toggle_detailed_logs(is_enabled)
            
            # Sauvegarder dans la base de données
            self._save_setting_to_db(setting_key, is_enabled)
            
            # Afficher une notification
            status = "activé" if is_enabled else "désactivé"
            self._show_setting_notification(f"{setting_key.replace('_', ' ').title()} {status}")
            
        except Exception as e:
            print(f"Erreur lors de l'application du paramètre {setting_key}: {e}")
            # Remettre le switch dans son état précédent
            switch.select() if not is_enabled else switch.deselect()

    def _apply_dark_mode(self, enabled):
        """Applique le mode sombre à toute l'application de manière intelligente"""
        try:
            if enabled:
                # Activer le mode sombre global
                ctk.set_appearance_mode("dark")
                ctk.set_default_color_theme("dark-blue")
                
                # Appliquer aux fenêtres principales
                if hasattr(self.parent, 'configure'):
                    self.parent.configure(fg_color="#1a1a1a")
                
                # Appliquer aux frames principaux
                if hasattr(self, 'main_content'):
                    self.main_content.configure(fg_color="#1a1a1a")
                
                # Appliquer aux sidebar
                if hasattr(self, 'sidebar'):
                    self.sidebar.configure(fg_color="#2d2d2d")
                
                # Notifier tous les composants du changement
                self._notify_theme_change("dark")
                
                # Forcer le rafraîchissement de l'interface
                self.parent.update()
                
            else:
                # Activer le mode clair global
                ctk.set_appearance_mode("light")
                ctk.set_default_color_theme("blue")
                
                # Appliquer aux fenêtres principales
                if hasattr(self.parent, 'configure'):
                    self.parent.configure(fg_color="white")
                
                # Appliquer aux frames principaux
                if hasattr(self, 'main_content'):
                    self.main_content.configure(fg_color="#f7fafd")
                
                # Appliquer aux sidebar
                if hasattr(self, 'sidebar'):
                    self.sidebar.configure(fg_color="white")
                
                # Notifier tous les composants du changement
                self._notify_theme_change("light")
                
                # Forcer le rafraîchissement de l'interface
                self.parent.update()
                
        except Exception as e:
            print(f"Erreur lors de l'application du thème: {e}")

    def _notify_theme_change(self, theme):
        """Notifie tous les composants du changement de thème"""
        try:
            # Notifier l'application principale
            if hasattr(self.parent, 'notify_theme_change'):
                self.parent.notify_theme_change(theme)
            
            # Notifier la sidebar
            if hasattr(self, 'sidebar') and hasattr(self.sidebar, 'apply_theme'):
                self.sidebar.apply_theme(theme)
            
            # Notifier tous les frames enfants
            for widget in self.winfo_children():
                if hasattr(widget, 'apply_theme'):
                    widget.apply_theme(theme)
                elif hasattr(widget, 'configure'):
                    # Appliquer les couleurs de base selon le thème
                    if theme == "dark":
                        widget.configure(fg_color="#2d2d2d")
                    else:
                        widget.configure(fg_color="white")
            
            print(f"Thème appliqué globalement: {theme}")
            
        except Exception as e:
            print(f"Erreur lors de la notification du thème: {e}")

    def _toggle_email_notifications(self, enabled):
        """Active/désactive les notifications email"""
        if enabled:
            print("Notifications email activées")
            # Logique pour activer les notifications
        else:
            print("Notifications email désactivées")
            # Logique pour désactiver les notifications

    def _toggle_auto_backup(self, enabled):
        """Active/désactive la sauvegarde automatique"""
        if enabled:
            print("Sauvegarde automatique activée")
            # Démarrer le processus de sauvegarde
            self._start_backup_process()
        else:
            print("Sauvegarde automatique désactivée")
            # Arrêter le processus de sauvegarde
            self._stop_backup_process()

    def _toggle_maintenance_mode(self, enabled):
        """Active/désactive le mode maintenance"""
        if enabled:
            print("Mode maintenance activé")
            # Logique pour activer le mode maintenance
            self._show_maintenance_warning()
        else:
            print("Mode maintenance désactivé")
            # Logique pour désactiver le mode maintenance

    def _toggle_auto_refresh(self, enabled):
        """Active/désactive le rafraîchissement automatique"""
        if enabled:
            print("Rafraîchissement automatique activé")
            # Démarrer le rafraîchissement automatique
        else:
            print("Rafraîchissement automatique désactivé")
            # Arrêter le rafraîchissement automatique

    def _toggle_detailed_logs(self, enabled):
        """Active/désactive les logs détaillés"""
        if enabled:
            print("Logs détaillés activés")
            # Activer les logs détaillés
        else:
            print("Logs détaillés désactivés")
            # Désactiver les logs détaillés

    def _start_backup_process(self):
        """Démarre le processus de sauvegarde automatique"""
        # Simulation d'un processus de sauvegarde
        print("Processus de sauvegarde démarré")

    def _stop_backup_process(self):
        """Arrête le processus de sauvegarde automatique"""
        # Simulation d'arrêt du processus
        print("Processus de sauvegarde arrêté")

    def _show_maintenance_warning(self):
        """Affiche un avertissement pour le mode maintenance"""
        import tkinter.messagebox as mbox
        mbox.showwarning("Mode Maintenance", 
                        "Le mode maintenance est activé.\n"
                        "L'accès sera restreint pour les utilisateurs non-administrateurs.")

    def _save_setting_to_db(self, setting_key, value):
        """Sauvegarde un paramètre dans la base de données"""
        try:
            # Ici on peut sauvegarder dans une table settings
            # Pour l'instant, on simule
            print(f"Sauvegarde: {setting_key} = {value}")
        except Exception as e:
            print(f"Erreur lors de la sauvegarde du paramètre: {e}")

    def _save_all_settings(self):
        """Sauvegarde tous les paramètres"""
        try:
            for setting_key, switch in self.settings_switches.items():
                value = switch.get()
                self._save_setting_to_db(setting_key, value)
            
            self._show_setting_notification("Tous les paramètres sauvegardés")
        except Exception as e:
            print(f"Erreur lors de la sauvegarde: {e}")

    def _reset_settings(self):
        """Réinitialise tous les paramètres"""
        try:
            for setting_key, switch in self.settings_switches.items():
                # Remettre les valeurs par défaut
                default_states = {
                    "dark_mode": True,
                    "email_notifications": False,
                    "auto_backup": True,
                    "maintenance_mode": False,
                    "auto_refresh": True,
                    "detailed_logs": False
                }
                
                is_enabled = default_states.get(setting_key, False)
                switch.select() if is_enabled else switch.deselect()
                
                # Appliquer le changement
                self._on_setting_changed(setting_key, switch)
            
            self._show_setting_notification("Paramètres réinitialisés")
        except Exception as e:
            print(f"Erreur lors de la réinitialisation: {e}")

    def _show_settings_stats(self):
        """Affiche les statistiques des paramètres"""
        try:
            stats_window = ctk.CTkToplevel(self)
            stats_window.title("Statistiques des Paramètres")
            stats_window.geometry("400x300")
            stats_window.grab_set()
            
            ctk.CTkLabel(stats_window, text="📊 Statistiques des Paramètres", 
                        font=ctk.CTkFont(size=18, weight="bold")).pack(pady=20)
            
            # Afficher les statistiques
            enabled_count = sum(1 for switch in self.settings_switches.values() if switch.get())
            total_count = len(self.settings_switches)
            
            ctk.CTkLabel(stats_window, text=f"Paramètres activés: {enabled_count}/{total_count}").pack(pady=5)
            ctk.CTkLabel(stats_window, text=f"Taux d'activation: {(enabled_count/total_count)*100:.1f}%").pack(pady=5)
            
        except Exception as e:
            print(f"Erreur lors de l'affichage des statistiques: {e}")

    def _show_setting_notification(self, message):
        """Affiche une notification pour un changement de paramètre"""
        try:
            # Créer une notification temporaire
            notif = ctk.CTkToplevel(self)
            notif.title("Notification")
            notif.geometry("300x80")
            notif.grab_set()
            notif.resizable(False, False)
            
            ctk.CTkLabel(notif, text=message, font=ctk.CTkFont(size=14)).pack(expand=True)
            
            # Fermer automatiquement après 2 secondes
            notif.after(2000, notif.destroy)
        except Exception as e:
            print(f"Erreur lors de l'affichage de la notification: {e}")

    def _get_audit_logs(self, filter_type="Tous", limit=50):
        """Récupère les logs d'activité depuis la base de données"""
        try:
            import psycopg2
            conn = psycopg2.connect(**PG_CONN)
            cursor = conn.cursor()
            
            # Requête de base
            base_query = """
                SELECT 
                    l.type_action,
                    l.description,
                    l.date_action,
                    l.utilisateur,
                    l.details
                FROM sge_cre.logs_activite l
                WHERE 1=1
            """
            
            # Ajouter le filtre par type
            if filter_type != "Tous":
                if filter_type == "Connexion":
                    base_query += " AND l.type_action IN ('Connexion', 'Déconnexion')"
                elif filter_type == "Modification":
                    base_query += " AND l.type_action IN ('Création', 'Modification', 'Mise à jour')"
                elif filter_type == "Suppression":
                    base_query += " AND l.type_action IN ('Suppression', 'Désactivation')"
            
            # Ajouter l'ordre et la limite
            base_query += " ORDER BY l.date_action DESC LIMIT %s"
            
            cursor.execute(base_query, (limit,))
            logs = cursor.fetchall()
            
            conn.close()
            
            # Formater les logs
            formatted_logs = []
            for log in logs:
                type_action, description, date_action, utilisateur, details = log
                
                # Déterminer la couleur selon le type d'action
                if type_action in ['Connexion', 'Déconnexion']:
                    color = "#10b981"  # Vert
                elif type_action in ['Création', 'Modification', 'Mise à jour']:
                    color = "#f59e42"  # Orange
                elif type_action in ['Suppression', 'Désactivation']:
                    color = "#ef4444"  # Rouge
                else:
                    color = "#3b82f6"  # Bleu par défaut
                
                # Formater la date
                try:
                    if isinstance(date_action, str):
                        dt = datetime.datetime.strptime(date_action, "%Y-%m-%d %H:%M:%S")
                    else:
                        dt = date_action
                    
                    if HAS_BABEL:
                        date_formatted = format_datetime(dt, "d MMMM yyyy HH:mm", locale="fr_FR")
                    else:
                        date_formatted = dt.strftime("%d/%m/%Y %H:%M")
                except:
                    date_formatted = str(date_action)
                
                formatted_logs.append({
                    'type': type_action,
                    'message': description,
                    'date': date_formatted,
                    'color': color,
                    'utilisateur': utilisateur,
                    'details': details
                })
            
            return formatted_logs
            
        except Exception as e:
            print(f"Erreur lors de la récupération des logs: {e}")
            # Retourner des logs par défaut en cas d'erreur
            return [
                {
                    'type': 'Erreur',
                    'message': 'Impossible de charger les logs depuis la base de données',
                    'date': datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
                    'color': '#ef4444',
                    'utilisateur': 'Système',
                    'details': str(e)
                }
            ]

    def _build_audit_tab(self, parent):
        # Stocker la référence du parent pour les mises à jour
        self.audit_parent = parent
        
        # Titre et description
        ctk.CTkLabel(parent, text="📋 Journal d'Activité", font=ctk.CTkFont(size=19, weight="bold"), text_color="#2563eb").pack(pady=(18, 2))
        ctk.CTkLabel(parent, text="Consultez l'historique des actions du système en temps réel. Filtrez par type ou utilisateur.", font=ctk.CTkFont(size=13), text_color="#666").pack(pady=(0, 10))
        
        # Conteneur pour les filtres
        filtres_container = ctk.CTkFrame(parent, fg_color="white")
        filtres_container.pack(fill="x", padx=30, pady=(0, 8))
        
        ctk.CTkLabel(filtres_container, text="🔍 Filtrer :", font=ctk.CTkFont(size=13, weight="bold"), text_color="#222").pack(side="left", padx=(15, 8))
        
        # Boutons de filtre avec stockage des références
        self.filter_buttons = {}
        filter_types = [("Tous", "#3b82f6"), ("Connexion", "#10b981"), ("Modification", "#f59e42"), ("Suppression", "#ef4444")]
        
        for txt, color in filter_types:
            btn = ctk.CTkButton(
                filtres_container, text=txt, fg_color=color, text_color="#fff", 
                width=110, height=32, corner_radius=8,
                command=lambda t=txt: self._filter_audit_logs(t)
            )
            btn.pack(side="left", padx=4)
            self.filter_buttons[txt] = btn
        
        # Conteneur pour la timeline
        self.timeline_container = ctk.CTkFrame(parent, fg_color="transparent")
        self.timeline_container.pack(fill="both", expand=True, padx=30, pady=10)
        
        # Charger les logs initiaux
        self._refresh_audit_logs("Tous")
        
        # Programmer la mise à jour automatique
        self._schedule_audit_update()
        
        # Message d'information
        info_label = ctk.CTkLabel(
            parent, 
            text="🔄 Mise à jour automatique toutes les 15 secondes • Pour plus de détails, exportez le journal ou contactez l'administrateur.", 
            font=ctk.CTkFont(size=12, slant="italic"), 
            text_color="#90A4AE"
        )
        info_label.pack(pady=(18, 0))

    def _filter_audit_logs(self, filter_type):
        """Filtre les logs d'activité selon le type sélectionné"""
        # Mettre à jour l'apparence des boutons
        for txt, btn in self.filter_buttons.items():
            if txt == filter_type:
                btn.configure(fg_color="#1e40af", text_color="white")
            else:
                # Restaurer la couleur originale
                colors = {"Tous": "#3b82f6", "Connexion": "#10b981", "Modification": "#f59e42", "Suppression": "#ef4444"}
                btn.configure(fg_color=colors.get(txt, "#3b82f6"), text_color="white")
        
        # Rafraîchir les logs avec le nouveau filtre
        self._refresh_audit_logs(filter_type)

    def _refresh_audit_logs(self, filter_type="Tous"):
        """Rafraîchit l'affichage des logs d'activité"""
        # Vider le conteneur de timeline
        for widget in self.timeline_container.winfo_children():
            widget.destroy()
        
        # Récupérer les logs depuis la base de données
        logs = self._get_audit_logs(filter_type, limit=20)
        
        if not logs:
            # Afficher un message si aucun log
            no_logs_label = ctk.CTkLabel(
                self.timeline_container, 
                text="📭 Aucune activité récente trouvée", 
                font=ctk.CTkFont(size=14), 
                text_color="#666"
            )
            no_logs_label.pack(pady=50)
            return
        
        # Créer la timeline avec les logs réels
        for log in logs:
            row = ctk.CTkFrame(self.timeline_container, fg_color="#fff", corner_radius=10, border_width=1, border_color="#e5e7eb")
            row.pack(fill="x", pady=6)
            
            # Badge du type d'action
            badge = ctk.CTkLabel(
                row, text=log['type'], 
                font=ctk.CTkFont(size=12, weight="bold"), 
                text_color="#fff", fg_color=log['color'], 
                corner_radius=8, width=100, anchor="center"
            )
            badge.pack(side="left", padx=8, pady=8)
            
            # Message principal
            message_text = f"👤 {log['utilisateur']}: {log['message']}"
            message_label = ctk.CTkLabel(
                row, text=message_text, 
                font=ctk.CTkFont(size=13), text_color="#222", 
                width=340, anchor="w", wraplength=320, justify="left"
            )
            message_label.pack(side="left", padx=2)
            
            # Date
            date_label = ctk.CTkLabel(
                row, text=log['date'], 
                font=ctk.CTkFont(size=12), text_color="#666", 
                width=160, anchor="e"
            )
            date_label.pack(side="left", padx=2)
            
            # Tooltip avec détails si disponibles
            if HAS_TOOLTIP and log.get('details'):
                CTkToolTip(badge, message=f"Type: {log['type']}\nDétails: {log['details']}")
            elif HAS_TOOLTIP:
                CTkToolTip(badge, message=f"Type d'action: {log['type']}")

    def _schedule_audit_update(self):
        """Programme la mise à jour automatique du journal d'activité"""
        def update_audit():
            try:
                # Déterminer le filtre actif
                active_filter = "Tous"
                for txt, btn in self.filter_buttons.items():
                    if btn.cget("fg_color") == "#1e40af":
                        active_filter = txt
                        break
                
                # Rafraîchir les logs
                self._refresh_audit_logs(active_filter)
                
                # Programmer la prochaine mise à jour
                self.after(15000, update_audit)  # 15 secondes
                
            except Exception as e:
                print(f"Erreur lors de la mise à jour du journal d'activité: {e}")
                # Réessayer dans 15 secondes même en cas d'erreur
                self.after(15000, update_audit)
        
        # Démarrer la première mise à jour
        self.after(15000, update_audit)

    def _log_activity(self, type_action, description, utilisateur, details=""):
        """Enregistre une action dans le journal d'activité"""
        try:
            import psycopg2
            conn = psycopg2.connect(**PG_CONN)
            cursor = conn.cursor()
            
            # Insérer le log dans la base de données
            cursor.execute("""
                INSERT INTO sge_cre.logs_activite (type_action, description, date_action, utilisateur, details)
                VALUES (%s, %s, NOW(), %s, %s)
            """, (type_action, description, utilisateur, details))
            
            conn.commit()
            conn.close()
            
            # Rafraîchir immédiatement le journal d'activité si l'onglet est ouvert
            if hasattr(self, 'timeline_container'):
                self._refresh_audit_logs("Tous")
                
        except Exception as e:
            print(f"Erreur lors de l'enregistrement du log d'activité: {e}")
            # En cas d'erreur, on peut créer la table si elle n'existe pas
            self._create_audit_table()

    def _refresh_all_interfaces(self):
        """Actualise toutes les interfaces après un changement"""
        try:
            # Actualiser les statistiques
            self._update_stats_immediately()
            
            # Actualiser la liste des utilisateurs si l'onglet est ouvert
            if hasattr(self, 'users_table') and self.users_table:
                self._refresh_users(self.users_table, "")
            
            # Actualiser le journal d'activité si l'onglet est ouvert
            if hasattr(self, 'timeline_container'):
                self._refresh_audit_logs("Tous")
            
            # Actualiser les matricules si l'onglet est ouvert
            if hasattr(self, 'matricules_parent') and self.matricules_parent:
                self._refresh_matricules_tab(self.matricules_parent)
                
        except Exception as e:
            print(f"Erreur lors de l'actualisation des interfaces: {e}")

    def _create_audit_table(self):
        """Crée la table de logs d'activité si elle n'existe pas"""
        try:
            import psycopg2
            conn = psycopg2.connect(**PG_CONN)
            cursor = conn.cursor()
            
            # Créer la table logs_activite
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS sge_cre.logs_activite (
                    id SERIAL PRIMARY KEY,
                    type_action VARCHAR(50) NOT NULL,
                    description TEXT NOT NULL,
                    date_action TIMESTAMP DEFAULT NOW(),
                    utilisateur VARCHAR(100) NOT NULL,
                    details TEXT,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            
            conn.commit()
            conn.close()
            print("Table logs_activite créée avec succès")
            
        except Exception as e:
            print(f"Erreur lors de la création de la table logs_activite: {e}")

    def _build_cli_tab(self, parent):
        # Titre et description
        ctk.CTkLabel(parent, text="💻 Terminal CLI", font=ctk.CTkFont(size=19, weight="bold"), text_color="#00ff00").pack(pady=(18, 2))
        ctk.CTkLabel(parent, text="Terminal système avancé - Gestion complète du SGE via commandes", font=ctk.CTkFont(size=13), text_color="#cccccc").pack(pady=(0, 10))
        
        # Terminal moderne et sombre
        cli_frame = ctk.CTkFrame(parent, fg_color="#0a0a0a", corner_radius=12, border_width=2, border_color="#333333")
        cli_frame.pack(fill="both", expand=True, padx=30, pady=10)
        
        # Zone de sortie avec scroll
        output = ctk.CTkTextbox(cli_frame, font=ctk.CTkFont(size=12, family="Consolas"), 
                               fg_color="#0a0a0a", text_color="#00ff00", 
                               border_width=1, border_color="#333333")
        output.pack(pady=8, fill="both", expand=True, padx=8)
        
        # Message de bienvenue
        output.insert("end", "╔══════════════════════════════════════════════════════════════╗\n")
        output.insert("end", "║                    SAC Terminal v2.0                        ║\n")
        output.insert("end", "║              Système de Gestion d'Entrepôts                 ║\n")
        output.insert("end", "╚══════════════════════════════════════════════════════════════╝\n\n")
        output.insert("end", "🔧 Tapez 'help' pour voir toutes les commandes disponibles\n")
        output.insert("end", "📚 Tapez 'help <commande>' pour l'aide sur une commande spécifique\n\n")
        output.insert("end", "root@sac:~$ ")
        
        # Barre de commande moderne
        cmd_frame = ctk.CTkFrame(cli_frame, fg_color="#1a1a1a", height=40, corner_radius=8)
        cmd_frame.pack(fill="x", padx=8, pady=(0, 8))
        cmd_frame.pack_propagate(False)
        
        # Label pour le prompt
        prompt_label = ctk.CTkLabel(cmd_frame, text="root@sac:~$ ", font=ctk.CTkFont(size=12, family="Consolas"), 
                                   text_color="#00ff00", fg_color="transparent")
        prompt_label.pack(side="left", padx=(10, 5))
        
        # Zone de saisie
        entry = ctk.CTkEntry(cmd_frame, placeholder_text="Entrez votre commande...", 
                            font=ctk.CTkFont(size=12, family="Consolas"), 
                            height=30, fg_color="#2a2a2a", text_color="#00ff00", 
                            placeholder_text_color="#666666", border_width=1, border_color="#444444")
        entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        # Stocker les références pour les utiliser dans les commandes
        self.cli_output = output
        self.cli_entry = entry
        
        def execute_command(cmd):
            """Exécute une commande CLI"""
            if not cmd:
                return
            
            output.insert("end", f"{cmd}\n")
            
            # Parser la commande
            parts = cmd.split()
            command = parts[0].lower()
            args = parts[1:] if len(parts) > 1 else []
            
            try:
                if command == "help":
                    self._cli_help(args)
                elif command == "user":
                    self._cli_user(args)
                elif command == "product":
                    self._cli_product(args)
                elif command == "report":
                    self._cli_report(args)
                elif command == "export":
                    self._cli_export(args)
                elif command == "download":
                    self._cli_download(args)
                elif command == "clear":
                    output.delete("1.0", "end")
                    output.insert("end", "╔══════════════════════════════════════════════════════════════╗\n")
                    output.insert("end", "║                    SAC Terminal v2.0                        ║\n")
                    output.insert("end", "║              Système de Gestion d'Entrepôts                 ║\n")
                    output.insert("end", "╚══════════════════════════════════════════════════════════════╝\n\n")
                    output.insert("end", "🔧 Tapez 'help' pour voir toutes les commandes disponibles\n")
                    output.insert("end", "📚 Tapez 'help <commande>' pour l'aide sur une commande spécifique\n\n")
                elif command == "date":
                    import datetime
                    now = datetime.datetime.now()
                    output.insert("end", f"📅 {now.strftime('%d/%m/%Y %H:%M:%S')}\n")
                elif command == "status":
                    self._cli_status()
                elif command == "exit" or command == "quit":
                    output.insert("end", "👋 Déconnexion du terminal...\n")
                else:
                    output.insert("end", f"❌ Commande '{command}' non reconnue. Tapez 'help' pour l'aide.\n")
            except Exception as e:
                output.insert("end", f"❌ Erreur: {str(e)}\n")
            output.insert("end", "root@sac:~$ ")
            output.see("end")
        
        def on_enter(event=None):
            cmd = entry.get().strip()
            entry.delete(0, "end")
            execute_command(cmd)
        
        entry.bind('<Return>', on_enter)
        
        # Bouton d'exécution
        ctk.CTkButton(cmd_frame, text="▶ Exécuter", width=100, height=30, fg_color="#00aa00", 
                     text_color="#ffffff", corner_radius=6, command=on_enter).pack(side="right", padx=10)
        
        # Informations d'aide
        help_frame = ctk.CTkFrame(parent, fg_color="#1a1a1a", corner_radius=8)
        help_frame.pack(fill="x", padx=30, pady=(0, 10))
        
        ctk.CTkLabel(help_frame, text="💡 Astuce: Utilisez Tab pour l'auto-complétion • Ctrl+L pour effacer • ↑↓ pour l'historique", 
                    font=ctk.CTkFont(size=11, slant="italic"), text_color="#888888").pack(pady=8)

    def _cli_help(self, args):
        """Affiche l'aide des commandes CLI"""
        output = self.cli_output
        
        if not args:
            # Aide générale
            output.insert("end", "\n📚 COMMANDES DISPONIBLES:\n")
            output.insert("end", "═" * 60 + "\n\n")
            
            output.insert("end", "👥 GESTION DES UTILISATEURS:\n")
            output.insert("end", "  user add <nom> <prenom> <email> <role>     - Ajouter un utilisateur\n")
            output.insert("end", "  user del <email>                           - Supprimer un utilisateur\n")
            output.insert("end", "  user mod <email> <champ> <valeur>          - Modifier un utilisateur\n")
            output.insert("end", "  user list                                  - Lister tous les utilisateurs\n")
            output.insert("end", "  user search <terme>                        - Rechercher un utilisateur\n\n")
            
            output.insert("end", "📦 GESTION DES PRODUITS:\n")
            output.insert("end", "  product add <nom> <description> <prix>     - Ajouter un produit\n")
            output.insert("end", "  product del <id>                           - Supprimer un produit\n")
            output.insert("end", "  product mod <id> <champ> <valeur>          - Modifier un produit\n")
            output.insert("end", "  product list                               - Lister tous les produits\n")
            output.insert("end", "  product search <terme>                     - Rechercher un produit\n\n")
            
            output.insert("end", "📊 RAPPORTS ET EXPORTS:\n")
            output.insert("end", "  report users                               - Rapport des utilisateurs\n")
            output.insert("end", "  report products                            - Rapport des produits\n")
            output.insert("end", "  report movements                           - Rapport des mouvements\n")
            output.insert("end", "  report packaging                           - Rapport des emballages\n\n")
            
            output.insert("end", "📥 EXPORTS ET TÉLÉCHARGEMENTS:\n")
            output.insert("end", "  export users [format]                      - Exporter les utilisateurs\n")
            output.insert("end", "  export products [format]                   - Exporter les produits\n")
            output.insert("end", "  export movements [format]                  - Exporter les mouvements\n")
            output.insert("end", "  download expedition <id>                   - Télécharger bon d'expédition\n")
            output.insert("end", "  download reception <id>                    - Télécharger bon de réception\n\n")
            
            output.insert("end", "🔧 COMMANDES SYSTÈME:\n")
            output.insert("end", "  clear                                      - Effacer l'écran\n")
            output.insert("end", "  date                                       - Afficher la date/heure\n")
            output.insert("end", "  status                                     - Statut du système\n")
            output.insert("end", "  exit/quit                                  - Quitter le terminal\n\n")
            
            output.insert("end", "📝 Formats d'export supportés: csv, xlsx, pdf, json\n")
            output.insert("end", "💡 Exemple: export users csv\n\n")
            
        else:
            # Aide spécifique pour une commande
            command = args[0].lower()
            if command == "user":
                output.insert("end", "\n👥 AIDE - GESTION DES UTILISATEURS:\n")
                output.insert("end", "═" * 50 + "\n")
                output.insert("end", "user add <nom> <prenom> <email> <role>\n")
                output.insert("end", "  Ajoute un nouvel utilisateur au système\n")
                output.insert("end", "  Rôles disponibles: admin, manager, operator\n")
                output.insert("end", "  Exemple: user add Dupont Jean jean@example.com admin\n\n")
                
                output.insert("end", "user del <email>\n")
                output.insert("end", "  Supprime un utilisateur par son email\n")
                output.insert("end", "  Exemple: user del jean@example.com\n\n")
                
                output.insert("end", "user mod <email> <champ> <valeur>\n")
                output.insert("end", "  Modifie un champ d'un utilisateur\n")
                output.insert("end", "  Champs: nom, prenom, email, role, actif\n")
                output.insert("end", "  Exemple: user mod jean@example.com role manager\n\n")
                
                output.insert("end", "user list\n")
                output.insert("end", "  Affiche la liste de tous les utilisateurs\n\n")
                
                output.insert("end", "user search <terme>\n")
                output.insert("end", "  Recherche un utilisateur par nom, email ou matricule\n")
                output.insert("end", "  Exemple: user search Dupont\n\n")
                
            elif command == "product":
                output.insert("end", "\n📦 AIDE - GESTION DES PRODUITS:\n")
                output.insert("end", "═" * 50 + "\n")
                output.insert("end", "product add <nom> <description> <prix>\n")
                output.insert("end", "  Ajoute un nouveau produit\n")
                output.insert("end", "  Exemple: product add \"Chocolat Noir\" \"Chocolat 70% cacao\" 5.99\n\n")
                
                output.insert("end", "product del <id>\n")
                output.insert("end", "  Supprime un produit par son ID\n")
                output.insert("end", "  Exemple: product del 123\n\n")
                
                output.insert("end", "product mod <id> <champ> <valeur>\n")
                output.insert("end", "  Modifie un champ d'un produit\n")
                output.insert("end", "  Champs: nom, description, prix, stock\n")
                output.insert("end", "  Exemple: product mod 123 prix 6.50\n\n")
                
                output.insert("end", "product list\n")
                output.insert("end", "  Affiche la liste de tous les produits\n\n")
                
                output.insert("end", "product search <terme>\n")
                output.insert("end", "  Recherche un produit par nom ou description\n")
                output.insert("end", "  Exemple: product search chocolat\n\n")
                
            elif command == "export":
                output.insert("end", "\n📤 AIDE - EXPORTS:\n")
                output.insert("end", "═" * 50 + "\n")
                output.insert("end", "export <table> [format]\n")
                output.insert("end", "  Exporte une table vers un fichier\n")
                output.insert("end", "  Tables: users, products, movements, packaging\n")
                output.insert("end", "  Formats: csv, xlsx, pdf, json (défaut: csv)\n")
                output.insert("end", "  Exemples:\n")
                output.insert("end", "    export users csv\n")
                output.insert("end", "    export products xlsx\n")
                output.insert("end", "    export movements pdf\n\n")
                
            elif command == "download":
                output.insert("end", "\n📥 AIDE - TÉLÉCHARGEMENTS:\n")
                output.insert("end", "═" * 50 + "\n")
                output.insert("end", "download expedition <id>\n")
                output.insert("end", "  Télécharge un bon d'expédition\n")
                output.insert("end", "  Exemple: download expedition 456\n\n")
                
                output.insert("end", "download reception <id>\n")
                output.insert("end", "  Télécharge un bon de réception\n")
                output.insert("end", "  Exemple: download reception 789\n\n")
                
            else:
                output.insert("end", f"❌ Aucune aide disponible pour la commande '{command}'\n")
                output.insert("end", "💡 Tapez 'help' pour voir toutes les commandes disponibles\n")

    def _cli_user(self, args):
        """Gestion des utilisateurs via CLI"""
        output = self.cli_output
        
        if not args:
            output.insert("end", "❌ Commande incomplète. Tapez 'help user' pour l'aide.\n")
            return
        
        action = args[0].lower()
        
        try:
            if action == "add" and len(args) >= 5:
                nom, prenom, email, role = args[1], args[2], args[3], args[4]
                
                # Vérifier que le rôle est valide
                valid_roles = [
                    "Le responsable des stocks",
                    "Le magasinier",
                    "Les emballeurs",
                    "Le responsable de la logistique",
                    "Les agents de logistique",
                    "Les livreurs",
                    "Le responsable informatique",
                    "Les techniciens informatiques",
                    "Le responsable de la sécurité physique",
                    "Les gardes de sécurité"
                ]
                if role not in valid_roles:
                    output.insert("end", f"❌ Rôle '{role}' non autorisé. Rôles valides: {', '.join(valid_roles)}\n")
                    return
                
                # Générer un matricule automatique
                from matricule_manager import MatriculeManager
                matricule = MatriculeManager.generate_matricule(role)
                
                # Ajouter l'utilisateur en base
                import psycopg2
                conn = psycopg2.connect(**PG_CONN)
                cursor = conn.cursor()
                
                # Vérifier si l'email existe déjà
                cursor.execute("SELECT email FROM sge_cre.individus WHERE email = %s", (email,))
                if cursor.fetchone():
                    output.insert("end", f"❌ L'email {email} existe déjà dans la base de données.\n")
                    conn.close()
                    return
                
                cursor.execute("""
                    INSERT INTO sge_cre.individus (nom, prenom, email, password, role, matricule, adresse, telephone)
                    VALUES (%s, %s, %s, %s, %s, %s, 'Adresse par défaut', 'Téléphone par défaut')
                """, (nom, prenom, email, "password123", role, matricule))
                
                conn.commit()
                conn.close()
                
                output.insert("end", f"✅ Utilisateur {prenom} {nom} ajouté avec succès!\n")
                output.insert("end", f"📧 Email: {email}\n")
                output.insert("end", f"🆔 Matricule: {matricule}\n")
                output.insert("end", f"👤 Rôle: {role}\n")
                
                # Log de l'activité
                self._log_activity("Création", f"Création utilisateur via CLI: {prenom} {nom}", "CLI", f"Email: {email}, Rôle: {role}")
                
                # Actualiser les interfaces
                self._refresh_all_interfaces()
                
            elif action == "del" and len(args) >= 2:
                email = args[1]
                
                import psycopg2
                conn = psycopg2.connect(**PG_CONN)
                cursor = conn.cursor()
                
                # Vérifier si l'utilisateur existe
                cursor.execute("SELECT nom, prenom FROM sge_cre.individus WHERE email = %s", (email,))
                user = cursor.fetchone()
                
                if user:
                    cursor.execute("DELETE FROM sge_cre.individus WHERE email = %s", (email,))
                    conn.commit()
                    output.insert("end", f"✅ Utilisateur {user[1]} {user[0]} supprimé avec succès!\n")
                    
                    # Log de l'activité
                    self._log_activity("Suppression", f"Suppression utilisateur via CLI: {user[1]} {user[0]}", "CLI", f"Email: {email}")
                    
                    # Actualiser les interfaces
                    self._refresh_all_interfaces()
                else:
                    output.insert("end", f"❌ Utilisateur avec l'email {email} non trouvé.\n")
                
                conn.close()
                
            elif action == "mod" and len(args) >= 4:
                email, champ, valeur = args[1], args[2], args[3]
                
                import psycopg2
                conn = psycopg2.connect(**PG_CONN)
                cursor = conn.cursor()
                
                # Vérifier si l'utilisateur existe
                cursor.execute("SELECT nom, prenom FROM sge_cre.individus WHERE email = %s", (email,))
                user = cursor.fetchone()
                
                if user:
                    # Vérifier que le champ existe
                    allowed_fields = ['nom', 'prenom', 'email', 'role', 'adresse', 'telephone']
                    if champ not in allowed_fields:
                        output.insert("end", f"❌ Champ '{champ}' non autorisé. Champs autorisés: {', '.join(allowed_fields)}\n")
                        conn.close()
                        return
                    
                    # Si on modifie le rôle, vérifier qu'il est valide
                    if champ == 'role':
                        valid_roles = [
                            "Le responsable des stocks",
                            "Le magasinier",
                            "Les emballeurs",
                            "Le responsable de la logistique",
                            "Les agents de logistique",
                            "Les livreurs",
                            "Le responsable informatique",
                            "Les techniciens informatiques",
                            "Le responsable de la sécurité physique",
                            "Les gardes de sécurité"
                        ]
                        if valeur not in valid_roles:
                            output.insert("end", f"❌ Rôle '{valeur}' non autorisé. Rôles valides: {', '.join(valid_roles)}\n")
                            conn.close()
                            return
                    
                    cursor.execute(f"UPDATE sge_cre.individus SET {champ} = %s WHERE email = %s", (valeur, email))
                    conn.commit()
                    output.insert("end", f"✅ Utilisateur {user[1]} {user[0]} modifié avec succès!\n")
                    output.insert("end", f"📝 Champ '{champ}' mis à jour vers '{valeur}'\n")
                    
                    # Log de l'activité
                    self._log_activity("Modification", f"Modification utilisateur via CLI: {user[1]} {user[0]}", "CLI", f"Champ: {champ}, Nouvelle valeur: {valeur}")
                    
                    # Actualiser les interfaces
                    self._refresh_all_interfaces()
                else:
                    output.insert("end", f"❌ Utilisateur avec l'email {email} non trouvé.\n")
                
                conn.close()
                
            elif action == "list":
                import psycopg2
                conn = psycopg2.connect(**PG_CONN)
                cursor = conn.cursor()
                
                cursor.execute("SELECT nom, prenom, email, role, matricule FROM sge_cre.individus ORDER BY nom")
                users = cursor.fetchall()
                
                if users:
                    output.insert("end", "\n👥 LISTE DES UTILISATEURS:\n")
                    output.insert("end", "─" * 80 + "\n")
                    output.insert("end", f"{'Nom':<15} {'Prénom':<15} {'Email':<25} {'Rôle':<20} {'Matricule':<10} {'Statut':<8}\n")
                    output.insert("end", "─" * 80 + "\n")
                    
                    for user in users:
                        nom, prenom, email, role, matricule, actif = user
                        statut = "🟢 Actif" if actif else "🔴 Inactif"
                        output.insert("end", f"{nom:<15} {prenom:<15} {email:<25} {role:<20} {matricule:<10} {statut:<8}\n")
                    
                    output.insert("end", f"\n📊 Total: {len(users)} utilisateur(s)\n")
                else:
                    output.insert("end", "📭 Aucun utilisateur trouvé.\n")
                
                conn.close()
                
            elif action == "search" and len(args) >= 2:
                terme = args[1]
                
                import psycopg2
                conn = psycopg2.connect(**PG_CONN)
                cursor = conn.cursor()
                
                cursor.execute("""
                    SELECT nom, prenom, email, role, matricule, actif 
                    FROM sge_cre.individus 
                    WHERE nom ILIKE %s OR prenom ILIKE %s OR email ILIKE %s OR matricule ILIKE %s
                    ORDER BY nom
                """, (f"%{terme}%", f"%{terme}%", f"%{terme}%", f"%{terme}%"))
                
                users = cursor.fetchall()
                
                if users:
                    output.insert("end", f"\n🔍 RÉSULTATS DE RECHERCHE POUR '{terme}':\n")
                    output.insert("end", "─" * 80 + "\n")
                    output.insert("end", f"{'Nom':<15} {'Prénom':<15} {'Email':<25} {'Rôle':<20} {'Matricule':<10} {'Statut':<8}\n")
                    output.insert("end", "─" * 80 + "\n")
                    
                    for user in users:
                        nom, prenom, email, role, matricule, actif = user
                        statut = "🟢 Actif" if actif else "🔴 Inactif"
                        output.insert("end", f"{nom:<15} {prenom:<15} {email:<25} {role:<20} {matricule:<10} {statut:<8}\n")
                    
                    output.insert("end", f"\n📊 {len(users)} résultat(s) trouvé(s)\n")
                else:
                    output.insert("end", f"❌ Aucun utilisateur trouvé pour '{terme}'.\n")
                
                conn.close()
                
            else:
                output.insert("end", "❌ Commande incomplète. Tapez 'help user' pour l'aide.\n")
                
        except Exception as e:
            output.insert("end", f"❌ Erreur: {str(e)}\n")

    def _cli_product(self, args):
        """Gestion des produits via CLI"""
        output = self.cli_output
        
        if not args:
            output.insert("end", "❌ Commande incomplète. Tapez 'help product' pour l'aide.\n")
            return
        
        action = args[0].lower()
        
        try:
            if action == "add" and len(args) >= 4:
                nom = args[1]
                description = args[2]
                prix = float(args[3])
                
                import psycopg2
                conn = psycopg2.connect(**PG_CONN)
                cursor = conn.cursor()
                
                # Vérifier si le produit existe déjà
                cursor.execute("SELECT nom FROM sge_cre.produits WHERE nom = %s", (nom,))
                if cursor.fetchone():
                    output.insert("end", f"❌ Le produit '{nom}' existe déjà dans la base de données.\n")
                    conn.close()
                    return
                
                cursor.execute("""
                    INSERT INTO sge_cre.produits (nom, description, prix, stock_disponible, categorie, fournisseur)
                    VALUES (%s, %s, %s, 0, 'Catégorie par défaut', 'Fournisseur par défaut')
                """, (nom, description, prix))
                
                conn.commit()
                conn.close()
                
                output.insert("end", f"✅ Produit '{nom}' ajouté avec succès!\n")
                output.insert("end", f"📝 Description: {description}\n")
                output.insert("end", f"💰 Prix: {prix}€\n")
                output.insert("end", f"📦 Stock initial: 0\n")
                
                # Log de l'activité
                self._log_activity("Création", f"Création produit via CLI: {nom}", "CLI", f"Prix: {prix}€")
                
                # Actualiser les interfaces
                self._refresh_all_interfaces()
                
            elif action == "del" and len(args) >= 2:
                nom = args[1]
                
                import psycopg2
                conn = psycopg2.connect(**PG_CONN)
                cursor = conn.cursor()
                
                # Vérifier si le produit existe
                cursor.execute("SELECT nom FROM sge_cre.produits WHERE nom = %s", (nom,))
                if cursor.fetchone():
                    cursor.execute("DELETE FROM sge_cre.produits WHERE nom = %s", (nom,))
                    conn.commit()
                    output.insert("end", f"✅ Produit '{nom}' supprimé avec succès!\n")
                    
                    # Log de l'activité
                    self._log_activity("Suppression", f"Suppression produit via CLI: {nom}", "CLI", "")
                    
                    # Actualiser les interfaces
                    self._refresh_all_interfaces()
                else:
                    output.insert("end", f"❌ Produit '{nom}' non trouvé.\n")
                
                conn.close()
                
            elif action == "mod" and len(args) >= 4:
                nom, champ, valeur = args[1], args[2], args[3]
                
                import psycopg2
                conn = psycopg2.connect(**PG_CONN)
                cursor = conn.cursor()
                
                # Vérifier si le produit existe
                cursor.execute("SELECT nom FROM sge_cre.produits WHERE nom = %s", (nom,))
                if cursor.fetchone():
                    # Vérifier que le champ existe
                    allowed_fields = ['nom', 'description', 'prix', 'stock_disponible', 'categorie', 'fournisseur']
                    if champ not in allowed_fields:
                        output.insert("end", f"❌ Champ '{champ}' non autorisé. Champs autorisés: {', '.join(allowed_fields)}\n")
                        conn.close()
                        return
                    
                    cursor.execute(f"UPDATE sge_cre.produits SET {champ} = %s WHERE nom = %s", (valeur, nom))
                    conn.commit()
                    output.insert("end", f"✅ Produit '{nom}' modifié avec succès!\n")
                    output.insert("end", f"📝 Champ '{champ}' mis à jour vers '{valeur}'\n")
                    
                    # Log de l'activité
                    self._log_activity("Modification", f"Modification produit via CLI: {nom}", "CLI", f"Champ: {champ}, Nouvelle valeur: {valeur}")
                    
                    # Actualiser les interfaces
                    self._refresh_all_interfaces()
                else:
                    output.insert("end", f"❌ Produit '{nom}' non trouvé.\n")
                
                conn.close()
                
            elif action == "list":
                import psycopg2
                conn = psycopg2.connect(**PG_CONN)
                cursor = conn.cursor()
                
                cursor.execute("SELECT id, nom, description, prix, stock_disponible FROM sge_cre.produits ORDER BY nom")
                products = cursor.fetchall()
                
                if products:
                    output.insert("end", "\n📦 LISTE DES PRODUITS:\n")
                    output.insert("end", "─" * 80 + "\n")
                    output.insert("end", f"{'ID':<5} {'Nom':<20} {'Description':<25} {'Prix':<8} {'Stock':<8}\n")
                    output.insert("end", "─" * 80 + "\n")
                    
                    for product in products:
                        id_prod, nom, description, prix, stock = product
                        output.insert("end", f"{id_prod:<5} {nom:<20} {description:<25} {prix:<8}€ {stock:<8}\n")
                    
                    output.insert("end", f"\n📊 Total: {len(products)} produit(s)\n")
                else:
                    output.insert("end", "📭 Aucun produit trouvé.\n")
                
                conn.close()
                
            elif action == "search" and len(args) >= 2:
                terme = args[1]
                
                import psycopg2
                conn = psycopg2.connect(**PG_CONN)
                cursor = conn.cursor()
                
                cursor.execute("""
                    SELECT id, nom, description, prix, stock_disponible 
                    FROM sge_cre.produits 
                    WHERE nom ILIKE %s OR description ILIKE %s
                    ORDER BY nom
                """, (f"%{terme}%", f"%{terme}%"))
                
                products = cursor.fetchall()
                
                if products:
                    output.insert("end", f"\n🔍 RÉSULTATS DE RECHERCHE POUR '{terme}':\n")
                    output.insert("end", "─" * 80 + "\n")
                    output.insert("end", f"{'ID':<5} {'Nom':<20} {'Description':<25} {'Prix':<8} {'Stock':<8}\n")
                    output.insert("end", "─" * 80 + "\n")
                    
                    for product in products:
                        id_prod, nom, description, prix, stock = product
                        output.insert("end", f"{id_prod:<5} {nom:<20} {description:<25} {prix:<8}€ {stock:<8}\n")
                    
                    output.insert("end", f"\n📊 {len(products)} résultat(s) trouvé(s)\n")
                else:
                    output.insert("end", f"❌ Aucun produit trouvé pour '{terme}'.\n")
                
                conn.close()
                
            else:
                output.insert("end", "❌ Commande incomplète. Tapez 'help product' pour l'aide.\n")
                
        except Exception as e:
            output.insert("end", f"❌ Erreur: {str(e)}\n")

    def _cli_report(self, args):
        """Génération de rapports via CLI"""
        output = self.cli_output
        
        if not args:
            output.insert("end", "❌ Commande incomplète. Tapez 'help report' pour l'aide.\n")
            return
        
        report_type = args[0].lower()
        
        try:
            import psycopg2
            conn = psycopg2.connect(**PG_CONN)
            cursor = conn.cursor()
            
            if report_type == "users":
                output.insert("end", "📊 Génération du rapport des utilisateurs...\n")
                
                # Statistiques des utilisateurs
                cursor.execute("SELECT COUNT(*) FROM sge_cre.individus")
                total_users = cursor.fetchone()[0]

                # Sans colonne actif, on considère tous les utilisateurs comme actifs
                active_users = total_users

                cursor.execute("SELECT role, COUNT(*) FROM sge_cre.individus GROUP BY role")
                roles_stats = cursor.fetchall()

                output.insert("end", f"✅ Rapport des utilisateurs généré!\n")
                output.insert("end", f"📊 Total utilisateurs: {total_users}\n")
                output.insert("end", f"🟢 Utilisateurs actifs: {active_users}\n")
                output.insert("end", f"📈 Répartition par rôle:\n")

                for role, count in roles_stats:
                    output.insert("end", f"   • {role}: {count} utilisateur(s)\n")

                # Log de l'activité
                self._log_activity("Rapport", f"Génération rapport utilisateurs via CLI", "CLI", f"Total: {total_users}, Actifs: {active_users}")
                
            elif report_type == "products":
                output.insert("end", "📊 Génération du rapport des produits...\n")
                
                # Statistiques des produits
                cursor.execute("SELECT COUNT(*) FROM sge_cre.produits")
                total_products = cursor.fetchone()[0]
                
                cursor.execute("SELECT COUNT(*) FROM sge_cre.produits WHERE stock_disponible > 0")
                in_stock = cursor.fetchone()[0]
                
                cursor.execute("SELECT AVG(prix) FROM sge_cre.produits")
                avg_price = cursor.fetchone()[0] or 0
                
                cursor.execute("SELECT SUM(stock_disponible) FROM sge_cre.produits")
                total_stock = cursor.fetchone()[0] or 0
                
                output.insert("end", f"✅ Rapport des produits généré!\n")
                output.insert("end", f"📦 Total produits: {total_products}\n")
                output.insert("end", f"📥 En stock: {in_stock}\n")
                output.insert("end", f"💰 Prix moyen: {avg_price:.2f}€\n")
                output.insert("end", f"📊 Stock total: {total_stock} unités\n")
                
                # Log de l'activité
                self._log_activity("Rapport", f"Génération rapport produits via CLI", "CLI", f"Total: {total_products}, En stock: {in_stock}")
                
            elif report_type == "movements":
                output.insert("end", "📊 Génération du rapport des mouvements...\n")
                
                # Statistiques des mouvements
                cursor.execute("SELECT COUNT(*) FROM sge_cre.mouvements_stock")
                total_movements = cursor.fetchone()[0]
                
                cursor.execute("SELECT type_mouvement, COUNT(*) FROM sge_cre.mouvements_stock GROUP BY type_mouvement")
                movement_types = cursor.fetchall()
                
                cursor.execute("SELECT DATE(date_mouvement), COUNT(*) FROM sge_cre.mouvements_stock GROUP BY DATE(date_mouvement) ORDER BY DATE(date_mouvement) DESC LIMIT 7")
                recent_movements = cursor.fetchall()
                
                output.insert("end", f"✅ Rapport des mouvements généré!\n")
                output.insert("end", f"📊 Total mouvements: {total_movements}\n")
                output.insert("end", f"📈 Types de mouvements:\n")
                
                for mvt_type, count in movement_types:
                    output.insert("end", f"   • {mvt_type}: {count} mouvement(s)\n")
                
                output.insert("end", f"📅 Mouvements récents (7 derniers jours):\n")
                for date, count in recent_movements:
                    output.insert("end", f"   • {date}: {count} mouvement(s)\n")
                
                # Log de l'activité
                self._log_activity("Rapport", f"Génération rapport mouvements via CLI", "CLI", f"Total: {total_movements}")
                
            elif report_type == "packaging":
                output.insert("end", "📊 Génération du rapport des emballages...\n")
                
                # Statistiques des emballages
                cursor.execute("SELECT COUNT(*) FROM sge_cre.emballages")
                total_packaging = cursor.fetchone()[0]
                
                cursor.execute("SELECT type_emballage, COUNT(*) FROM sge_cre.emballages GROUP BY type_emballage")
                packaging_types = cursor.fetchall()
                
                output.insert("end", f"✅ Rapport des emballages généré!\n")
                output.insert("end", f"📦 Total emballages: {total_packaging}\n")
                output.insert("end", f"📈 Types d'emballages:\n")
                
                for pkg_type, count in packaging_types:
                    output.insert("end", f"   • {pkg_type}: {count} emballage(s)\n")
                
                # Log de l'activité
                self._log_activity("Rapport", f"Génération rapport emballages via CLI", "CLI", f"Total: {total_packaging}")
                
            else:
                output.insert("end", f"❌ Type de rapport '{report_type}' non reconnu.\n")
            
            conn.close()
                
        except Exception as e:
            output.insert("end", f"❌ Erreur: {str(e)}\n")

    def _cli_export(self, args):
        """Export de données via CLI"""
        output = self.cli_output
        
        if not args:
            output.insert("end", "❌ Commande incomplète. Tapez 'help export' pour l'aide.\n")
            return
        
        table = args[0].lower()
        format_export = args[1].lower() if len(args) > 1 else "csv"
        
        try:
            import psycopg2
            import csv
            import json
            import datetime
            import os
            if not os.path.exists("exports"):
                os.makedirs("exports")
            # Obtenir le dossier Téléchargements du système
            downloads_path = self.get_downloads_folder()
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{downloads_path}/SGE_{table}_{timestamp}.{format_export}"
            # Créer aussi une copie dans le dossier exports local
            local_filename = f"exports/{table}_{timestamp}.{format_export}"
            try:
                conn = psycopg2.connect(**PG_CONN)
                cursor = conn.cursor()

                if table == "users":
                    output.insert("end", f"📤 Export des utilisateurs au format {format_export.upper()}...\n")
                    
                    cursor.execute("SELECT nom, prenom, email, role, matricule, adresse, telephone FROM sge_cre.individus ORDER BY nom")
                    users = cursor.fetchall()
                    
                    if format_export == "csv":
                        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                            writer = csv.writer(csvfile, delimiter=';')
                            # En-têtes avec formatage
                            writer.writerow(['NOM', 'PRÉNOM', 'EMAIL', 'RÔLE', 'MATRICULE', 'ADRESSE', 'TÉLÉPHONE'])
                            # Séparateur visuel
                            writer.writerow(['---', '---', '---', '---', '---', '---', '---'])
                            # Données
                            for user in users:
                                writer.writerow(user)
                    elif format_export == "xlsx":
                        try:
                            import pandas as pd
                            from openpyxl import Workbook
                            from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
                            
                            # Créer un DataFrame
                            df = pd.DataFrame(users, columns=['NOM', 'PRÉNOM', 'EMAIL', 'RÔLE', 'MATRICULE', 'ADRESSE', 'TÉLÉPHONE'])
                            
                            # Exporter vers Excel avec formatage
                            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                                df.to_excel(writer, sheet_name='Utilisateurs', index=False)
                                
                                # Récupérer le workbook et la worksheet
                                workbook = writer.book
                                worksheet = writer.sheets['Utilisateurs']
                                
                                # Style pour les en-têtes
                                header_font = Font(bold=True, color="FFFFFF", size=12)
                                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                                header_alignment = Alignment(horizontal="center", vertical="center")
                                
                                # Appliquer le style aux en-têtes
                                for col in range(1, len(df.columns) + 1):
                                    cell = worksheet.cell(row=1, column=col)
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = header_alignment
                                
                                # Style alterné pour les lignes de données
                                for row in range(2, len(df) + 2):
                                    for col in range(1, len(df.columns) + 1):
                                        cell = worksheet.cell(row=row, column=col)
                                        if row % 2 == 0:
                                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                                
                                # Ajuster la largeur des colonnes
                                for column in worksheet.columns:
                                    max_length = 0
                                    column_letter = column[0].column_letter
                                    for cell in column:
                                        try:
                                            if len(str(cell.value)) > max_length:
                                                max_length = len(str(cell.value))
                                        except:
                                            pass
                                    adjusted_width = min(max_length + 2, 50)
                                    worksheet.column_dimensions[column_letter].width = adjusted_width
                                    
                        except ImportError:
                            output.insert("end", "❌ Erreur: pandas et openpyxl requis pour l'export Excel\n")
                            return
                            
                    elif format_export == "json":
                        data = []
                        for user in users:
                            data.append({
                                'nom': user[0],
                                'prenom': user[1],
                                'email': user[2],
                                'role': user[3],
                                'matricule': user[4],
                                'adresse': user[5],
                                'telephone': user[6]
                            })
                        with open(filename, 'w', encoding='utf-8') as jsonfile:
                            json.dump(data, jsonfile, indent=2, ensure_ascii=False)
                    
                    # Créer une copie locale
                    try:
                        import shutil
                        shutil.copy2(filename, local_filename)
                        output.insert("end", f"✅ Export terminé: {filename}\n")
                        output.insert("end", f"📁 Copie locale: {local_filename}\n")
                        output.insert("end", f"📊 {len(users)} utilisateur(s) exporté(s)\n")
                    except Exception as e:
                        output.insert("end", f"✅ Export terminé: {filename}\n")
                        output.insert("end", f"⚠️ Impossible de créer la copie locale: {e}\n")
                        output.insert("end", f"📊 {len(users)} utilisateur(s) exporté(s)\n")
                    
                    # Ouvrir le fichier automatiquement
                    try:
                        import os
                        import subprocess
                        import platform
                        
                        # Obtenir le chemin absolu du fichier
                        abs_path = os.path.abspath(filename)
                        output.insert("end", f"📁 Fichier sauvegardé: {abs_path}\n")
                        
                        # Ouvrir le fichier selon le système d'exploitation
                        system = platform.system()
                        if system == "Windows":
                            os.startfile(abs_path)
                        elif system == "Darwin":  # macOS
                            subprocess.run(["open", abs_path])
                        else:  # Linux
                            subprocess.run(["xdg-open", abs_path])
                        
                        output.insert("end", f"🚀 Fichier ouvert automatiquement!\n")
                        
                    except Exception as e:
                        output.insert("end", f"⚠️ Impossible d'ouvrir le fichier automatiquement: {e}\n")
                    
                    # Log de l'activité
                    self._log_activity("Export", f"Export utilisateurs via CLI", "CLI", f"Format: {format_export}, Fichier: {filename}")
                    
                elif table == "products":
                    output.insert("end", f"📤 Export des produits au format {format_export.upper()}...\n")
                    
                    cursor.execute("SELECT nom, description, marque, modele, fournisseur, stock FROM sge_cre.produits ORDER BY nom")
                    products = cursor.fetchall()
                    
                    if format_export == "csv":
                        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                            writer = csv.writer(csvfile, delimiter=';')
                            # En-têtes avec formatage
                            writer.writerow(['NOM', 'DESCRIPTION', 'MARQUE', 'MODÈLE', 'FOURNISSEUR', 'STOCK'])
                            # Séparateur visuel
                            writer.writerow(['---', '---', '---', '---', '---', '---'])
                            # Données
                            for product in products:
                                writer.writerow(product)
                    elif format_export == "xlsx":
                        try:
                            import pandas as pd
                            from openpyxl import Workbook
                            from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
                            
                            # Créer un DataFrame
                            df = pd.DataFrame(products, columns=['NOM', 'DESCRIPTION', 'MARQUE', 'MODÈLE', 'FOURNISSEUR', 'STOCK'])
                            
                            # Exporter vers Excel avec formatage
                            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                                df.to_excel(writer, sheet_name='Produits', index=False)
                                
                                # Récupérer le workbook et la worksheet
                                workbook = writer.book
                                worksheet = writer.sheets['Produits']
                                
                                # Style pour les en-têtes
                                header_font = Font(bold=True, color="FFFFFF", size=12)
                                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                                header_alignment = Alignment(horizontal="center", vertical="center")
                                
                                # Appliquer le style aux en-têtes
                                for col in range(1, len(df.columns) + 1):
                                    cell = worksheet.cell(row=1, column=col)
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = header_alignment
                                
                                # Style alterné pour les lignes de données
                                for row in range(2, len(df) + 2):
                                    for col in range(1, len(df.columns) + 1):
                                        cell = worksheet.cell(row=row, column=col)
                                        if row % 2 == 0:
                                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                                
                                # Ajuster la largeur des colonnes
                                for column in worksheet.columns:
                                    max_length = 0
                                    column_letter = column[0].column_letter
                                    for cell in column:
                                        try:
                                            if len(str(cell.value)) > max_length:
                                                max_length = len(str(cell.value))
                                        except:
                                            pass
                                    adjusted_width = min(max_length + 2, 50)
                                    worksheet.column_dimensions[column_letter].width = adjusted_width
                                    
                        except ImportError:
                            output.insert("end", "❌ Erreur: pandas et openpyxl requis pour l'export Excel\n")
                            return
                            
                    elif format_export == "json":
                        data = []
                        for product in products:
                            data.append({
                                'nom': product[0],
                                'description': product[1],
                                'marque': product[2],
                                'modele': product[3],
                                'fournisseur': product[4],
                                'stock': product[5]
                            })
                        with open(filename, 'w', encoding='utf-8') as jsonfile:
                            json.dump(data, jsonfile, indent=2, ensure_ascii=False)
                    
                    # Créer une copie locale
                    try:
                        import shutil
                        shutil.copy2(filename, local_filename)
                        output.insert("end", f"✅ Export terminé: {filename}\n")
                        output.insert("end", f"📁 Copie locale: {local_filename}\n")
                        output.insert("end", f"📊 {len(products)} produit(s) exporté(s)\n")
                    except Exception as e:
                        output.insert("end", f"✅ Export terminé: {filename}\n")
                        output.insert("end", f"⚠️ Impossible de créer la copie locale: {e}\n")
                        output.insert("end", f"📊 {len(products)} produit(s) exporté(s)\n")
                    
                    # Ouvrir le fichier automatiquement
                    try:
                        import os
                        import subprocess
                        import platform
                        
                        # Obtenir le chemin absolu du fichier
                        abs_path = os.path.abspath(filename)
                        output.insert("end", f"📁 Fichier sauvegardé: {abs_path}\n")
                        
                        # Ouvrir le fichier selon le système d'exploitation
                        system = platform.system()
                        if system == "Windows":
                            os.startfile(abs_path)
                        elif system == "Darwin":  # macOS
                            subprocess.run(["open", abs_path])
                        else:  # Linux
                            subprocess.run(["xdg-open", abs_path])
                        
                        output.insert("end", f"🚀 Fichier ouvert automatiquement!\n")
                        
                    except Exception as e:
                        output.insert("end", f"⚠️ Impossible d'ouvrir le fichier automatiquement: {e}\n")
                    
                    # Log de l'activité
                    self._log_activity("Export", f"Export produits via CLI", "CLI", f"Format: {format_export}, Fichier: {filename}")
                    
                elif table == "movements":
                    output.insert("end", f"📤 Export des mouvements au format {format_export.upper()}...\n")
                    
                    cursor.execute("SELECT type, quantite, date_mouvement, produit_nom, reference FROM sge_cre.mouvements ORDER BY date_mouvement DESC")
                    movements = cursor.fetchall()
                    
                    if format_export == "csv":
                        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                            writer = csv.writer(csvfile, delimiter=';')
                            # En-têtes avec formatage
                            writer.writerow(['TYPE', 'QUANTITÉ', 'DATE', 'PRODUIT', 'RÉFÉRENCE'])
                            # Séparateur visuel
                            writer.writerow(['---', '---', '---', '---', '---'])
                            # Données
                            for movement in movements:
                                writer.writerow(movement)
                    elif format_export == "xlsx":
                        try:
                            import pandas as pd
                            from openpyxl import Workbook
                            from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
                            
                            # Créer un DataFrame
                            df = pd.DataFrame(movements, columns=['TYPE', 'QUANTITÉ', 'DATE', 'PRODUIT', 'RÉFÉRENCE'])
                            
                            # Exporter vers Excel avec formatage
                            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                                df.to_excel(writer, sheet_name='Mouvements', index=False)
                                
                                # Récupérer le workbook et la worksheet
                                workbook = writer.book
                                worksheet = writer.sheets['Mouvements']
                                
                                # Style pour les en-têtes
                                header_font = Font(bold=True, color="FFFFFF", size=12)
                                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                                header_alignment = Alignment(horizontal="center", vertical="center")
                                
                                # Appliquer le style aux en-têtes
                                for col in range(1, len(df.columns) + 1):
                                    cell = worksheet.cell(row=1, column=col)
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = header_alignment
                                
                                # Style alterné pour les lignes de données
                                for row in range(2, len(df) + 2):
                                    for col in range(1, len(df.columns) + 1):
                                        cell = worksheet.cell(row=row, column=col)
                                        if row % 2 == 0:
                                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                                
                                # Ajuster la largeur des colonnes
                                for column in worksheet.columns:
                                    max_length = 0
                                    column_letter = column[0].column_letter
                                    for cell in column:
                                        try:
                                            if len(str(cell.value)) > max_length:
                                                max_length = len(str(cell.value))
                                        except:
                                            pass
                                    adjusted_width = min(max_length + 2, 50)
                                    worksheet.column_dimensions[column_letter].width = adjusted_width
                                    
                        except ImportError:
                            output.insert("end", "❌ Erreur: pandas et openpyxl requis pour l'export Excel\n")
                            return
                            
                    elif format_export == "json":
                        data = []
                        for movement in movements:
                            data.append({
                                'type': movement[0],
                                'quantite': movement[1],
                                'date': movement[2].isoformat() if movement[2] else None,
                                'produit': movement[3],
                                'reference': movement[4]
                            })
                        with open(filename, 'w', encoding='utf-8') as jsonfile:
                            json.dump(data, jsonfile, indent=2, ensure_ascii=False)
                        
                        
                        # Créer une copie locale
                        try:
                            import shutil
                            shutil.copy2(filename, local_filename)
                            output.insert("end", f"✅ Export terminé: {filename}\n")
                            output.insert("end", f"📁 Copie locale: {local_filename}\n")
                            output.insert("end", f"📊 {len(users)} utilisateur(s) exporté(s)\n")
                        except Exception as e:
                            output.insert("end", f"✅ Export terminé: {filename}\n")
                            output.insert("end", f"⚠️ Impossible de créer la copie locale: {e}\n")
                            output.insert("end", f"📊 {len(users)} utilisateur(s) exporté(s)\n")
                        
                        # Ouvrir le fichier automatiquement
                        try:
                            import os
                            import subprocess
                            import platform
                            
                            # Obtenir le chemin absolu du fichier
                            abs_path = os.path.abspath(filename)
                            output.insert("end", f"📁 Fichier sauvegardé: {abs_path}\n")
                            
                            # Ouvrir le fichier selon le système d'exploitation
                            system = platform.system()
                            if system == "Windows":
                                os.startfile(abs_path)
                            elif system == "Darwin":  # macOS
                                subprocess.run(["open", abs_path])
                            else:  # Linux
                                subprocess.run(["xdg-open", abs_path])
                            
                            output.insert("end", f"🚀 Fichier ouvert automatiquement!\n")
                            
                        except Exception as e:
                            output.insert("end", f"⚠️ Impossible d'ouvrir le fichier automatiquement: {e}\n")

                        
                        # Ouvrir le fichier automatiquement
                        try:
                            import os
                            import subprocess
                            import platform
                            
                            # Obtenir le chemin absolu du fichier
                            abs_path = os.path.abspath(filename)
                            output.insert("end", f"📁 Fichier sauvegardé: {abs_path}\n")
                            
                            # Ouvrir le fichier selon le système d'exploitation
                            system = platform.system()
                            if system == "Windows":
                                os.startfile(abs_path)
                            elif system == "Darwin":  # macOS
                                subprocess.run(["open", abs_path])
                            else:  # Linux
                                subprocess.run(["xdg-open", abs_path])
                            
                            output.insert("end", f"🚀 Fichier ouvert automatiquement!\n")
                            
                        except Exception as e:
                            output.insert("end", f"⚠️ Impossible d'ouvrir le fichier automatiquement: {e}\n")
                        
                        # Log de l'activité
                        self._log_activity("Export", f"Export mouvements via CLI", "CLI", f"Format: {format_export}, Fichier: {filename}")
                        
                    elif table == "packaging":
                        output.insert("end", f"📤 Export des emballages au format {format_export.upper()}...\n")
                        
                        cursor.execute("SELECT type_emballage, etat_emballage FROM sge_cre.materiel_emballage ORDER BY type_emballage")
                        packaging = cursor.fetchall()
                        
                        if format_export == "csv":
                            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                                writer = csv.writer(csvfile, delimiter=';')
                                # En-têtes avec formatage
                                writer.writerow(['TYPE', 'ÉTAT'])
                                # Séparateur visuel
                                writer.writerow(['---', '---'])
                                # Données
                                for pack in packaging:
                                    writer.writerow(pack)
                        elif format_export == "xlsx":
                            try:
                                import pandas as pd
                                from openpyxl import Workbook
                                from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
                                
                                # Créer un DataFrame
                                df = pd.DataFrame(packaging, columns=['TYPE', 'ÉTAT'])
                                
                                # Exporter vers Excel avec formatage
                                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                                    df.to_excel(writer, sheet_name='Emballages', index=False)
                                    
                                    # Récupérer le workbook et la worksheet
                                    workbook = writer.book
                                    worksheet = writer.sheets['Emballages']
                                    
                                    # Style pour les en-têtes
                                    header_font = Font(bold=True, color="FFFFFF", size=12)
                                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                                    header_alignment = Alignment(horizontal="center", vertical="center")
                                    
                                    # Appliquer le style aux en-têtes
                                    for col in range(1, len(df.columns) + 1):
                                        cell = worksheet.cell(row=1, column=col)
                                        cell.font = header_font
                                        cell.fill = header_fill
                                        cell.alignment = header_alignment
                                    
                                    # Style alterné pour les lignes de données
                                    for row in range(2, len(df) + 2):
                                        for col in range(1, len(df.columns) + 1):
                                            cell = worksheet.cell(row=row, column=col)
                                            if row % 2 == 0:
                                                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                                    
                                    # Ajuster la largeur des colonnes
                                    for column in worksheet.columns:
                                        max_length = 0
                                        column_letter = column[0].column_letter
                                        for cell in column:
                                            try:
                                                if len(str(cell.value)) > max_length:
                                                    max_length = len(str(cell.value))
                                            except:
                                                pass
                                    adjusted_width = min(max_length + 2, 50)
                                    worksheet.column_dimensions[column_letter].width = adjusted_width
                                    
                            except ImportError:
                                output.insert("end", "❌ Erreur: pandas et openpyxl requis pour l'export Excel\n")
                                return
                                
                        elif format_export == "json":
                            data = []
                            for pkg in packaging:
                                data.append({
                                    'type': pkg[0],
                                    'etat': pkg[1]
                                })
                            with open(filename, 'w', encoding='utf-8') as jsonfile:
                                json.dump(data, jsonfile, indent=2, ensure_ascii=False)
                            
                            
                            # Créer une copie locale
                            try:
                                import shutil
                                shutil.copy2(filename, local_filename)
                                output.insert("end", f"✅ Export terminé: {filename}\n")
                                output.insert("end", f"📁 Copie locale: {local_filename}\n")
                                output.insert("end", f"📊 {len(users)} utilisateur(s) exporté(s)\n")
                            except Exception as e:
                                output.insert("end", f"✅ Export terminé: {filename}\n")
                                output.insert("end", f"⚠️ Impossible de créer la copie locale: {e}\n")
                                output.insert("end", f"📊 {len(users)} utilisateur(s) exporté(s)\n")
                            
                            # Ouvrir le fichier automatiquement
                            try:
                                import os
                                import subprocess
                                import platform
                                
                                # Obtenir le chemin absolu du fichier
                                abs_path = os.path.abspath(filename)
                                output.insert("end", f"📁 Fichier sauvegardé: {abs_path}\n")
                                
                                # Ouvrir le fichier selon le système d'exploitation
                                system = platform.system()
                                if system == "Windows":
                                    os.startfile(abs_path)
                                elif system == "Darwin":  # macOS
                                    subprocess.run(["open", abs_path])
                                else:  # Linux
                                    subprocess.run(["xdg-open", abs_path])
                                
                                output.insert("end", f"🚀 Fichier ouvert automatiquement!\n")
                                
                            except Exception as e:
                                output.insert("end", f"⚠️ Impossible d'ouvrir le fichier automatiquement: {e}\n")

                            
                            # Ouvrir le fichier automatiquement
                            try:
                                import os
                                import subprocess
                                import platform
                                
                                # Obtenir le chemin absolu du fichier
                                abs_path = os.path.abspath(filename)
                                output.insert("end", f"📁 Fichier sauvegardé: {abs_path}\n")
                                
                                # Ouvrir le fichier selon le système d'exploitation
                                system = platform.system()
                                if system == "Windows":
                                    os.startfile(abs_path)
                                elif system == "Darwin":  # macOS
                                    subprocess.run(["open", abs_path])
                                else:  # Linux
                                    subprocess.run(["xdg-open", abs_path])
                                
                                output.insert("end", f"🚀 Fichier ouvert automatiquement!\n")
                                
                            except Exception as e:
                                output.insert("end", f"⚠️ Impossible d'ouvrir le fichier automatiquement: {e}\n")
                            
                            # Log de l'activité
                            self._log_activity("Export", f"Export emballages via CLI", "CLI", f"Format: {format_export}, Fichier: {filename}")
                            
                        else:
                            output.insert("end", f"❌ Table '{table}' non reconnue.\n")
                        
                        conn.close()
                            


            except Exception as e:
                output.insert("end", f"❌ Erreur: {str(e)}\n")

            except Exception as e:
                output.insert("end", f"❌ Erreur: {str(e)}\n")

        except Exception as e:
            output.insert("end", f"❌ Erreur: {str(e)}\n")

        except Exception as e:
            output.insert("end", f"❌ Erreur: {str(e)}\n")

    def _cli_download(self, args):
        """Téléchargement de documents via CLI"""
        output = self.cli_output
        
        if not args:
            output.insert("end", "❌ Commande incomplète. Tapez 'help download' pour l'aide.\n")
            return
        
        doc_type = args[0].lower()
        
        if len(args) < 2:
            output.insert("end", "❌ ID du document manquant.\n")
            return
        
        doc_id = args[1]
        
        try:
            import psycopg2
            import datetime
            import os
            
            conn = psycopg2.connect(**PG_CONN)
            cursor = conn.cursor()
            
            # Créer le dossier downloads s'il n'existe pas
            if not os.path.exists("downloads"):
                os.makedirs("downloads")
            
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if doc_type == "expedition":
                output.insert("end", f"📥 Téléchargement du bon d'expédition #{doc_id}...\n")
                
                # Vérifier si l'expédition existe
                cursor.execute("""
                    SELECT e.id, e.date_expedition, e.destinataire, e.adresse_livraison, 
                           p.nom as produit_nom, e.quantite, e.statut
                    FROM sge_cre.expeditions e
                    LEFT JOIN sge_cre.produits p ON e.produit_id = p.id
                    WHERE e.id = %s
                """, (doc_id,))
                
                expedition = cursor.fetchone()
                
                if expedition:
                    filename = f"downloads/bon_expedition_{doc_id}_{timestamp}.txt"
                    
                    with open(filename, 'w', encoding='utf-8') as f:
                        f.write("=" * 50 + "\n")
                        f.write("BON D'EXPÉDITION\n")
                        f.write("=" * 50 + "\n\n")
                        f.write(f"Numéro: {expedition[0]}\n")
                        f.write(f"Date: {expedition[1]}\n")
                        f.write(f"Destinataire: {expedition[2]}\n")
                        f.write(f"Adresse: {expedition[3]}\n")
                        f.write(f"Produit: {expedition[4]}\n")
                        f.write(f"Quantité: {expedition[5]}\n")
                        f.write(f"Statut: {expedition[6]}\n")
                        f.write("\n" + "=" * 50 + "\n")
                    
                    output.insert("end", f"✅ Bon d'expédition #{doc_id} téléchargé avec succès!\n")
                    output.insert("end", f"📁 Fichier: {filename}\n")
                    
                    # Log de l'activité
                    self._log_activity("Téléchargement", f"Téléchargement bon expédition #{doc_id} via CLI", "CLI", f"Fichier: {filename}")
                else:
                    output.insert("end", f"❌ Expédition #{doc_id} non trouvée.\n")
                
            elif doc_type == "reception":
                output.insert("end", f"📥 Téléchargement du bon de réception #{doc_id}...\n")
                
                # Vérifier si la réception existe
                cursor.execute("""
                    SELECT r.id, r.date_reception, r.fournisseur, r.adresse_reception,
                           p.nom as produit_nom, r.quantite, r.statut
                    FROM sge_cre.receptions r
                    LEFT JOIN sge_cre.produits p ON r.produit_id = p.id
                    WHERE r.id = %s
                """, (doc_id,))
                
                reception = cursor.fetchone()
                
                if reception:
                    filename = f"downloads/bon_reception_{doc_id}_{timestamp}.txt"
                    
                    with open(filename, 'w', encoding='utf-8') as f:
                        f.write("=" * 50 + "\n")
                        f.write("BON DE RÉCEPTION\n")
                        f.write("=" * 50 + "\n\n")
                        f.write(f"Numéro: {reception[0]}\n")
                        f.write(f"Date: {reception[1]}\n")
                        f.write(f"Fournisseur: {reception[2]}\n")
                        f.write(f"Adresse: {reception[3]}\n")
                        f.write(f"Produit: {reception[4]}\n")
                        f.write(f"Quantité: {reception[5]}\n")
                        f.write(f"Statut: {reception[6]}\n")
                        f.write("\n" + "=" * 50 + "\n")
                    
                    output.insert("end", f"✅ Bon de réception #{doc_id} téléchargé avec succès!\n")
                    output.insert("end", f"📁 Fichier: {filename}\n")
                    
                    # Log de l'activité
                    self._log_activity("Téléchargement", f"Téléchargement bon réception #{doc_id} via CLI", "CLI", f"Fichier: {filename}")
                else:
                    output.insert("end", f"❌ Réception #{doc_id} non trouvée.\n")
                
            else:
                output.insert("end", f"❌ Type de document '{doc_type}' non reconnu.\n")
            
            conn.close()
                
        except Exception as e:
            output.insert("end", f"❌ Erreur: {str(e)}\n")

    def _cli_status(self):
        """Affiche le statut du système"""
        output = self.cli_output
        
        try:
            import psycopg2
            conn = psycopg2.connect(**PG_CONN)
            cursor = conn.cursor()
            
            # Statistiques des utilisateurs (corrigé : on compte tous les utilisateurs)
            cursor.execute("SELECT COUNT(*) FROM sge_cre.individus")
            users_actifs = cursor.fetchone()[0]
            
            # Statistiques des produits
            cursor.execute("SELECT COUNT(*) FROM sge_cre.produits")
            total_products = cursor.fetchone()[0]
            
            # Statistiques des mouvements
            cursor.execute("SELECT COUNT(*) FROM sge_cre.mouvements_stock")
            total_movements = cursor.fetchone()[0]
            
            conn.close()
            
            output.insert("end", "\n🔧 STATUT DU SYSTÈME:\n")
            output.insert("end", "═" * 50 + "\n")
            output.insert("end", f"👥 Utilisateurs (total): {users_actifs}\n")
            output.insert("end", f"📦 Produits en base: {total_products}\n")
            output.insert("end", f"📊 Mouvements enregistrés: {total_movements}\n")
            output.insert("end", f"🟢 Système: Opérationnel\n")
            import datetime
            output.insert("end", f"📅 Dernière vérification: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
            
        except Exception as e:
            output.insert("end", f"❌ Erreur lors de la vérification du statut: {str(e)}\n")

    def _open_user_modal(self):
        import threading
        modal = ctk.CTkToplevel(self)
        modal.title("Ajouter un Utilisateur")
        modal.geometry("500x700")  # Augmenté la hauteur pour accommoder les boutons
        modal.grab_set()
        modal.resizable(False, False)
        modal.configure(fg_color="#f0f6ff")  # Fond doux bleu clair

        # Barre de titre colorée avec icône
        title_bar = ctk.CTkFrame(modal, fg_color="#2563eb", corner_radius=18)
        title_bar.pack(fill="x", pady=(0, 0))
        ctk.CTkLabel(
            title_bar, text="👤 Ajouter un Utilisateur",
            font=ctk.CTkFont(size=23, weight="bold"),
            text_color="#fff"
        ).pack(pady=18)

        # Illustration ou icône décorative
        ctk.CTkLabel(
            modal, text="📝", font=ctk.CTkFont(size=48), text_color="#2563eb"
        ).pack(pady=(10, 0))

        scroll = ctk.CTkScrollableFrame(modal, fg_color="#f7fafd", corner_radius=22, width=460, height=420)  # Réduit la hauteur
        scroll.pack(fill="both", expand=True, padx=20, pady=8)
        entries = {}
        errors = {}
        fields = [
            ("Nom *", "user", ""),
            ("Prénom *", "user", ""),
            ("Email *", "mail", ""),
            ("Matricule *", "id", ""),
            ("Rôle *", "role", "combo"),
            ("Mot de passe *", "lock", "password"),
            ("Confirmation *", "lock", "password"),
            ("Adresse *", "home", ""),
            ("Téléphone", "phone", "")
        ]
        role_values = [
            "Le responsable des stocks",
            "Le magasinier",
            "Les emballeurs",
            "Le responsable de la logistique",
            "Les agents de logistique",
            "Les livreurs",
            "Le responsable informatique",
            "Les techniciens informatiques",
            "Le responsable de la sécurité physique",
            "Les gardes de sécurité"
        ]
        for label, icon, val in fields:
            row = ctk.CTkFrame(scroll, fg_color="transparent")
            row.pack(fill="x", pady=(14,0))
            # Astérisque rouge pour les champs obligatoires
            if "*" in label:
                label_widget = ctk.CTkLabel(row, text=label, font=ctk.CTkFont(size=16, weight="bold"), text_color="#374151")
                label_widget.pack(side="left", padx=(0,2))
                ctk.CTkLabel(row, text="*", font=ctk.CTkFont(size=16, weight="bold"), text_color="#ef4444").pack(side="left")
            else:
                ctk.CTkLabel(row, text=label, font=ctk.CTkFont(size=16, weight="bold"), text_color="#374151").pack(side="left", padx=(0,10))
            if val == "combo":
                e = ctk.CTkComboBox(row, values=role_values, width=270, font=ctk.CTkFont(size=15))
                e.set(role_values[0])
            elif val == "password":
                e = ctk.CTkEntry(row, height=40, font=ctk.CTkFont(size=15), show="*")
            else:
                e = ctk.CTkEntry(row, height=40, font=ctk.CTkFont(size=15))
            e.pack(side="right", fill="x", expand=True)
            entries[label] = e
            err = ctk.CTkLabel(scroll, text="", text_color="#ef4444", font=ctk.CTkFont(size=12))
            err.pack(anchor="w", padx=10)
            errors[label] = err
        # Génération automatique du matricule avec le nouveau module
        from matricule_manager import MatriculeManager
        def on_role_change(event=None):
            role = entries["Rôle *"].get()
            if role:
                auto_matricule = MatriculeManager.generate_matricule(role)
            if not getattr(entries["Matricule *"], '_user_modified', False):
                entries["Matricule *"].delete(0, 'end')
                entries["Matricule *"].insert(0, auto_matricule)
                prefix = MatriculeManager.get_role_prefix(role)
                if hasattr(entries["Matricule *"], '_prefix_label'):
                    entries["Matricule *"]._prefix_label.configure(text=f"Préfixe: {prefix}")
                else:
                    prefix_label = ctk.CTkLabel(scroll, text=f"Préfixe: {prefix}", font=ctk.CTkFont(size=12), text_color="#6b7280")
                    prefix_label.pack(anchor="w", padx=10, pady=(0, 5))
                    entries["Matricule *"]._prefix_label = prefix_label
        entries["Rôle *"].bind("<<ComboboxSelected>>", on_role_change)
        def on_matricule_edit(event=None):
            entries["Matricule *"]._user_modified = True
        entries["Matricule *"].bind('<Key>', on_matricule_edit)
        on_role_change()
        # Gestion dynamique des erreurs : efface l'erreur dès que l'utilisateur modifie le champ
        def clear_error(event, label):
            errors[label].configure(text="")
            entries[label].configure(border_color="#D1D5DB")
        for label in entries:
            entries[label].bind('<Key>', lambda e, l=label: clear_error(e, l))
            if label == "Rôle *":
                entries[label].bind("<<ComboboxSelected>>", lambda e, l=label: clear_error(e, l))
        def save():
            # --- ÉDITION UTILISATEUR ---
            if 'uid' in locals() or 'uid' in globals():
                nom = entries["Nom *"].get()
                prenom = entries["Prénom *"].get()
                email = entries["Email *"].get()
                role = entries["Rôle *"].get()
                matricule = entries["Matricule *"].get()
                valid = True
                for err in errors.values(): err.configure(text="")
                for field, entry in entries.items(): entry.configure(border_color="#D1D5DB")
                # Validation des champs obligatoires
                if not nom:
                    errors["Nom *"].configure(text="Champ obligatoire")
                    entries["Nom *"].configure(border_color="#ef4444")
                    valid = False
                if not prenom:
                    errors["Prénom *"].configure(text="Champ obligatoire")
                    entries["Prénom *"].configure(border_color="#ef4444")
                    valid = False
                if not email or "@" not in email:
                    errors["Email *"].configure(text="Email invalide")
                    entries["Email *"].configure(border_color="#ef4444")
                    valid = False
                if not matricule:
                    errors["Matricule *"].configure(text="Champ obligatoire")
                    entries["Matricule *"].configure(border_color="#ef4444")
                    valid = False
                if not role:
                    errors["Rôle *"].configure(text="Champ obligatoire")
                    entries["Rôle *"].configure(border_color="#ef4444")
                    valid = False
                if not valid:
                    return
                try:
                    from database import update_user
                    update_user(uid, nom, prenom, email, role, "", "", matricule)
                    modal.destroy()
                    self._refresh_users(table, search_term)
                except Exception as e:
                    errors["Email *"].configure(text=f"Erreur : {e}")
                    entries["Email *"].configure(border_color="#ef4444")
            # --- CRÉATION UTILISATEUR ---
            else:
                nom = entries["Nom *"].get()
                prenom = entries["Prénom *"].get()
                email = entries["Email *"].get()
                matricule = entries["Matricule *"].get()
                role = entries["Rôle *"].get()
                mdp = entries["Mot de passe *"].get()
                conf = entries["Confirmation *"].get()
                adresse = entries["Adresse *"].get()
                telephone = entries["Téléphone"].get()
                valid = True
                for err in errors.values(): err.configure(text="")
                for field, entry in entries.items(): entry.configure(border_color="#D1D5DB")
                # Validation des champs obligatoires
                if not nom:
                    errors["Nom *"].configure(text="Champ obligatoire")
                    entries["Nom *"].configure(border_color="#ef4444")
                    valid = False
                if not prenom:
                    errors["Prénom *"].configure(text="Champ obligatoire")
                    entries["Prénom *"].configure(border_color="#ef4444")
                    valid = False
                if not email or "@" not in email:
                    errors["Email *"].configure(text="Email invalide")
                    entries["Email *"].configure(border_color="#ef4444")
                    valid = False
                if not matricule:
                    errors["Matricule *"].configure(text="Champ obligatoire")
                    entries["Matricule *"].configure(border_color="#ef4444")
                    valid = False
                if not mdp or not conf:
                    errors["Mot de passe *"].configure(text="Champ obligatoire")
                    errors["Confirmation *"].configure(text="Champ obligatoire")
                    entries["Mot de passe *"].configure(border_color="#ef4444")
                    entries["Confirmation *"].configure(border_color="#ef4444")
                    valid = False
                if mdp != conf:
                    errors["Confirmation *"].configure(text="Les mots de passe ne correspondent pas")
                    entries["Confirmation *"].configure(border_color="#ef4444")
                    valid = False
                if not role:
                    errors["Rôle *"].configure(text="Champ obligatoire")
                    entries["Rôle *"].configure(border_color="#ef4444")
                    valid = False
                if not adresse:
                    errors["Adresse *"].configure(text="Champ obligatoire")
                    entries["Adresse *"].configure(border_color="#ef4444")
                    valid = False
                # Validation et unicité du matricule
                if matricule:
                    is_valid, validation_msg = MatriculeManager.validate_matricule(matricule)
                    if not is_valid:
                        errors["Matricule *"].configure(text=validation_msg)
                        entries["Matricule *"].configure(border_color="#ef4444")
                        valid = False
                    else:
                        is_available, availability_msg = MatriculeManager.is_matricule_available(matricule)
                        if not is_available:
                            errors["Matricule *"].configure(text=availability_msg)
                            entries["Matricule *"].configure(border_color="#ef4444")
                            valid = False
                        expected_role = MatriculeManager.get_role_from_matricule(matricule)
                        if expected_role and expected_role != role:
                            errors["Matricule *"].configure(text=f"Ce matricule correspond au rôle '{expected_role}'")
                            entries["Matricule *"].configure(border_color="#ef4444")
                            valid = False
                if not valid:
                    return
                try:
                    import psycopg2, datetime
                    conn = psycopg2.connect(**PG_CONN)
                    cursor = conn.cursor()
                    cursor.execute("""
                        INSERT INTO sge_cre.individus (nom, prenom, email, password, role, matricule, adresse, telephone)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        nom,
                        prenom,
                        email,
                        mdp,  # mot de passe en clair (à hasher en production)
                        role,
                        matricule,
                        adresse,
                        telephone
                    ))
                    conn.commit()
                    conn.close()
                    # Popup de succès
                    success_modal = ctk.CTkToplevel(modal)
                    success_modal.title("🎉 Succès")
                    success_modal.geometry("480x420")
                    success_modal.grab_set()
                    success_modal.resizable(False, False)
                    success_modal.configure(fg_color="#ffffff")
                    # Centrer la modale
                    success_modal.update_idletasks()
                    x = (success_modal.winfo_screenwidth() // 2) - (480 // 2)
                    y = (success_modal.winfo_screenheight() // 2) - (420 // 2)
                    success_modal.geometry(f"480x420+{x}+{y}")
                    # Conteneur principal
                    main_container = ctk.CTkFrame(success_modal, fg_color="#ffffff", corner_radius=24, border_width=3, border_color="#e5e7eb")
                    main_container.pack(fill="both", expand=True, padx=25, pady=25)
                    # Animation de succès
                    success_circle = ctk.CTkFrame(main_container, fg_color="#dcfce7", corner_radius=60, width=120, height=120, border_width=3, border_color="#bbf7d0")
                    success_circle.pack(pady=(35, 25))
                    ctk.CTkLabel(success_circle, text="🎉", font=ctk.CTkFont(size=52), text_color="#16a34a").pack(expand=True)
                    ctk.CTkLabel(main_container, text="🎊 Utilisateur Créé avec Succès !", font=ctk.CTkFont(size=26, weight="bold"), text_color="#1f2937").pack(pady=(0, 18))
                    ctk.CTkLabel(main_container, text=f"✨ L'utilisateur {prenom} {nom} a été ajouté avec succès au système de gestion d'entrepôts.", font=ctk.CTkFont(size=15), text_color="#6b7280", wraplength=400).pack(pady=(0, 28))
                    info_frame = ctk.CTkFrame(main_container, fg_color="#f8fafc", corner_radius=16, border_width=2, border_color="#e2e8f0")
                    info_frame.pack(fill="x", padx=25, pady=(0, 28))
                    matricule_container = ctk.CTkFrame(info_frame, fg_color="transparent")
                    matricule_container.pack(fill="x", padx=18, pady=(15, 8))
                    ctk.CTkLabel(matricule_container, text="🆔 Matricule", font=ctk.CTkFont(size=15, weight="bold"), text_color="#374151").pack(side="left")
                    ctk.CTkLabel(matricule_container, text=matricule, font=ctk.CTkFont(size=20, weight="bold"), text_color="#059669").pack(side="right")
                    role_container = ctk.CTkFrame(info_frame, fg_color="transparent")
                    role_container.pack(fill="x", padx=18, pady=(8, 15))
                    ctk.CTkLabel(role_container, text="👤 Rôle", font=ctk.CTkFont(size=15, weight="bold"), text_color="#374151").pack(side="left")
                    ctk.CTkLabel(role_container, text=role, font=ctk.CTkFont(size=17), text_color="#6b7280").pack(side="right")
                    ctk.CTkButton(main_container, text="🎯 Parfait ! Continuer", fg_color="#10b981", hover_color="#059669", text_color="white", corner_radius=16, height=52, font=ctk.CTkFont(size=17, weight="bold"), border_width=2, border_color="#34d399", command=lambda: [success_modal.destroy(), modal.destroy()]).pack(pady=(0, 25), padx=25, fill="x")
                    success_modal.after(5000, lambda: [success_modal.destroy(), modal.destroy()])
                except Exception as e:
                    errors["Email *"].configure(text=f"Erreur : {e}")
                    entries["Email *"].configure(border_color="#ef4444")
                
                if hasattr(self, 'tab_contents') and "Utilisateurs" in self.tab_contents:
                    for w in self.tab_contents["Utilisateurs"].winfo_children():
                        w.destroy()
                    self._build_users_tab(self.tab_contents["Utilisateurs"])
                    self._update_stats_immediately()
                    self._log_activity("Création", f"Création de l'utilisateur {prenom} {nom}", f"{prenom} {nom}", f"Matricule: {matricule}, Rôle: {role}")
            

        # Zone des boutons en bas de la modale - Plus visible et moderne
        buttons_frame = ctk.CTkFrame(modal, fg_color="#ffffff", corner_radius=15, border_width=1, border_color="#e5e7eb")
        buttons_frame.pack(fill="x", pady=(15, 20), padx=20)
        
        # Titre de la section boutons avec icône
        title_container = ctk.CTkFrame(buttons_frame, fg_color="transparent")
        title_container.pack(fill="x", padx=20, pady=(15, 10))
        
        ctk.CTkLabel(
            title_container, text="⚡", font=ctk.CTkFont(size=16), text_color="#f59e0b"
        ).pack(side="left", padx=(0, 8))
        
        ctk.CTkLabel(
            title_container, text="Ajouter ou Annuler", font=ctk.CTkFont(size=16, weight="bold"), text_color="#1f2937"
        ).pack(side="left")
        
        # Conteneur pour les boutons
        button_container = ctk.CTkFrame(buttons_frame, fg_color="transparent")
        button_container.pack(fill="x", padx=20, pady=(0, 15))
        
        # Bouton Annuler (à gauche) - Plus visible
        ctk.CTkButton(
            button_container, text="❌ Annuler", fg_color="#fef2f2", hover_color="#fee2e2", 
            text_color="#dc2626", corner_radius=12, height=50, font=ctk.CTkFont(size=16, weight="bold"),
            border_width=2, border_color="#fecaca", command=modal.destroy
        ).pack(side="left", padx=(0, 10), expand=True, fill="x")
        
        # Bouton Valider (à droite) - Plus visible
        ctk.CTkButton(
            button_container, text="✔ Valider & Créer", fg_color="#10b981", hover_color="#059669", 
            text_color="white", corner_radius=12, height=50, font=ctk.CTkFont(size=16, weight="bold"),
            command=save
        ).pack(side="right", padx=(10, 0), expand=True, fill="x")